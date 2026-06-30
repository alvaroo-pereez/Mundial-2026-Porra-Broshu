"""
Audita pronósticos R32 (73-88) comparando Excel fuente vs maestro.
Prioridad: Pepe y Patri; extensible al resto de importados.
"""
from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from config.groups import load_group
from import_r32_predictions import (
    IMPORTS,
    PT_FIRST_ROW,
    R32_FIRST_ID,
    R32_LAST_ID,
    build_pair_index,
    load_master_r32,
    parse_source,
)
from scoring import calc_match_points, classify_match_tier

ROOT = Path(__file__).parent
REPORT_PATH = ROOT / "output" / "audit_predictions_report.txt"
FINISHED_R32 = {73, 74, 75, 76}
FOCUS_PLAYERS = {"Pepe", "Patri"}


@dataclass
class AuditFlag:
    kind: str
    player: str
    mid: int | None
    detail: str


@dataclass
class PlayerAudit:
    label: str
    player: str
    flags: list[AuditFlag] = field(default_factory=list)
    rows: list[str] = field(default_factory=list)


def _norm_clasif(v) -> str | None:
    if v in (None, "", "-"):
        return None
    s = str(v).strip()
    if s.lower() in ("local", "visitante"):
        return s.capitalize()
    return s


def read_master_r32(master_path: Path, player: str) -> dict[int, tuple[int, int, str | None]]:
    wb = load_workbook(master_path, data_only=True)
    if player not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb[player]
    out: dict[int, tuple[int, int, str | None]] = {}
    for mid in range(R32_FIRST_ID, R32_LAST_ID + 1):
        row = PT_FIRST_ROW + mid - 1
        gh = ws.cell(row, 6).value
        ga = ws.cell(row, 7).value
        cl = ws.cell(row, 8).value
        try:
            gh_i = int(gh) if gh is not None and gh != "" else None
            ga_i = int(ga) if ga is not None and ga != "" else None
        except (TypeError, ValueError):
            gh_i, ga_i = None, None
        out[mid] = (gh_i, ga_i, _norm_clasif(cl))
    wb.close()
    return out


def read_master_results(master_path: Path) -> dict[int, tuple[int | None, int | None, str | None]]:
    wb = load_workbook(master_path, data_only=True)
    pt = wb["Partidos"]
    out: dict[int, tuple[int | None, int | None, str | None]] = {}
    for mid in range(R32_FIRST_ID, R32_LAST_ID + 1):
        row = PT_FIRST_ROW + mid - 1
        try:
            rh = int(pt.cell(row, 6).value) if pt.cell(row, 6).value not in (None, "") else None
            ra = int(pt.cell(row, 7).value) if pt.cell(row, 7).value not in (None, "") else None
        except (TypeError, ValueError):
            rh, ra = None, None
        rc = _norm_clasif(pt.cell(row, 11).value)
        out[mid] = (rh, ra, rc)
    wb.close()
    return out


def audit_player(
    source_path: str,
    group_id: str,
    player: str,
    label: str,
    master_path: Path,
    master_pairs: dict[int, tuple[str, str]],
    pair_index: dict,
    results: dict[int, tuple[int | None, int | None, str | None]],
) -> PlayerAudit:
    audit = PlayerAudit(label=label, player=player)
    src = Path(source_path)

    if not src.exists():
        audit.flags.append(
            AuditFlag("MISSING_SOURCE", player, None, f"Archivo no encontrado: {source_path}")
        )
        return audit

    parsed = parse_source(src, master_pairs, pair_index, label)
    master = read_master_r32(master_path, player)

    for flag in parsed.flags:
        kind = "INFERRED_CLASIF"
        if "pareja invertida" in flag.lower():
            kind = "SWAPPED"
        elif "no match" in flag.lower() or "faltan" in flag.lower():
            kind = "PARSE"
        elif "clasif" in flag.lower() or "inferido" in flag.lower():
            kind = "INFERRED_CLASIF"
        elif "id desfasado" in flag.lower():
            kind = "ID_OFFSET"
        audit.flags.append(AuditFlag(kind, player, None, flag))

    for mid in range(R32_FIRST_ID, R32_LAST_ID + 1):
        ml, mv = master_pairs[mid]
        src_pred = parsed.matched.get(mid)
        m_pred = master.get(mid)
        line_flags: list[str] = []

        if src_pred is None and m_pred and m_pred[0] is not None:
            line_flags.append("MISSING_SOURCE")
        if src_pred is not None and (m_pred is None or m_pred[0] is None):
            line_flags.append("MISSING_MASTER")

        if src_pred and m_pred and m_pred[0] is not None:
            s_gh, s_ga, s_cl = src_pred
            m_gh, m_ga, m_cl = m_pred
            if (s_gh, s_ga, s_cl) != (m_gh, m_ga, m_cl):
                line_flags.append("MISMATCH")
                audit.flags.append(
                    AuditFlag(
                        "MISMATCH",
                        player,
                        mid,
                        f"fuente {s_gh}-{s_ga} cl={s_cl or '?'} vs "
                        f"maestro {m_gh}-{m_ga} cl={m_cl or '?'}",
                    )
                )

        status = "OK" if not line_flags else ",".join(line_flags)
        pred = m_pred or src_pred
        extra = ""
        if mid in FINISHED_R32 and pred and pred[0] is not None:
            rh, ra, rc = results.get(mid, (None, None, None))
            if rh is not None and ra is not None:
                pts = calc_match_points("Dieciseisavos", pred[0], pred[1], rh, ra, pred[2], rc)
                tier = classify_match_tier(
                    "Dieciseisavos", pred[0], pred[1], rh, ra, pred[2], rc
                )
                extra = f" | real {rh}-{ra} cl={rc or '-'} -> {pred[0]}-{pred[1]} tier={tier} pts={pts}"

        gh, ga, cl = pred if pred else (None, None, None)
        audit.rows.append(
            f"  {mid:2d} {ml} vs {mv}: {gh}-{ga} cl={cl or '-'} [{status}]{extra}"
        )

    return audit


def run_audit(players_filter: set[str] | None = None) -> tuple[list[PlayerAudit], int]:
    audits: list[PlayerAudit] = []
    mismatch_count = 0
    groups_cache: dict[str, Path] = {}
    pairs_cache: dict[str, dict] = {}
    results_cache: dict[str, dict] = {}

    for source, group_id, player, label in IMPORTS:
        if players_filter and player not in players_filter:
            continue

        if group_id not in groups_cache:
            g = load_group(group_id)
            groups_cache[group_id] = g["excel_path"]
            pairs_cache[group_id] = load_master_r32(g["excel_path"])
            results_cache[group_id] = read_master_results(g["excel_path"])

        master_path = groups_cache[group_id]
        audit = audit_player(
            source,
            group_id,
            player,
            label,
            master_path,
            pairs_cache[group_id],
            build_pair_index(pairs_cache[group_id]),
            results_cache[group_id],
        )
        audits.append(audit)
        mismatch_count += sum(1 for f in audit.flags if f.kind == "MISMATCH")

    return audits, mismatch_count


def format_report(audits: list[PlayerAudit], mismatch_count: int) -> str:
    lines = [
        "AUDITORÍA PRONÓSTICOS R32 (73-88)",
        "=" * 60,
        "",
    ]

    focus = [a for a in audits if a.player in FOCUS_PLAYERS]
    others = [a for a in audits if a.player not in FOCUS_PLAYERS]

    for section, group in [("FOCUS: Pepe / Patri", focus), ("RESTO", others)]:
        if not group:
            continue
        lines.append(section)
        lines.append("-" * 40)
        for audit in group:
            lines.append(f"\n{audit.label} ({audit.player})")
            critical = [f for f in audit.flags if f.kind in ("MISMATCH", "MISSING_SOURCE", "MISSING_MASTER", "PARSE")]
            info = [f for f in audit.flags if f.kind not in ("MISMATCH", "MISSING_SOURCE", "MISSING_MASTER", "PARSE")]
            if critical:
                lines.append("  FLAGS CRÍTICOS:")
                for f in critical:
                    mid_s = f"#{f.mid} " if f.mid else ""
                    lines.append(f"    [{f.kind}] {mid_s}{f.detail}")
            if info:
                lines.append("  FLAGS INFORMATIVOS:")
                for f in info:
                    lines.append(f"    [{f.kind}] {f.detail}")
            if not critical and not info:
                lines.append("  Sin flags.")
            lines.extend(audit.rows)
        lines.append("")

    lines.append("=" * 60)
    if mismatch_count:
        lines.append(f"RESULTADO: {mismatch_count} MISMATCH(es) — revisar Excel maestro.")
    else:
        lines.append("RESULTADO: sin MISMATCH entre fuente y maestro.")
        lines.append("Pepe/Patri: datos coherentes con Excel original (si fuente disponible).")
        lines.append("La confusión visual en R32 jugados venía del tier 'clasificado' (morado).")
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(description="Auditar pronósticos R32 vs Excel fuente")
    parser.add_argument(
        "--focus-only",
        action="store_true",
        help="Solo Pepe y Patri",
    )
    parser.add_argument(
        "--all",
        action="store_true",
        help="Todos los jugadores importados (default)",
    )
    args = parser.parse_args()

    players_filter = FOCUS_PLAYERS if args.focus_only else None
    audits, mismatch_count = run_audit(players_filter)
    report = format_report(audits, mismatch_count)

    print(report)
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text(report, encoding="utf-8")
    print(f"\nInforme guardado en {REPORT_PATH}")

    if mismatch_count:
        sys.exit(1)


if __name__ == "__main__":
    main()
