"""
Genera dashboard_<grupo>.html leyendo Porra_Mundial_2026_<Grupo>.xlsx.
SPA de una sola página con rutas hash: #/ (inicio), #/partidos, #/jugador/<slug>.
"""
from __future__ import annotations

import json
import shutil
import unicodedata
import webbrowser
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

from config.groups import load_group, resolve_group_ids
from scoring import (
    TIER_LABELS,
    calc_match_points,
    calc_round_bonus,
    calc_special_points,
    classify_match_tier,
    load_scoring_config,
)

ROOT = Path(__file__).parent
OUTPUT = ROOT / "output"
MATCHES_JSON = ROOT / "data" / "matches_2026.json"
PT_FIRST_ROW = 4
SPECIAL_RESULTS_FIRST_ROW = 34

# Fallback si no se puede leer la hoja Resumen.
SPECIAL_BET_CATEGORIES_FALLBACK = [
    "Campeón del Mundial",
    "Subcampeón",
    "Tercer equipo",
    "Balón de oro de la FIFA",
    "Premio de la FIFA al mejor jugador joven",
    "Guante de oro de la FIFA",
    "Bota de oro de la FIFA",
    "Máximo goleador España",
]

def strip_accents(name: str) -> str:
    norm = unicodedata.normalize("NFKD", name)
    return "".join(c for c in norm if not unicodedata.combining(c))


def _player_photo_filename(name: str, photos_src: Path, photo_overrides: dict) -> str | None:
    """Nombre de jugador -> archivo PNG en photos_src, o None si no existe."""
    if name in photo_overrides:
        filename = photo_overrides[name]
    else:
        filename = f"{strip_accents(name)}.png"
    path = photos_src / filename
    if path.exists():
        return filename
    print(f"  Aviso: foto no encontrada para {name} ({path.name})")
    return None


def _save_jpeg(img: Image.Image, dst: Path, max_long_edge: int | None = None) -> None:
    img = img.convert("RGB")
    if max_long_edge:
        w, h = img.size
        if max(w, h) > max_long_edge:
            scale = max_long_edge / max(w, h)
            img = img.resize((int(w * scale), int(h * scale)), Image.Resampling.LANCZOS)
    dst.parent.mkdir(parents=True, exist_ok=True)
    img.save(dst, "JPEG", quality=85, optimize=True)


def copy_dashboard_assets(players: list[str], group: dict) -> dict:
    """Copia y optimiza portada + fotos de jugadores a output/<assets_dir>/."""
    result: dict = {"portada": "", "players": {}}
    assets_path = group["assets_path"]
    assets_path.mkdir(parents=True, exist_ok=True)
    photos_src = group["photos_path"]
    photo_overrides = group["photo_overrides"]
    assets_prefix = group["assets_dir"]

    portada_src = photos_src / "Portada.png"
    if portada_src.exists():
        with Image.open(portada_src) as im:
            im = im.convert("RGB")
            w, h = im.size
            if w > 1600:
                im = im.resize((1600, int(h * 1600 / w)), Image.Resampling.LANCZOS)
                w, h = im.size
            dst = assets_path / "portada.jpg"
            im.save(dst, "JPEG", quality=85, optimize=True)
            result["portada"] = f"{assets_prefix}/portada.jpg"
            result["portada_aspect"] = round(w / h, 4)
            print(f"  Portada: {dst}")
    else:
        src = ROOT / "assets" / "wc2026_stadium.jpg"
        if not src.exists():
            for alt in ("football.jpg", "wc2026_crowd.jpg"):
                p = ROOT / "assets" / alt
                if p.exists() and p.stat().st_size > 1000:
                    src = p
                    break
        dst = assets_path / "hero.jpg"
        if src.exists() and src.stat().st_size > 1000:
            shutil.copy2(src, dst)
            result["portada"] = f"{assets_prefix}/hero.jpg"
            print(f"  Portada (fallback): {dst}")

    photos_dir = assets_path / "photos"
    photos_dir.mkdir(parents=True, exist_ok=True)
    for player in players:
        filename = _player_photo_filename(player, photos_src, photo_overrides)
        if not filename:
            continue
        slug = slugify(player)
        dst = photos_dir / f"{slug}.jpg"
        with Image.open(photos_src / filename) as im:
            _save_jpeg(im, dst, max_long_edge=480)
        result["players"][slug] = f"{assets_prefix}/photos/{slug}.jpg"
        print(f"  Foto {player}: {result['players'][slug]}")

    return result


def load_json(path: Path):
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def slugify(name: str) -> str:
    """Nombre de jugador -> slug para la ruta hash (minúsculas, sin acentos)."""
    import unicodedata

    norm = unicodedata.normalize("NFKD", name)
    ascii_name = "".join(c for c in norm if not unicodedata.combining(c))
    return "".join(c if c.isalnum() else "-" for c in ascii_name.lower()).strip("-")


def assign_player_initials(players: list[str]) -> dict[str, str]:
    """Iniciales únicas por jugador dentro del grupo (prefijo más corto sin colisión)."""
    result: dict[str, str] = {}
    used: set[str] = set()
    for player in players:
        base = strip_accents(player)
        for length in range(1, len(base) + 1):
            candidate = base[:length].upper()
            if candidate not in used:
                result[player] = candidate
                used.add(candidate)
                break
        else:
            candidate = base.upper()
            n = 1
            while candidate in used:
                n += 1
                candidate = f"{base[:2].upper()}{n}"
            result[player] = candidate
            used.add(candidate)
    return result


def is_finished(res_h, res_a) -> bool:
    return res_h is not None and res_a is not None


def read_special_categories(wb) -> list[str]:
    """Lee el orden y nombres de apuestas especiales desde Resumen (col. A)."""
    if "Resumen" not in wb.sheetnames:
        return list(SPECIAL_BET_CATEGORIES_FALLBACK)
    summary = wb["Resumen"]
    categories: list[str] = []
    row = SPECIAL_RESULTS_FIRST_ROW
    while True:
        val = summary.cell(row, 1).value
        if val is None or str(val).strip() == "":
            break
        categories.append(str(val).strip())
        row += 1
    return categories or list(SPECIAL_BET_CATEGORIES_FALLBACK)


def read_workbook_data(excel_path: Path, players: list[str], n_matches: int) -> dict:
    wb = load_workbook(excel_path, data_only=True)
    pt = wb["Partidos"]
    last_row = PT_FIRST_ROW + n_matches - 1

    matches_meta: dict[int, dict] = {}
    results: dict[int, tuple[int | None, int | None]] = {}
    clasificados: dict[int, str | None] = {}
    for r in range(PT_FIRST_ROW, last_row + 1):
        mid = pt.cell(r, 1).value
        if mid is None:
            continue
        mid = int(mid)
        results[mid] = (pt.cell(r, 6).value, pt.cell(r, 7).value)
        clasificados[mid] = pt.cell(r, 11).value
        matches_meta[mid] = {
            "id": mid,
            "fecha": pt.cell(r, 2).value or "",
            "local": pt.cell(r, 4).value or "",
            "visitante": pt.cell(r, 5).value or "",
            "jornada": pt.cell(r, 9).value,
            "fase": pt.cell(r, 10).value or "",
        }

    special_categories = read_special_categories(wb)
    official_specials: list = []
    if "Resumen" in wb.sheetnames:
        summary = wb["Resumen"]
        official_specials = [
            summary.cell(SPECIAL_RESULTS_FIRST_ROW + idx, 2).value
            for idx in range(len(special_categories))
        ]

    predictions: dict[tuple[str, int], tuple] = {}
    special_bets: dict[tuple[str, int], tuple] = {}
    special_first_row = last_row + 4
    for player in players:
        if player not in wb.sheetnames:
            continue
        ws = wb[player]
        for r in range(PT_FIRST_ROW, last_row + 1):
            mid = ws.cell(r, 1).value
            if mid is None:
                continue
            predictions[(player, int(mid))] = (
                ws.cell(r, 6).value,
                ws.cell(r, 7).value,
                ws.cell(r, 8).value,
            )
        for idx, cat in enumerate(special_categories):
            row = special_first_row + idx
            sheet_cat = ws.cell(row, 1).value
            if sheet_cat and str(sheet_cat).strip() != cat:
                print(
                    f"  Aviso: en {player} fila {row} "
                    f"se esperaba '{cat}' y hay '{sheet_cat}'"
                )
            official = official_specials[idx] if idx < len(official_specials) else None
            pred = ws.cell(row, 2).value
            special_bets[(player, idx)] = (pred, official)

    wb.close()

    cfg = load_scoring_config()
    points: dict[tuple[str, int], int | None] = {}
    for (player, mid), (ph, pa, pc) in predictions.items():
        rh, ra = results.get(mid, (None, None))
        fase = matches_meta.get(mid, {}).get("fase", "")
        rc = clasificados.get(mid)
        if not is_finished(rh, ra):
            points[(player, mid)] = None
        else:
            points[(player, mid)] = calc_match_points(fase, ph, pa, rh, ra, pc, rc, cfg)

    return {
        "matches_meta": matches_meta,
        "results": results,
        "predictions": predictions,
        "points": points,
        "clasificados": clasificados,
        "special_categories": special_categories,
        "special_bets": special_bets,
    }


def build_stats(
    players: list[str],
    raw: dict,
    n_matches: int,
    initials_map: dict[str, str] | None = None,
) -> dict:
    matches_meta = raw["matches_meta"]
    results = raw["results"]
    points = raw["points"]
    clasificados = raw["clasificados"]
    predictions = raw["predictions"]
    special_bets = raw["special_bets"]
    special_categories = raw["special_categories"]
    cfg = load_scoring_config()
    initials_map = initials_map or {}

    finished = [
        mid
        for mid, (h, a) in results.items()
        if h is not None and a is not None
    ]
    finished.sort()

    player_stats = []
    for player in players:
        total_match = 0
        exact = partial = fails = 0
        for mid in finished:
            p = points.get((player, mid))
            if p is None:
                continue
            total_match += p
            if p in (cfg["grupos"]["exacto"], cfg["eliminatorias"]["exacto"]):
                exact += 1
            elif p == 0:
                fails += 1
            else:
                partial += 1

        bonus_total = 0
        for fase_name, n_expected in cfg.get("rondas_bonus", {}).items():
            mids = [m for m in finished if matches_meta.get(m, {}).get("fase") == fase_name]
            finished_in_round = len(mids)
            correct = 0
            for mid in mids:
                rc = clasificados.get(mid)
                pc = predictions.get((player, mid), (None, None, None))[2]
                if rc and pc and str(rc).strip() == str(pc).strip():
                    correct += 1
            bkey = {"Octavos": "octavos", "Cuartos": "cuartos", "Semifinal": "semifinal"}.get(fase_name)
            bpts = cfg.get("bonus_ronda", {}).get(bkey or "", 0)
            bonus_total += calc_round_bonus(correct, finished_in_round, n_expected, bpts)

        special_total = 0
        for idx in range(len(special_categories)):
            pred_off = special_bets.get((player, idx))
            if pred_off:
                pred, official = pred_off
                special_total += calc_special_points(pred, official, cfg)

        total = total_match + bonus_total + special_total
        n_fin = len(finished)
        pct = (exact + partial) / n_fin if n_fin else 0
        player_stats.append(
            {
                "name": player,
                "points": total,
                "points_match": total_match,
                "points_bonus": bonus_total,
                "points_special": special_total,
                "exact": exact,
                "partial": partial,
                "fails": fails,
                "pct": round(pct * 100, 1),
            }
        )

    player_stats.sort(key=lambda x: (-x["points"], -x["exact"], x["fails"]))

    evolution_labels = list(range(1, min(16, n_matches) + 1))
    evolution: dict[str, list] = {}
    for ps in player_stats[:3]:
        name = ps["name"]
        cum = []
        for n in evolution_labels:
            s = 0
            for mid in finished:
                if mid <= n:
                    p = points.get((name, mid))
                    if p is not None:
                        s += p
            cum.append(s)
        evolution[name] = cum

    last_matches = []
    for mid in reversed(finished[-5:]):
        m = matches_meta[mid]
        h, a = results[mid]
        match_players = []
        for player in players:
            ph, pa, pc = predictions.get((player, mid), (None, None, None))
            tier = classify_match_tier(
                m["fase"], ph, pa, h, a, pc, clasificados.get(mid)
            )
            pts = points.get((player, mid))
            if ph is not None and pa is not None:
                pred_str = f"{int(ph)}-{int(pa)}"
            else:
                pred_str = "—"
            match_players.append(
                {
                    "name": player,
                    "initials": initials_map.get(player, player[:1]),
                    "tier": tier,
                    "points": pts,
                    "pred": pred_str,
                }
            )
        last_matches.append(
            {
                "id": mid,
                "local": m["local"],
                "visitante": m["visitante"],
                "score": f"{int(h)}-{int(a)}",
                "fecha": str(m["fecha"]),
                "players": match_players,
            }
        )

    leader = player_stats[0] if player_stats else {"name": "—", "points": 0}
    n_fin = len(finished)

    return {
        "generated_at": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "leader": leader,
        "finished_count": n_fin,
        "total_matches": n_matches,
        "progress_pct": round(100 * n_fin / n_matches, 1) if n_matches else 0,
        "avg_pts": round(leader["points"] / n_fin, 2) if n_fin else 0,
        "ranking": player_stats,
        "last_matches": last_matches,
        "evolution_labels": evolution_labels,
        "evolution": evolution,
    }


def build_upcoming_matches(
    players: list[str],
    matches_meta: dict,
    results: dict,
    predictions: dict,
    initials_map: dict[str, str],
    limit: int = 5,
) -> list[dict]:
    """Próximos partidos pendientes con pronósticos de cada jugador."""
    upcoming: list[dict] = []
    for mid in sorted(matches_meta):
        rh, ra = results.get(mid, (None, None))
        if is_finished(rh, ra):
            continue
        m = matches_meta[mid]
        preds = []
        for player in players:
            ph, pa, pc = predictions.get((player, mid), (None, None, None))
            if ph is not None and pa is not None:
                pred = f"{int(ph)}-{int(pa)}"
            else:
                pred = "—"
            preds.append(
                {
                    "name": player,
                    "initials": initials_map.get(player, player[:1]),
                    "pred": pred,
                    "pred_clasificado": pc or "",
                }
            )
        upcoming.append(
            {
                "id": mid,
                "fecha": str(m["fecha"]),
                "fase": m["fase"],
                "local": m["local"],
                "visitante": m["visitante"],
                "predictions": preds,
            }
        )
        if len(upcoming) >= limit:
            break
    return upcoming


def build_full_payload(
    players: list[str],
    raw: dict,
    n_matches: int,
    group: dict,
    asset_paths: dict | None = None,
) -> dict:
    """Payload completo del SPA: home + todos los partidos + ficha por jugador."""
    initials_map = assign_player_initials(players)
    stats = build_stats(players, raw, n_matches, initials_map)
    matches_meta = raw["matches_meta"]
    results = raw["results"]
    predictions = raw["predictions"]
    points = raw["points"]
    clasificados = raw["clasificados"]
    special_bets = raw["special_bets"]
    special_categories = raw["special_categories"]
    cfg = load_scoring_config()

    rank_pos = {ps["name"]: i + 1 for i, ps in enumerate(stats["ranking"])}

    # ---- Lista de partidos (orden por id) ----
    matches = []
    for mid in sorted(matches_meta):
        m = matches_meta[mid]
        rh, ra = results.get(mid, (None, None))
        finished = is_finished(rh, ra)
        matches.append(
            {
                "id": mid,
                "fecha": str(m["fecha"]),
                "fase": m["fase"],
                "jornada": m["jornada"],
                "local": m["local"],
                "visitante": m["visitante"],
                "res_h": rh if finished else None,
                "res_a": ra if finished else None,
                "score": f"{int(rh)}-{int(ra)}" if finished else "—",
                "clasificado": clasificados.get(mid) or "",
                "finished": finished,
            }
        )

    # ---- Resumen de apuestas especiales (oficial + quién acertó) ----
    specials_summary = []
    for idx, cat in enumerate(special_categories):
        official = None
        for p in players:
            pred_off = special_bets.get((p, idx))
            if pred_off and pred_off[1] not in (None, ""):
                official = pred_off[1]
                break
        hitters = []
        if official not in (None, ""):
            for p in players:
                pred_off = special_bets.get((p, idx))
                if pred_off and calc_special_points(pred_off[0], pred_off[1], cfg) > 0:
                    hitters.append(p)
        specials_summary.append(
            {
                "categoria": cat,
                "oficial": official if official not in (None, "") else "",
                "aciertos": hitters,
            }
        )

    asset_paths = asset_paths or {"portada": "", "players": {}}
    player_photos = asset_paths.get("players", {})

    # ---- Ficha por jugador ----
    players_data = {}
    for player in players:
        ps = next((x for x in stats["ranking"] if x["name"] == player), None)
        player_matches = []
        for mid in sorted(matches_meta):
            m = matches_meta[mid]
            rh, ra = results.get(mid, (None, None))
            ph, pa, pc = predictions.get((player, mid), (None, None, None))
            tier = classify_match_tier(
                m["fase"], ph, pa, rh, ra, pc, clasificados.get(mid)
            )
            finished = is_finished(rh, ra)
            player_matches.append(
                {
                    "id": mid,
                    "fecha": str(m["fecha"]),
                    "fase": m["fase"],
                    "local": m["local"],
                    "visitante": m["visitante"],
                    "pred_h": ph if ph not in (None, "") else None,
                    "pred_a": pa if pa not in (None, "") else None,
                    "pred_clasificado": pc or "",
                    "real_h": rh if finished else None,
                    "real_a": ra if finished else None,
                    "real_clasificado": clasificados.get(mid) or "",
                    "tier": tier,
                    "tier_label": TIER_LABELS.get(tier, tier),
                    "points": points.get((player, mid)),
                }
            )

        specials = []
        for idx, cat in enumerate(special_categories):
            pred, official = special_bets.get((player, idx), (None, None))
            has_official = official not in (None, "")
            pts = calc_special_points(pred, official, cfg) if has_official else 0
            specials.append(
                {
                    "categoria": cat,
                    "prediccion": pred if pred not in (None, "") else "",
                    "oficial": official if has_official else "",
                    "hit": pts > 0,
                    "points": pts,
                }
            )

        slug = slugify(player)
        players_data[player] = {
            "name": player,
            "slug": slug,
            "initials": initials_map.get(player, player[:1]),
            "rank": rank_pos.get(player),
            "photo": player_photos.get(slug, ""),
            "stats": ps
            or {
                "name": player,
                "points": 0,
                "points_match": 0,
                "points_bonus": 0,
                "points_special": 0,
                "exact": 0,
                "partial": 0,
                "fails": 0,
                "pct": 0,
            },
            "matches": player_matches,
            "specials": specials,
        }

    players_index = [
        {
            "name": p,
            "slug": slugify(p),
            "rank": rank_pos.get(p),
            "initials": initials_map.get(p, p[:1]),
        }
        for p in players
    ]

    upcoming_matches = build_upcoming_matches(
        players, matches_meta, results, predictions, initials_map
    )

    return {
        "meta": {
            "generated_at": stats["generated_at"],
            "total_matches": n_matches,
            "finished_count": stats["finished_count"],
            "progress_pct": stats["progress_pct"],
            "avg_pts": stats["avg_pts"],
            "special_categories": special_categories,
            "group_id": group["id"],
            "group_label": group["label"],
            "excel_name": group["excel"],
            "evo_top": min(3, len(players)),
        },
        "home": {
            "leader": stats["leader"],
            "ranking": stats["ranking"],
            "ranking_by_pct": sorted(
                stats["ranking"], key=lambda x: (-x["pct"], -x["points"])
            ),
            "ranking_by_exact": sorted(
                stats["ranking"], key=lambda x: (-x["exact"], -x["points"])
            ),
            "last_matches": stats["last_matches"],
            "upcoming_matches": upcoming_matches,
            "evolution_labels": stats["evolution_labels"],
            "evolution": stats["evolution"],
            "specials_summary": specials_summary,
        },
        "matches": matches,
        "players_index": players_index,
        "players": players_data,
        "assets": {
            "portada": asset_paths.get("portada", ""),
            "portada_aspect": asset_paths.get("portada_aspect"),
        },
    }


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Porra Mundial 2026 — __GROUP_LABEL__</title>
  <link rel="preconnect" href="https://fonts.googleapis.com" />
  <link href="https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Inter:wght@400;600;700&display=swap" rel="stylesheet" />
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
  <style>
    :root {
      --navy: #051C2C;
      --blue: #0066CC;
      --gold: #C5A028;
      --green: #2E7D32;
      --red: #C62828;
      --bg: #0a1628;
      --card: rgba(255,255,255,0.06);
      --border: rgba(255,255,255,0.12);
      --text: #f5f5f5;
      --muted: #94a3b8;
      --t-exact: #15803d;
      --t-diff: #0e7490;
      --t-winner: #b45309;
      --t-clasif: #6d28d9;
      --t-miss: #7f1d1d;
      --t-pending: #334155;
    }
    * { box-sizing: border-box; margin: 0; padding: 0; }
    body {
      font-family: 'Inter', system-ui, sans-serif;
      background: var(--bg);
      color: var(--text);
      min-height: 100vh;
      line-height: 1.5;
    }
    a { color: inherit; text-decoration: none; }
    .hero {
      position: relative;
      aspect-ratio: __HERO_ASPECT__;
      min-height: __HERO_MIN_HEIGHT__;
      background: __HERO_CSS__;
      background-position: center center;
      display: flex;
      flex-direction: column;
      justify-content: flex-end;
      padding: 1.75rem clamp(1rem, 4vw, 3rem);
      overflow: hidden;
    }
    .hero::after {
      content: '';
      position: absolute;
      inset: 0;
      background: linear-gradient(to top, var(--bg) 0%, transparent 65%),
                  linear-gradient(90deg, rgba(5,28,44,.55) 0%, transparent 60%);
    }
    .hero-content { position: relative; z-index: 1; }
    .hero-badge {
      display: inline-block;
      background: var(--gold);
      color: var(--navy);
      font-size: 0.7rem;
      font-weight: 700;
      letter-spacing: 0.15em;
      padding: 0.35rem 0.75rem;
      border-radius: 2px;
      margin-bottom: 0.6rem;
    }
    .hero h1 {
      font-family: 'Bebas Neue', sans-serif;
      font-size: clamp(2.4rem, 7vw, 4.5rem);
      letter-spacing: 0.04em;
      line-height: 0.95;
    }
    .hero-sub { color: #cbd5e1; font-size: 0.95rem; margin-top: 0.25rem; }
    .hosts { margin-top: 0.5rem; display: flex; gap: 1rem; font-size: 0.85rem; color: #cbd5e1; }

    nav.tabs {
      position: sticky;
      top: 0;
      z-index: 20;
      display: flex;
      align-items: center;
      flex-wrap: wrap;
      gap: 0.25rem;
      padding: 0.5rem clamp(1rem, 4vw, 3rem);
      background: rgba(5,28,44,0.92);
      backdrop-filter: blur(8px);
      border-bottom: 1px solid var(--border);
    }
    nav.tabs a.tab, nav.tabs .dropdown > button {
      font: inherit;
      color: var(--muted);
      background: none;
      border: none;
      cursor: pointer;
      padding: 0.5rem 0.9rem;
      border-radius: 6px;
      font-weight: 600;
      font-size: 0.9rem;
    }
    nav.tabs a.tab:hover, nav.tabs .dropdown > button:hover { color: var(--text); background: rgba(255,255,255,0.06); }
    nav.tabs a.tab.active { color: var(--gold); background: rgba(197,160,40,0.12); }
    .dropdown { position: relative; }
    .dropdown-menu {
      position: absolute;
      top: 110%;
      left: 0;
      min-width: 200px;
      background: #0d2034;
      border: 1px solid var(--border);
      border-radius: 8px;
      padding: 0.4rem;
      display: none;
      grid-template-columns: 1fr 1fr;
      gap: 0.2rem;
      box-shadow: 0 12px 30px rgba(0,0,0,.4);
    }
    .dropdown.open .dropdown-menu { display: grid; }
    .dropdown-menu a {
      display: flex;
      justify-content: space-between;
      gap: 0.5rem;
      padding: 0.45rem 0.6rem;
      border-radius: 5px;
      font-size: 0.85rem;
      color: var(--text);
    }
    .dropdown-menu a:hover { background: rgba(255,255,255,0.08); }
    .dropdown-menu a .rk { color: var(--muted); font-size: 0.75rem; }

    main { padding: 1.5rem clamp(1rem, 4vw, 3rem) 3rem; max-width: 1180px; margin: 0 auto; }
    h2 { font-size: 1.05rem; margin-bottom: 0.75rem; letter-spacing: 0.01em; }
    .section-title { font-family: 'Bebas Neue', sans-serif; font-size: 1.8rem; letter-spacing: 0.03em; margin: 0.5rem 0 1rem; }
    .card {
      background: var(--card);
      border: 1px solid var(--border);
      border-radius: 12px;
      padding: 1.25rem;
      margin-bottom: 1.25rem;
    }
    .grid-2 { display: grid; grid-template-columns: 1fr 1fr; gap: 1.25rem; }
    .grid-3 { display: grid; grid-template-columns: repeat(3, 1fr); gap: 1.25rem; }
    @media (max-width: 860px) { .grid-2, .grid-3 { grid-template-columns: 1fr; } }

    .kpis { display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 0.75rem; margin-bottom: 1.25rem; }
    .kpi { background: var(--card); border: 1px solid var(--border); border-radius: 10px; padding: 0.9rem 1rem; }
    .kpi-label { color: var(--muted); font-size: 0.72rem; text-transform: uppercase; letter-spacing: 0.08em; }
    .kpi-value { font-family: 'Bebas Neue', sans-serif; font-size: 1.9rem; color: var(--gold); line-height: 1.1; }

    table { width: 100%; border-collapse: collapse; font-size: 0.88rem; }
    th, td { text-align: left; padding: 0.5rem 0.6rem; border-bottom: 1px solid var(--border); }
    th { color: var(--muted); font-size: 0.72rem; text-transform: uppercase; letter-spacing: 0.06em; font-weight: 600; }
    td.pos, th.pos { width: 2rem; color: var(--muted); }
    tr.me td { background: rgba(197,160,40,0.10); }
    .medal-1 { color: #ffd700; font-weight: 700; }
    .medal-2 { color: #c0c0c0; font-weight: 700; }
    .medal-3 { color: #cd7f32; font-weight: 700; }

    .chart-wrap { position: relative; height: 300px; }
    .chart-wrap.tall { height: 360px; }

    .match-row { display: flex; justify-content: space-between; align-items: flex-start; padding: 0.6rem 0; border-bottom: 1px solid var(--border); gap: 0.75rem; }
    .match-row:last-child { border-bottom: none; }
    .match-row-main { flex: 1; min-width: 0; }
    .match-score { font-family: 'Bebas Neue', sans-serif; font-size: 1.4rem; color: var(--gold); min-width: 3rem; text-align: center; flex-shrink: 0; }
    .last-players { display: flex; flex-wrap: wrap; gap: 0.3rem; margin-top: 0.35rem; }
    .init-badge {
      display: inline-flex; align-items: center; justify-content: center;
      width: 26px; height: 26px; border-radius: 50%; font-size: 0.62rem; font-weight: 700;
      border: 2px solid var(--t-pending); color: var(--text); background: transparent;
      flex-shrink: 0;
    }
    .init-badge.exact { border-color: var(--t-exact); }
    .init-badge.difference { border-color: var(--t-diff); }
    .init-badge.winner { border-color: var(--t-winner); }
    .init-badge.clasificado { border-color: var(--t-clasif); }
    .init-badge.miss { border-color: var(--t-miss); }
    .init-badge.pending { border-color: var(--t-pending); opacity: 0.65; }
    .tier-legend-mini { display: flex; flex-wrap: wrap; gap: 0.45rem 0.75rem; margin-top: 0.65rem; padding-top: 0.5rem; border-top: 1px solid var(--border); }
    .tier-legend-mini span { display: inline-flex; align-items: center; gap: 0.3rem; font-size: 0.68rem; color: var(--muted); }
    .tier-legend-mini i { width: 10px; height: 10px; border-radius: 50%; border: 2px solid currentColor; display: inline-block; background: transparent; }
    .tier-legend-mini .exact i { color: var(--t-exact); }
    .tier-legend-mini .difference i { color: var(--t-diff); }
    .tier-legend-mini .winner i { color: var(--t-winner); }
    .tier-legend-mini .clasificado i { color: var(--t-clasif); }
    .tier-legend-mini .miss i { color: var(--t-miss); }

    .upcoming-list { display: flex; flex-direction: column; gap: 0.85rem; }
    .upcoming-block {
      border: 1px solid var(--border); border-radius: 10px; padding: 0.85rem 1rem;
      background: rgba(255,255,255,0.03);
    }
    .upcoming-head { display: flex; flex-wrap: wrap; justify-content: space-between; align-items: baseline; gap: 0.4rem; margin-bottom: 0.65rem; }
    .upcoming-teams { font-weight: 700; font-size: 1rem; }
    .upcoming-meta { color: var(--muted); font-size: 0.78rem; }
    .upcoming-preds { display: grid; grid-template-columns: repeat(auto-fill, minmax(72px, 1fr)); gap: 0.45rem; }
    .pred-mini {
      display: flex; flex-direction: column; align-items: center; gap: 0.15rem;
      padding: 0.35rem 0.25rem; border-radius: 8px; background: rgba(255,255,255,0.04);
      border: 1px solid var(--border); font-size: 0.78rem;
    }
    .pred-mini .ini { font-weight: 700; font-size: 0.68rem; color: var(--gold); }
    .pred-mini .sc-pred { font-family: 'Bebas Neue', sans-serif; font-size: 1.05rem; letter-spacing: 0.02em; }
    .pred-mini .sc-pred.muted { color: var(--muted); font-family: inherit; font-size: 0.85rem; }

    .mt-row.clickable { cursor: pointer; border-radius: 6px; transition: background 0.15s; }
    .mt-row.clickable:hover { background: rgba(255,255,255,0.05); }

    .modal-overlay {
      position: fixed; inset: 0; z-index: 100;
      background: rgba(5,28,44,0.82); backdrop-filter: blur(4px);
      display: flex; align-items: center; justify-content: center;
      padding: 1rem; overflow-y: auto;
    }
    .modal-box {
      background: #0d2034; border: 1px solid var(--border); border-radius: 12px;
      width: min(560px, 100%); max-height: min(85vh, 640px); overflow-y: auto;
      padding: 1.1rem 1.25rem; position: relative;
      box-shadow: 0 20px 50px rgba(0,0,0,0.45);
    }
    .modal-close {
      position: absolute; top: 0.65rem; right: 0.75rem;
      background: none; border: none; color: var(--muted); font-size: 1.4rem;
      cursor: pointer; line-height: 1; padding: 0.2rem;
    }
    .modal-close:hover { color: var(--text); }
    .modal-head { padding-right: 1.5rem; margin-bottom: 0.85rem; }
    .modal-head h3 { font-family: 'Bebas Neue', sans-serif; font-size: 1.45rem; letter-spacing: 0.03em; color: var(--gold); }
    .modal-head .modal-meta { color: var(--muted); font-size: 0.82rem; margin-top: 0.15rem; }
    .modal-score { font-family: 'Bebas Neue', sans-serif; font-size: 1.6rem; color: var(--gold); margin-top: 0.35rem; }
    .modal-table { width: 100%; font-size: 0.85rem; }
    .modal-table th { font-size: 0.68rem; }
    .modal-table td { vertical-align: middle; }

    .filters { display: flex; flex-wrap: wrap; gap: 0.6rem; margin-bottom: 1rem; }
    .filters select, .filters input {
      font: inherit; font-size: 0.85rem; background: #0d2034; color: var(--text);
      border: 1px solid var(--border); border-radius: 8px; padding: 0.45rem 0.7rem;
    }
    .phase-group { margin-bottom: 1.25rem; }
    .phase-head { font-family: 'Bebas Neue', sans-serif; font-size: 1.3rem; color: var(--gold); letter-spacing: 0.03em; margin: 0.5rem 0; }
    .mt-row {
      display: grid; grid-template-columns: 92px 1fr auto auto; align-items: center; gap: 0.6rem;
      padding: 0.55rem 0.4rem; border-bottom: 1px solid var(--border); font-size: 0.88rem;
    }
    .mt-row .teams { font-weight: 600; }
    .mt-row .date { color: var(--muted); font-size: 0.78rem; }
    .mt-row .sc { font-family: 'Bebas Neue', sans-serif; font-size: 1.25rem; color: var(--gold); min-width: 3rem; text-align: center; }
    .mt-row .clasif { color: var(--muted); font-size: 0.78rem; }
    .pill { display: inline-block; padding: 0.1rem 0.5rem; border-radius: 999px; font-size: 0.7rem; font-weight: 600; }
    .pill.fin { background: rgba(46,125,50,0.25); color: #86efac; }
    .pill.pend { background: rgba(148,163,184,0.18); color: var(--muted); }

    /* Tarjetas pronóstico por jugador */
    .pred-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(260px, 1fr)); gap: 0.75rem; }
    .pred-card {
      border: 1px solid var(--border); border-radius: 10px; padding: 0.7rem 0.8rem;
      border-left: 5px solid var(--t-pending); background: rgba(255,255,255,0.03);
    }
    .pred-card.exact { border-left-color: var(--t-exact); }
    .pred-card.difference { border-left-color: var(--t-diff); }
    .pred-card.winner { border-left-color: var(--t-winner); }
    .pred-card.clasificado { border-left-color: var(--t-clasif); }
    .pred-card.miss { border-left-color: var(--t-miss); }
    .pred-card.pending { border-left-color: var(--t-pending); opacity: 0.85; }
    .pred-card .pc-teams { font-weight: 600; font-size: 0.9rem; }
    .pred-card .pc-date { color: var(--muted); font-size: 0.72rem; margin-bottom: 0.4rem; }
    .pred-line { display: flex; justify-content: space-between; align-items: center; font-size: 0.85rem; margin-top: 0.2rem; }
    .pred-line .lbl { color: var(--muted); }
    .score-chip { font-family: 'Bebas Neue', sans-serif; font-size: 1.15rem; letter-spacing: 0.03em; }
    .pc-foot { display: flex; justify-content: space-between; align-items: center; margin-top: 0.5rem; }
    .tier-tag { font-size: 0.7rem; font-weight: 700; padding: 0.12rem 0.5rem; border-radius: 999px; }
    .tier-tag.exact { background: var(--t-exact); }
    .tier-tag.difference { background: var(--t-diff); }
    .tier-tag.winner { background: var(--t-winner); }
    .tier-tag.clasificado { background: var(--t-clasif); }
    .tier-tag.miss { background: var(--t-miss); }
    .tier-tag.pending { background: var(--t-pending); color: #cbd5e1; }
    .pts-badge { font-weight: 700; color: var(--gold); font-size: 0.9rem; }
    .clasif-line { font-size: 0.74rem; color: var(--muted); margin-top: 0.35rem; }

    .legend { display: flex; flex-wrap: wrap; gap: 0.6rem; margin: 0.5rem 0 1rem; }
    .legend span { display: inline-flex; align-items: center; gap: 0.35rem; font-size: 0.78rem; color: var(--muted); }
    .legend i { width: 12px; height: 12px; border-radius: 3px; display: inline-block; }

    .spec-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(220px, 1fr)); gap: 0.75rem; }
    .spec-card { border: 1px solid var(--border); border-radius: 10px; padding: 0.75rem 0.85rem; border-top: 4px solid var(--t-pending); }
    .spec-card.hit { border-top-color: var(--t-exact); }
    .spec-card.fail { border-top-color: var(--t-miss); }
    .spec-card .cat { font-size: 0.72rem; text-transform: uppercase; letter-spacing: 0.06em; color: var(--muted); }
    .spec-card .pred { font-weight: 700; font-size: 1rem; margin: 0.2rem 0; }
    .spec-card .off { font-size: 0.8rem; color: var(--muted); }
    .spec-status { font-size: 0.74rem; font-weight: 700; margin-top: 0.4rem; }
    .spec-status.hit { color: #86efac; }
    .spec-status.fail { color: #fca5a5; }
    .spec-status.pend { color: var(--muted); }

    .chips { display: flex; flex-wrap: wrap; gap: 0.4rem; }
    .chip { padding: 0.2rem 0.55rem; border-radius: 999px; background: rgba(255,255,255,0.08); font-size: 0.78rem; }

    .player-top {
      position: relative;
      margin-bottom: 1rem;
      padding-right: 200px;
    }
    .player-photo {
      position: absolute;
      top: 0;
      right: 0;
      width: auto;
      height: auto;
      max-width: 180px;
      max-height: 260px;
      object-fit: contain;
      border-radius: 12px;
      border: 3px solid var(--gold);
      box-shadow: 0 8px 24px rgba(0,0,0,0.45);
    }
    @media (max-width: 520px) {
      .player-top { padding-right: 0; padding-top: 280px; }
      .player-photo { right: 0; left: 0; margin: 0 auto; max-width: 160px; max-height: 230px; }
    }
    .player-head { display: flex; flex-wrap: wrap; align-items: baseline; gap: 0.75rem; margin-bottom: 0.5rem; }
    .player-head .pname { font-family: 'Bebas Neue', sans-serif; font-size: 2.6rem; letter-spacing: 0.03em; }
    .player-head .prank { color: var(--gold); font-weight: 700; }
    .breakdown { display: flex; flex-wrap: wrap; gap: 1.25rem; margin-top: 0.25rem; }
    .breakdown div span { display: block; }
    .breakdown .bk-val { font-family: 'Bebas Neue', sans-serif; font-size: 1.7rem; color: var(--gold); }
    .breakdown .bk-lbl { color: var(--muted); font-size: 0.72rem; text-transform: uppercase; letter-spacing: 0.06em; }

    .sync-note {
      background: rgba(0,102,204,.15); border: 1px solid var(--blue); border-radius: 8px;
      padding: 0.6rem 0.9rem; font-size: 0.78rem; color: var(--muted); margin-bottom: 1.25rem;
    }
    .footer { text-align: center; color: var(--muted); font-size: 0.75rem; margin-top: 2rem; padding-top: 1rem; border-top: 1px solid var(--border); }
    .muted { color: var(--muted); }
    .toggle-cuts { display: flex; gap: 0.4rem; margin-bottom: 0.75rem; }
    .toggle-cuts button {
      font: inherit; font-size: 0.8rem; font-weight: 600; cursor: pointer;
      background: rgba(255,255,255,0.06); color: var(--muted);
      border: 1px solid var(--border); border-radius: 999px; padding: 0.35rem 0.8rem;
    }
    .toggle-cuts button.active { background: var(--gold); color: var(--navy); border-color: var(--gold); }
  </style>
</head>
<body>
  <header class="hero">
    <div class="hero-content">
      <span class="hero-badge">FIFA WORLD CUP 26</span>
      <h1>__GROUP_TITLE__</h1>
      <p class="hero-sub">USA · Canadá · México</p>
      <div class="hosts"><span>USA</span><span>CANADÁ</span><span>MÉXICO</span></div>
    </div>
  </header>

  <nav class="tabs" id="tabs"></nav>

  <main id="app"></main>

  <footer class="footer" id="footer"></footer>

  <script>
    const DATA_URL = '__DATA_URL__';
    let DATA = null;
    let SLUG2NAME = {};

    function rebuildSlugMap() {
      SLUG2NAME = {};
      if (DATA && DATA.players_index) {
        DATA.players_index.forEach(p => { SLUG2NAME[p.slug] = p.name; });
      }
    }

    function updateFooter() {
      if (!DATA || !DATA.meta) return;
      document.getElementById('footer').innerHTML =
        `Actualizado: ${esc(DATA.meta.generated_at)} · Puntuación configurable en el Excel (hoja Puntuacion).`;
    }

    async function loadData(isPoll) {
      const prevAt = DATA && DATA.meta ? DATA.meta.generated_at : null;
      try {
        const res = await fetch(DATA_URL + '?t=' + Date.now());
        if (!res.ok) throw new Error('HTTP ' + res.status);
        const next = await res.json();
        const changed = !prevAt || prevAt !== next.meta.generated_at;
        DATA = next;
        rebuildSlugMap();
        updateFooter();
        if (!isPoll || changed) router();
      } catch (e) {
        if (!DATA) {
          document.getElementById('app').innerHTML =
            '<p class="muted">No se pudieron cargar los datos. Comprueba la conexión y recarga.</p>';
        }
      }
    }

    let charts = [];
    function destroyCharts() { charts.forEach(c => c.destroy()); charts = []; }

    function esc(s) {
      return String(s == null ? '' : s)
        .replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;');
    }
    const TIER_COLORS = {
      exact: '#15803d', difference: '#0e7490', winner: '#b45309',
      clasificado: '#6d28d9', miss: '#7f1d1d', pending: '#334155'
    };
    const TIER_LABELS = {
      exact: 'Resultado exacto', difference: 'Diferencia de goles', winner: 'Solo el ganador',
      clasificado: 'Solo el clasificado', miss: 'Sin acierto', pending: 'Pendiente'
    };

    function playerInitials(name) {
      const p = DATA.players_index.find(x => x.name === name);
      return p ? p.initials : name.slice(0, 1);
    }

    function initBadgeHtml(pl, tier) {
      const pts = pl.points != null ? ` · +${pl.points} pts` : '';
      const tip = `${pl.name}: ${pl.pred || '—'}${pts}`;
      return `<span class="init-badge ${tier}" title="${esc(tip)}">${esc(pl.initials || playerInitials(pl.name))}</span>`;
    }

    function tierLegendMiniHtml() {
      return ['exact', 'difference', 'winner', 'clasificado', 'miss'].map(t =>
        `<span class="${t}"><i></i>${TIER_LABELS[t]}</span>`
      ).join('');
    }

    function getMatchPredictions(matchId) {
      return DATA.players_index.map(p => {
        const m = DATA.players[p.name].matches.find(x => x.id === matchId);
        return {
          name: p.name,
          initials: p.initials,
          pred_h: m ? m.pred_h : null,
          pred_a: m ? m.pred_a : null,
          pred_clasificado: m ? m.pred_clasificado : '',
          tier: m ? m.tier : 'pending',
          tier_label: m ? m.tier_label : TIER_LABELS.pending,
          points: m ? m.points : null,
        };
      });
    }

    function closeMatchModal() {
      const el = document.getElementById('matchModal');
      if (el) el.remove();
      document.removeEventListener('keydown', onModalEscape);
    }

    function onModalEscape(e) {
      if (e.key === 'Escape') closeMatchModal();
    }

    function openMatchModal(matchId) {
      closeMatchModal();
      const match = DATA.matches.find(m => m.id === matchId);
      if (!match) return;
      const preds = getMatchPredictions(matchId);
      const finished = match.finished;
      const clasifLine = match.clasificado
        ? `<div class="modal-meta">Clasificado: ${esc(match.clasificado)}</div>` : '';
      const ptsHead = finished ? '<th>Pts</th>' : '';
      const rows = preds.map(pl => {
        const pred = (pl.pred_h != null && pl.pred_a != null) ? `${pl.pred_h}-${pl.pred_a}` : '—';
        let clasifCol = '';
        if (pl.pred_clasificado) {
          clasifCol = `<br><small class="muted">Clasif.: ${esc(pl.pred_clasificado)}</small>`;
        }
        const ptsCol = finished
          ? `<td>${pl.points != null ? `<span class="pts-badge">+${pl.points}</span>` : '<span class="muted">—</span>'}</td>`
          : '';
        return `<tr>
          <td><strong>${esc(pl.name)}</strong></td>
          <td>${esc(pred)}${clasifCol}</td>
          ${ptsCol}
          <td><span class="tier-tag ${pl.tier}">${esc(pl.tier_label)}</span></td>
        </tr>`;
      }).join('');

      const overlay = document.createElement('div');
      overlay.id = 'matchModal';
      overlay.className = 'modal-overlay';
      overlay.innerHTML = `
        <div class="modal-box" role="dialog" aria-modal="true">
          <button type="button" class="modal-close" aria-label="Cerrar">&times;</button>
          <div class="modal-head">
            <h3>${esc(match.local)} vs ${esc(match.visitante)}</h3>
            <div class="modal-meta">${esc(match.fecha)} · ${esc(match.fase || '')}</div>
            ${finished ? `<div class="modal-score">${esc(match.score)}</div>` : '<div class="modal-meta">Partido pendiente</div>'}
            ${clasifLine}
          </div>
          <table class="modal-table">
            <thead><tr><th>Jugador</th><th>Pronóstico</th>${ptsHead}<th>Acierto</th></tr></thead>
            <tbody>${rows}</tbody>
          </table>
        </div>`;
      overlay.addEventListener('click', e => { if (e.target === overlay) closeMatchModal(); });
      overlay.querySelector('.modal-close').addEventListener('click', closeMatchModal);
      overlay.querySelector('.modal-box').addEventListener('click', e => e.stopPropagation());
      document.body.appendChild(overlay);
      document.addEventListener('keydown', onModalEscape);
    }

    /* ---------------- Navegación ---------------- */
    function renderTabs() {
      const route = location.hash || '#/';
      const isHome = route === '#/' || route === '#' || route === '';
      const isMatches = route.startsWith('#/partidos');
      const players = DATA.players_index;
      const menu = players.map(p =>
        `<a href="#/jugador/${p.slug}"><span>${esc(p.name)}</span><span class="rk">#${p.rank}</span></a>`
      ).join('');
      document.getElementById('tabs').innerHTML = `
        <a class="tab ${isHome ? 'active' : ''}" href="#/">Inicio</a>
        <a class="tab ${isMatches ? 'active' : ''}" href="#/partidos">Partidos</a>
        <div class="dropdown" id="dd">
          <button id="ddBtn">Jugadores ▾</button>
          <div class="dropdown-menu">${menu}</div>
        </div>`;
      const dd = document.getElementById('dd');
      document.getElementById('ddBtn').addEventListener('click', e => {
        e.stopPropagation();
        dd.classList.toggle('open');
      });
    }
    document.addEventListener('click', () => {
      const dd = document.getElementById('dd');
      if (dd) dd.classList.remove('open');
    });

    /* ---------------- Helpers de tabla ---------------- */
    function rankingTable(rows, meName, cols) {
      const head = ['#', 'Jugador'].concat(cols.map(c => c.label));
      const body = rows.map((p, i) => {
        const pos = i + 1;
        const cls = pos === 1 ? 'medal-1' : pos === 2 ? 'medal-2' : pos === 3 ? 'medal-3' : '';
        const me = p.name === meName ? 'me' : '';
        const slug = DATA.players[p.name].slug;
        const tds = cols.map(c => `<td>${c.fmt(p)}</td>`).join('');
        return `<tr class="${me}">
          <td class="pos">${pos}</td>
          <td class="${cls}"><a href="#/jugador/${slug}">${esc(p.name)}</a></td>
          ${tds}
        </tr>`;
      }).join('');
      return `<table><thead><tr>${head.map(h => `<th>${h}</th>`).join('')}</tr></thead><tbody>${body}</tbody></table>`;
    }

    /* ---------------- Vista: Inicio ---------------- */
    function renderHome() {
      const h = DATA.home, m = DATA.meta;
      const kpis = [
        { label: 'Líder', value: h.leader.name },
        { label: 'Puntos líder', value: h.leader.points },
        { label: 'Partidos jugados', value: m.finished_count + ' / ' + m.total_matches },
        { label: 'Progreso', value: m.progress_pct + '%' },
        { label: 'Media pts/partido', value: m.avg_pts },
      ];
      const kpiHtml = kpis.map(k =>
        `<div class="kpi"><div class="kpi-label">${esc(k.label)}</div><div class="kpi-value">${esc(k.value)}</div></div>`
      ).join('');

      const lastHtml = h.last_matches.length
        ? h.last_matches.map(x => {
            const badges = (x.players || []).map(pl =>
              initBadgeHtml(pl, pl.tier || 'pending')
            ).join('');
            return `
            <div class="match-row">
              <div class="match-row-main">
                <div><strong>${esc(x.local)}</strong> vs ${esc(x.visitante)}<br><small class="muted">${esc(x.fecha)}</small></div>
                <div class="last-players">${badges}</div>
              </div>
              <div class="match-score">${esc(x.score)}</div>
            </div>`;
          }).join('') + `<div class="tier-legend-mini">${tierLegendMiniHtml()}</div>`
        : '<p class="muted">Aún no hay partidos finalizados.</p>';

      const upcomingHtml = h.upcoming_matches && h.upcoming_matches.length
        ? `<div class="upcoming-list">${h.upcoming_matches.map(u => {
            const preds = u.predictions.map(pl => {
              const muted = pl.pred === '—' ? ' muted' : '';
              const clasif = pl.pred_clasificado
                ? `<small class="muted">${esc(pl.pred_clasificado.slice(0, 3))}</small>` : '';
              return `<div class="pred-mini">
                <span class="ini">${esc(pl.initials)}</span>
                <span class="sc-pred${muted}">${esc(pl.pred)}</span>
                ${clasif}
              </div>`;
            }).join('');
            return `<div class="upcoming-block">
              <div class="upcoming-head">
                <span class="upcoming-teams">${esc(u.local)} <span class="muted">vs</span> ${esc(u.visitante)}</span>
                <span class="upcoming-meta">${esc(u.fecha)} · ${esc(u.fase || '')}</span>
              </div>
              <div class="upcoming-preds">${preds}</div>
            </div>`;
          }).join('')}</div>`
        : '<p class="muted">No hay partidos pendientes.</p>';

      const specHtml = h.specials_summary.map(s => {
        const winners = s.aciertos.length
          ? s.aciertos.map(n => `<span class="chip">${esc(n)}</span>`).join('')
          : '<span class="muted">—</span>';
        return `<div class="spec-card ${s.oficial ? (s.aciertos.length ? 'hit' : 'fail') : ''}">
          <div class="cat">${esc(s.categoria)}</div>
          <div class="pred">${s.oficial ? esc(s.oficial) : '<span class="muted">Sin resultado</span>'}</div>
          <div class="off">Aciertan:</div>
          <div class="chips" style="margin-top:.3rem">${winners}</div>
        </div>`;
      }).join('');

      document.getElementById('app').innerHTML = `
        <p class="sync-note">Datos desde <strong>__EXCEL_NAME__</strong>. Tras editar el Excel ejecuta <code>__BUILD_CMD__</code> y recarga.</p>
        <div class="kpis">${kpiHtml}</div>

        <div class="card">
          <h2>Clasificación</h2>
          <div class="toggle-cuts" id="cuts">
            <button data-cut="points" class="active">Por puntos</button>
            <button data-cut="pct">Por % acierto</button>
            <button data-cut="exact">Por exactos</button>
          </div>
          <div id="rankingBox"></div>
        </div>

        <div class="grid-2">
          <div class="card"><h2>Puntos totales</h2><div class="chart-wrap"><canvas id="chartPoints"></canvas></div></div>
          <div class="card"><h2>Origen de los puntos</h2><div class="chart-wrap"><canvas id="chartSource"></canvas></div></div>
        </div>

        <div class="grid-2">
          <div class="card"><h2>__EVO_TITLE__</h2><div class="chart-wrap"><canvas id="chartEvo"></canvas></div></div>
          <div class="card"><h2>Últimos resultados</h2>${lastHtml}</div>
        </div>

        <div class="card"><h2>Próximos partidos</h2>${upcomingHtml}</div>

        <div class="card"><h2>Composición de aciertos</h2><div class="chart-wrap tall"><canvas id="chartBreakdown"></canvas></div></div>

        <div class="card"><h2>Apuestas especiales</h2><div class="spec-grid">${specHtml}</div></div>
      `;

      // Cortes de clasificación
      const cutDefs = {
        points: { rows: h.ranking, cols: [
          { label: 'Pts', fmt: p => `<strong>${p.points}</strong>` },
          { label: 'Exactos', fmt: p => p.exact },
          { label: 'Parciales', fmt: p => p.partial },
          { label: '% acierto', fmt: p => p.pct + '%' },
        ]},
        pct: { rows: h.ranking_by_pct, cols: [
          { label: '% acierto', fmt: p => `<strong>${p.pct}%</strong>` },
          { label: 'Exactos', fmt: p => p.exact },
          { label: 'Parciales', fmt: p => p.partial },
          { label: 'Pts', fmt: p => p.points },
        ]},
        exact: { rows: h.ranking_by_exact, cols: [
          { label: 'Exactos', fmt: p => `<strong>${p.exact}</strong>` },
          { label: 'Parciales', fmt: p => p.partial },
          { label: 'Fallos', fmt: p => p.fails },
          { label: 'Pts', fmt: p => p.points },
        ]},
      };
      function drawCut(key) {
        const d = cutDefs[key];
        document.getElementById('rankingBox').innerHTML = rankingTable(d.rows, null, d.cols);
      }
      drawCut('points');
      document.querySelectorAll('#cuts button').forEach(b => {
        b.addEventListener('click', () => {
          document.querySelectorAll('#cuts button').forEach(x => x.classList.remove('active'));
          b.classList.add('active');
          drawCut(b.dataset.cut);
        });
      });

      drawHomeCharts();
    }

    function drawHomeCharts() {
      Chart.defaults.color = '#94a3b8';
      Chart.defaults.borderColor = 'rgba(255,255,255,0.08)';
      Chart.defaults.font.family = "'Inter', sans-serif";
      const r = DATA.home.ranking;
      const names = r.map(p => p.name);

      charts.push(new Chart(document.getElementById('chartPoints'), {
        type: 'bar',
        data: { labels: names, datasets: [{ label: 'Puntos', data: r.map(p => p.points),
          backgroundColor: 'rgba(0,102,204,0.75)', borderColor: '#0066CC', borderWidth: 1, borderRadius: 4 }] },
        options: { indexAxis: 'y', responsive: true, maintainAspectRatio: false,
          plugins: { legend: { display: false } },
          scales: { x: { grid: { display: false } }, y: { grid: { display: false }, ticks: { color: '#f5f5f5', font: { size: 11 } } } } }
      }));

      charts.push(new Chart(document.getElementById('chartSource'), {
        type: 'bar',
        data: { labels: names, datasets: [
          { label: 'Partidos', data: r.map(p => p.points_match), backgroundColor: '#0066CC' },
          { label: 'Bonus ronda', data: r.map(p => p.points_bonus), backgroundColor: '#C5A028' },
          { label: 'Apuestas', data: r.map(p => p.points_special), backgroundColor: '#6d28d9' },
        ]},
        options: { indexAxis: 'y', responsive: true, maintainAspectRatio: false,
          scales: { x: { stacked: true, grid: { display: false } }, y: { stacked: true, grid: { display: false }, ticks: { color: '#f5f5f5' } } },
          plugins: { legend: { position: 'bottom' } } }
      }));

      const evoNames = Object.keys(DATA.home.evolution);
      const evoColors = ['#C5A028', '#0066CC', '#2E7D32'];
      charts.push(new Chart(document.getElementById('chartEvo'), {
        type: 'line',
        data: { labels: DATA.home.evolution_labels.map(n => 'P' + n),
          datasets: evoNames.map((name, i) => ({ label: name, data: DATA.home.evolution[name],
            borderColor: evoColors[i], backgroundColor: evoColors[i] + '33', tension: 0.25, fill: false, pointRadius: 2 })) },
        options: { responsive: true, maintainAspectRatio: false,
          plugins: { legend: { position: 'bottom' } },
          scales: { x: { grid: { display: false } }, y: { grid: { display: false }, beginAtZero: true } } }
      }));

      charts.push(new Chart(document.getElementById('chartBreakdown'), {
        type: 'bar',
        data: { labels: names, datasets: [
          { label: 'Exactos', data: r.map(p => p.exact), backgroundColor: '#15803d' },
          { label: 'Parciales', data: r.map(p => p.partial), backgroundColor: '#0e7490' },
          { label: 'Fallos', data: r.map(p => p.fails), backgroundColor: '#546E7A' },
        ]},
        options: { indexAxis: 'y', responsive: true, maintainAspectRatio: false,
          scales: { x: { stacked: true, grid: { display: false } }, y: { stacked: true, grid: { display: false }, ticks: { color: '#f5f5f5' } } },
          plugins: { legend: { position: 'bottom' } } }
      }));
    }

    /* ---------------- Vista: Partidos ---------------- */
    function renderPartidos() {
      const fases = [...new Set(DATA.matches.map(m => m.fase).filter(Boolean))];
      const faseOpts = ['<option value="">Todas las fases</option>']
        .concat(fases.map(f => `<option value="${esc(f)}">${esc(f)}</option>`)).join('');
      document.getElementById('app').innerHTML = `
        <h2 class="section-title">Resultados del Mundial</h2>
        <div class="filters">
          <select id="fFase">${faseOpts}</select>
          <select id="fEstado">
            <option value="all">Todos</option>
            <option value="fin">Jugados</option>
            <option value="pend">Pendientes</option>
          </select>
          <input id="fTeam" type="search" placeholder="Buscar equipo..." />
        </div>
        <div id="matchList"></div>`;

      function draw() {
        const fFase = document.getElementById('fFase').value;
        const fEstado = document.getElementById('fEstado').value;
        const fTeam = document.getElementById('fTeam').value.trim().toLowerCase();
        let rows = DATA.matches.filter(m => {
          if (fFase && m.fase !== fFase) return false;
          if (fEstado === 'fin' && !m.finished) return false;
          if (fEstado === 'pend' && m.finished) return false;
          if (fTeam && !(m.local.toLowerCase().includes(fTeam) || m.visitante.toLowerCase().includes(fTeam))) return false;
          return true;
        });
        if (!rows.length) { document.getElementById('matchList').innerHTML = '<p class="muted">Sin partidos para este filtro.</p>'; return; }
        const groups = {};
        rows.forEach(m => { (groups[m.fase || 'Otros'] = groups[m.fase || 'Otros'] || []).push(m); });
        document.getElementById('matchList').innerHTML = Object.keys(groups).map(fase => {
          const items = groups[fase].map(m => {
            const estado = m.finished ? '<span class="pill fin">Jugado</span>' : '<span class="pill pend">Pendiente</span>';
            const clasif = m.clasificado ? `<span class="clasif">${esc(m.clasificado)}</span>` : '<span></span>';
            return `<div class="mt-row clickable" data-match-id="${m.id}" role="button" tabindex="0">
              <div class="date">${esc(m.fecha)}<br>${estado}</div>
              <div class="teams">${esc(m.local)} <span class="muted">vs</span> ${esc(m.visitante)}</div>
              <div class="sc">${esc(m.score)}</div>
              ${clasif}
            </div>`;
          }).join('');
          return `<div class="phase-group"><div class="phase-head">${esc(fase)}</div>${items}</div>`;
        }).join('');
        document.getElementById('matchList').querySelectorAll('.mt-row.clickable').forEach(row => {
          const mid = parseInt(row.dataset.matchId, 10);
          row.addEventListener('click', () => openMatchModal(mid));
          row.addEventListener('keydown', e => {
            if (e.key === 'Enter' || e.key === ' ') { e.preventDefault(); openMatchModal(mid); }
          });
        });
      }
      draw();
      ['fFase', 'fEstado'].forEach(id => document.getElementById(id).addEventListener('change', draw));
      document.getElementById('fTeam').addEventListener('input', draw);
    }

    /* ---------------- Vista: Jugador ---------------- */
    function renderJugador(slug) {
      const name = SLUG2NAME[slug];
      const p = name ? DATA.players[name] : null;
      if (!p) {
        document.getElementById('app').innerHTML = '<p class="muted">Jugador no encontrado. <a href="#/">Volver al inicio</a>.</p>';
        return;
      }
      const s = p.stats;
      const legend = ['exact', 'difference', 'winner', 'clasificado', 'miss', 'pending'].map(t =>
        `<span><i style="background:${TIER_COLORS[t]}"></i>${TIER_LABELS[t]}</span>`).join('');

      const specHtml = p.specials.map(sp => {
        const state = sp.oficial ? (sp.hit ? 'hit' : 'fail') : 'pend';
        const stTxt = sp.oficial ? (sp.hit ? '✓ Acierto (+' + sp.points + ')' : '✗ Fallo') : 'Pendiente';
        return `<div class="spec-card ${sp.oficial ? (sp.hit ? 'hit' : 'fail') : ''}">
          <div class="cat">${esc(sp.categoria)}</div>
          <div class="pred">${sp.prediccion ? esc(sp.prediccion) : '<span class="muted">Sin apuesta</span>'}</div>
          <div class="off">Oficial: ${sp.oficial ? esc(sp.oficial) : '—'}</div>
          <div class="spec-status ${state}">${stTxt}</div>
        </div>`;
      }).join('');

      // Pronósticos agrupados por fase
      const groups = {};
      p.matches.forEach(m => { (groups[m.fase || 'Otros'] = groups[m.fase || 'Otros'] || []).push(m); });
      const matchHtml = Object.keys(groups).map(fase => {
        const cards = groups[fase].map(m => {
          const pred = (m.pred_h != null && m.pred_a != null) ? `${m.pred_h}-${m.pred_a}` : '—';
          const real = (m.real_h != null && m.real_a != null) ? `${m.real_h}-${m.real_a}` : '—';
          const pts = m.points != null ? `<span class="pts-badge">+${m.points}</span>` : '<span class="muted">—</span>';
          let clasifLine = '';
          if (m.pred_clasificado || m.real_clasificado) {
            clasifLine = `<div class="clasif-line">Clasificado — tú: ${esc(m.pred_clasificado || '—')} · real: ${esc(m.real_clasificado || '—')}</div>`;
          }
          return `<div class="pred-card ${m.tier}">
            <div class="pc-teams">${esc(m.local)} <span class="muted">vs</span> ${esc(m.visitante)}</div>
            <div class="pc-date">${esc(m.fecha)}</div>
            <div class="pred-line"><span class="lbl">Tu pronóstico</span><span class="score-chip">${pred}</span></div>
            <div class="pred-line"><span class="lbl">Resultado real</span><span class="score-chip">${real}</span></div>
            ${clasifLine}
            <div class="pc-foot"><span class="tier-tag ${m.tier}">${esc(m.tier_label)}</span>${pts}</div>
          </div>`;
        }).join('');
        return `<div class="phase-group"><div class="phase-head">${esc(fase)}</div><div class="pred-grid">${cards}</div></div>`;
      }).join('');

      document.getElementById('app').innerHTML = `
        <div class="player-top">
          ${p.photo ? `<img class="player-photo" src="${p.photo}" alt="${esc(p.name)}">` : ''}
          <div class="player-head">
            <span class="pname">${esc(p.name)}</span>
            <span class="prank">#${p.rank}</span>
          </div>
        </div>
        <div class="card">
          <div class="breakdown">
            <div><span class="bk-val">${s.points}</span><span class="bk-lbl">Puntos totales</span></div>
            <div><span class="bk-val">${s.points_match}</span><span class="bk-lbl">Partidos</span></div>
            <div><span class="bk-val">${s.points_bonus}</span><span class="bk-lbl">Bonus ronda</span></div>
            <div><span class="bk-val">${s.points_special}</span><span class="bk-lbl">Apuestas</span></div>
            <div><span class="bk-val">${s.exact}</span><span class="bk-lbl">Exactos</span></div>
            <div><span class="bk-val">${s.pct}%</span><span class="bk-lbl">% acierto</span></div>
          </div>
        </div>

        <div class="card">
          <h2>Apuestas especiales</h2>
          <div class="spec-grid">${specHtml}</div>
        </div>

        <h2 class="section-title">Pronósticos por partido</h2>
        <div class="legend">${legend}</div>
        ${matchHtml}
      `;
    }

    /* ---------------- Router ---------------- */
    function router() {
      if (!DATA) return;
      destroyCharts();
      closeMatchModal();
      renderTabs();
      window.scrollTo(0, 0);
      const route = location.hash || '#/';
      if (route.startsWith('#/partidos')) {
        renderPartidos();
      } else if (route.startsWith('#/jugador/')) {
        renderJugador(decodeURIComponent(route.split('/')[2] || ''));
      } else {
        renderHome();
      }
    }

    window.addEventListener('hashchange', router);
    loadData(false).then(() => {
      if (DATA) setInterval(() => loadData(true), 5 * 60 * 1000);
    });
  </script>
</body>
</html>
"""


def render_html(data: dict, group: dict) -> str:
    assets_dir = group["assets_dir"]
    data_url = f"{assets_dir}/data.json"
    assets = data.get("assets", {})
    hero_path = assets.get("portada", "")
    portada_aspect = assets.get("portada_aspect")
    meta = data.get("meta", {})
    evo_top = meta.get("evo_top", 3)
    if hero_path and portada_aspect:
        hero_css = f"url('{hero_path}') center center / 100% 100% no-repeat"
        hero_aspect = str(portada_aspect)
        hero_min_height = "0"
    elif hero_path:
        hero_css = f"url('{hero_path}') center center / contain no-repeat"
        hero_aspect = "auto"
        hero_min_height = "380px"
    else:
        hero_css = "linear-gradient(135deg,#051C2C 0%,#1E3A5F 50%,#0066CC 100%)"
        hero_aspect = "auto"
        hero_min_height = "380px"
    group_title = f"PORRA MUNDIAL 2026 — {group['label'].upper()}"
    return (
        HTML_TEMPLATE.replace("__HERO_CSS__", hero_css)
        .replace("__HERO_ASPECT__", hero_aspect)
        .replace("__HERO_MIN_HEIGHT__", hero_min_height)
        .replace("__GROUP_LABEL__", group["label"])
        .replace("__GROUP_TITLE__", group_title)
        .replace("__EXCEL_NAME__", group["excel"])
        .replace("__BUILD_CMD__", f"py build_dashboard.py --group {group['id']}")
        .replace("__EVO_TITLE__", f"Evolución — Top {evo_top}")
        .replace("__DATA_URL__", data_url)
    )


def main(group_id: str = "broshu", open_browser: bool = True) -> None:
    group = load_group(group_id)
    excel_path = group["excel_path"]
    html_out = group["dashboard_path"]

    if not excel_path.exists():
        raise SystemExit(
            f"No existe {excel_path}. Ejecuta primero: py generate_porra.py --group {group_id}"
        )

    players = group["players"]
    matches = load_json(MATCHES_JSON)
    n_matches = len(matches)

    OUTPUT.mkdir(parents=True, exist_ok=True)
    asset_paths = copy_dashboard_assets(players, group)

    raw = read_workbook_data(excel_path, players, n_matches)
    data = build_full_payload(players, raw, n_matches, group, asset_paths)

    data_json_path = group["assets_path"] / "data.json"
    data_json_path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"  Datos: {data_json_path}")

    html_out.write_text(render_html(data, group), encoding="utf-8")
    print(f"Dashboard: {html_out}")

    if open_browser:
        webbrowser.open(html_out.as_uri())


if __name__ == "__main__":
    import sys

    open_browser = "--no-open" not in sys.argv
    for gid in resolve_group_ids(sys.argv):
        main(gid, open_browser=open_browser and gid == resolve_group_ids(sys.argv)[-1])
