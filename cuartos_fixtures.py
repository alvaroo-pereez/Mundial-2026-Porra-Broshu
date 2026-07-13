"""Equipos canónicos de cuartos (cuadro FIFA / openfootball)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from config.groups import list_groups, load_group
from worldcup_data import load_calendar, save_json, MATCHES_JSON

PT_FIRST_ROW = 4
CUARTOS_FIRST_ID = 97
CUARTOS_LAST_ID = 100

# Id calendario → (local, visitante) en español
CUARTOS_CANONICAL: dict[int, tuple[str, str]] = {
    97: ("Francia", "Marruecos"),
    98: ("España", "Bélgica"),
    99: ("Noruega", "Inglaterra"),
    100: ("Argentina", "Suiza"),
}

# Octavos emparejados → id cuartos (cuadro FIFA, no secuencial)
CUARTOS_OCTAVOS_PAIRS: list[tuple[int, int]] = [
    (89, 90),  # 97
    (93, 94),  # 98
    (91, 92),  # 99
    (95, 96),  # 100
]

CUARTOS_LIST_ORDER: list[tuple[int, str, str]] = [
    (mid, *CUARTOS_CANONICAL[mid]) for mid in range(CUARTOS_FIRST_ID, CUARTOS_LAST_ID + 1)
]


def apply_cuartos_to_json(dry_run: bool = False) -> list[int]:
    calendar = load_calendar()
    changed: list[int] = []
    cal_by_id = {m["id"]: m for m in calendar}
    for mid, (local, visitante) in CUARTOS_CANONICAL.items():
        m = cal_by_id.get(mid)
        if not m:
            continue
        if m["local"] == local and m["visitante"] == visitante:
            continue
        print(f"  JSON #{mid}: {m['local']} vs {m['visitante']} -> {local} vs {visitante}")
        m["local"] = local
        m["visitante"] = visitante
        changed.append(mid)
    if changed and not dry_run:
        save_json(MATCHES_JSON, calendar)
    return changed


def apply_cuartos_to_excel(excel_path: Path, dry_run: bool = False) -> list[int]:
    wb = load_workbook(excel_path)
    pt = wb["Partidos"]
    changed: list[int] = []
    for mid, (local, visitante) in CUARTOS_CANONICAL.items():
        row = PT_FIRST_ROW + mid - 1
        old_l, old_v = pt.cell(row, 4).value, pt.cell(row, 5).value
        if old_l == local and old_v == visitante:
            continue
        print(f"  #{mid}: {old_l} vs {old_v} -> {local} vs {visitante}")
        if not dry_run:
            pt.cell(row, 4, value=local)
            pt.cell(row, 5, value=visitante)
        changed.append(mid)
    if changed and not dry_run:
        wb.save(excel_path)
    wb.close()
    return changed


def apply_cuartos_canonical_teams(dry_run: bool = False) -> list[int]:
    """Fija equipos cuartos 97-100 en JSON y todos los Excel maestros."""
    print("Aplicando equipos canónicos cuartos (97-100)...")
    all_changed: set[int] = set()
    print("matches_2026.json:")
    all_changed.update(apply_cuartos_to_json(dry_run))
    for group_id in list_groups():
        excel_path = load_group(group_id)["excel_path"]
        if not excel_path.exists():
            continue
        print(f"Excel {group_id} ({excel_path.name}):")
        all_changed.update(apply_cuartos_to_excel(excel_path, dry_run))
    return sorted(all_changed)
