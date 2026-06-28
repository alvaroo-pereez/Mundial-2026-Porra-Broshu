"""
Genera Porra_Mundial_2026_<Grupo>.xlsx — porra automatizada Mundial 2026.
Fórmulas compatibles con Excel 2010+ (INDEX/MATCH, sin XLOOKUP/LET/SORT/FILTER).
Requiere: Python 3.10+, openpyxl.
"""
from __future__ import annotations

import json
import re
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.groups import load_group, resolve_group_ids
from scoring import load_scoring_config

ROOT = Path(__file__).parent
ASSETS = ROOT / "assets"

# --- Paleta McKinsey-style (consulting deck) ---
NAVY = "051C2C"
NAVY_MID = "1E3A5F"
BLUE_ACCENT = "0066CC"
BLUE_LIGHT = "E8F0FA"
GRAY_DARK = "424242"
GRAY_MID = "9E9E9E"
GRAY_LIGHT = "F5F5F5"
WHITE = "FFFFFF"
GOLD = "C5A028"
GOLD_DARK = "8A6D1B"
SILVER = "ECEFF1"
BRONZE = "F3E5D5"
RED_SOFT = "FFEBEE"
YELLOW = "FFF8E1"
GREEN_PTS = "E3F2FD"
RED_ACCENT = "B71C1C"
BG_DASH = "FAFAFA"
FORMULA_FILL = PatternFill("solid", fgColor="E8F0FA")
REFERENCE_FILL = PatternFill("solid", fgColor="ECEFF1")
# Alias compatibilidad con estilos existentes
GREEN_DARK = NAVY
GREEN_MID = NAVY_MID
GREEN_LIGHT = BLUE_LIGHT
BLUE_HOST = BLUE_ACCENT

# Referencias celdas hoja Puntuacion (editables; filas alineadas con build_puntuacion)
PTS_GR_EXACT = "Puntuacion!$B$4"
PTS_GR_DIFF = "Puntuacion!$B$5"
PTS_GR_WIN = "Puntuacion!$B$6"
PTS_KO_EXACT = "Puntuacion!$B$9"
PTS_KO_DIFF = "Puntuacion!$B$10"
PTS_KO_CLAS = "Puntuacion!$B$11"
PTS_KO_EMP_EX_FAIL = "Puntuacion!$B$12"
PTS_KO_EMP_DIFF_OK = "Puntuacion!$B$13"
PTS_KO_EMP_DIFF_FAIL = "Puntuacion!$B$14"
BONUS_OCT = "Puntuacion!$B$17"
BONUS_CUA = "Puntuacion!$B$18"
BONUS_SEM = "Puntuacion!$B$19"
PTS_APUESTA = "Puntuacion!$B$22"
SPECIAL_RESULTS_FIRST_ROW = 34

TAB_INTRO = GOLD
TAB_CORE = BLUE_ACCENT
TAB_WORK = GRAY_MID

# Demo: partidos 1-5 con resultado; pronósticos variados
SPECIAL_BET_CATEGORIES = [
    "Campeón del Mundial",
    "Subcampeón",
    "Tercer equipo",
    "Balón de oro de la FIFA",
    "Premio de la FIFA al mejor jugador joven",
    "Guante de oro de la FIFA",
    "Bota de oro de la FIFA",
    "Máximo goleador España",
]

DEMO_RESULTS: dict[int, tuple[int, int]] = {
    1: (2, 1),   # México vs Sudáfrica
    2: (1, 1),   # República de Corea vs República Checa
    3: (2, 0),   # Canadá vs Bosnia y Herzegovina
    4: (1, 0),   # Estados Unidos vs Paraguay
    5: (1, 2),   # Catar vs Suiza
}

DEMO_PREDICTIONS_BROSHU: dict[int, dict[str, tuple[int, int]]] = {
    1: {
        "Álvaro": (1, 0),
        "Pepe": (2, 1),
        "Patri": (0, 0),
        "Kike": (2, 1),
        "Quintero": (1, 2),
        "Nacho": (2, 0),
        "Luis": (3, 1),
        "Felipe": (2, 1),
        "Simón": (1, 1),
        "Muni": (0, 1),
        "Fer": (2, 1),
    },
    2: {
        "Álvaro": (2, 0),
        "Pepe": (1, 1),
        "Patri": (0, 1),
        "Kike": (1, 0),
        "Quintero": (1, 1),
        "Nacho": (2, 2),
        "Luis": (0, 0),
        "Felipe": (1, 1),
        "Simón": (3, 3),
        "Muni": (0, 0),
        "Fer": (2, 1),
    },
    3: {
        "Álvaro": (3, 0),
        "Pepe": (2, 1),
        "Patri": (1, 0),
        "Kike": (3, 0),
        "Quintero": (1, 1),
        "Nacho": (2, 0),
        "Luis": (1, 2),
        "Felipe": (3, 1),
        "Simón": (0, 1),
        "Muni": (3, 0),
        "Fer": (2, 2),
    },
    4: {
        "Álvaro": (1, 0),
        "Pepe": (0, 2),
        "Patri": (2, 1),
        "Kike": (0, 1),
        "Quintero": (0, 2),
        "Nacho": (1, 1),
        "Luis": (0, 3),
        "Felipe": (0, 2),
        "Simón": (2, 0),
        "Muni": (1, 0),
        "Fer": (0, 2),
    },
    5: {
        "Álvaro": (2, 1),
        "Pepe": (1, 3),
        "Patri": (0, 1),
        "Kike": (1, 2),
        "Quintero": (1, 3),
        "Nacho": (3, 0),
        "Luis": (1, 1),
        "Felipe": (2, 2),
        "Simón": (0, 2),
        "Muni": (1, 4),
        "Fer": (1, 3),
    },
}

DEMO_PREDICTIONS_PAPINENES: dict[int, dict[str, tuple[int, int]]] = {
    1: {"Álvaro": (1, 0), "Papá": (2, 0), "Diego": (0, 1)},
    2: {"Álvaro": (2, 0), "Papá": (1, 2), "Diego": (1, 1)},
    3: {"Álvaro": (3, 0), "Papá": (2, 1), "Diego": (1, 0)},
    4: {"Álvaro": (1, 0), "Papá": (2, 1), "Diego": (0, 2)},
    5: {"Álvaro": (2, 1), "Papá": (1, 1), "Diego": (2, 2)},
}

DEMO_PREDICTIONS_BY_GROUP: dict[str, dict[int, dict[str, tuple[int, int]]]] = {
    "broshu": DEMO_PREDICTIONS_BROSHU,
    "papinenes": DEMO_PREDICTIONS_PAPINENES,
}


def player_range(players: list[str]) -> tuple[int, int]:
    """Filas de estadísticas por jugador en _Helpers (cols A:J, fila 2+)."""
    first = 2
    last = first + len(players) - 1
    return first, last


def pron_last_row(n_matches: int, n_players: int) -> int:
    return PT_FIRST_ROW + n_matches * n_players - 1


def match_pts_sum_formula(player_cell: str, pron_last: int) -> str:
    """Suma puntos por jugador (SUMIFS simple, sin criterio sobre G que provoca ciclos)."""
    return (
        f"=SUMIFS(Pronosticos!$G${PT_FIRST_ROW}:$G${pron_last},"
        f"Pronosticos!$D${PT_FIRST_ROW}:$D${pron_last},{player_cell})"
    )


def count_pts_formula(player_cell: str, pts_ref: str, pron_last: int) -> str:
    return (
        f"COUNTIFS(Pronosticos!$D${PT_FIRST_ROW}:$D${pron_last},{player_cell},"
        f"Pronosticos!$G${PT_FIRST_ROW}:$G${pron_last},{pts_ref})"
    )


PT_FIRST_ROW = 4  # filas 1-2 = banner (título + descripción), fila 3 = cabecera
HEADER_ROW = PT_FIRST_ROW - 1  # cabecera de columnas en hojas de datos


def pt_last_row(n_matches: int) -> int:
    return PT_FIRST_ROW + n_matches - 1


def pt_match_idx(match_cell: str, last_row: int) -> str:
    return f"MATCH({match_cell},Partidos!$A${PT_FIRST_ROW}:$A${last_row},0)"


def pt_value(col: str, match_cell: str, last_row: int) -> str:
    """Valor de Partidos columna col (F,G,H,D,E...) para el nº partido en match_cell."""
    idx = pt_match_idx(match_cell, last_row)
    return f"INDEX(Partidos!${col}${PT_FIRST_ROW}:${col}${last_row},{idx})"


def idx_match(lookup: str, keys: str, values: str) -> str:
    return f"INDEX({values},MATCH({lookup},{keys},0))"


def load_json(path: Path) -> list | dict:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def thin_border() -> Border:
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)


def header_font() -> Font:
    return Font(name="Calibri", size=11, bold=True, color=WHITE)


def header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=NAVY)


def section_title(ws, cell: str, text: str, merge_to: str | None = None) -> None:
    if merge_to:
        row = ws[cell].row
        for col in range(ws[cell].column, ws[merge_to].column + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=GRAY_LIGHT)
    c = ws[cell]
    c.value = text
    c.font = Font(name="Calibri", size=12, bold=True, color=NAVY)
    c.fill = PatternFill("solid", fgColor=GRAY_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = Border(left=Side(style="thick", color=BLUE_ACCENT))


def apply_header_row(ws, headers: list[str], row: int = 1) -> None:
    for col, title in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=title)
        c.font = header_font()
        c.fill = header_fill()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()


def center_across(ws, row: int, start_col: int, end_col: int, value: str, font: Font, fill_color: str) -> None:
    """Visual span without merged cells: Excel-friendly Center Across Selection."""
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.alignment = Alignment(horizontal="centerContinuous", vertical="center", wrap_text=True)
    c = ws.cell(row=row, column=start_col, value=value)
    c.font = font


def style_data_range(ws, min_row: int, max_row: int, max_col: int) -> None:
    alt = PatternFill("solid", fgColor=GRAY_LIGHT)
    for r in range(min_row, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name="Calibri", size=11, color=GRAY_DARK)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if r % 2 == 0 and cell.fill.fill_type is None:
                cell.fill = alt


def style_block(ws, r1: int, r2: int, c1: int, c2: int) -> None:
    """Aplica fuente y bordes a un rectángulo concreto (sin tocar el resto)."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name="Calibri", size=11, color=GRAY_DARK)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def match_is_ko(m: dict) -> bool:
    return m.get("fase", "Grupos") != "Grupos"


def ko_points_formula(row: int, last_row: int) -> str:
    """Puntos eliminatorias con reglas de empate a 90'."""
    r = row
    rh = pt_value("F", f"$A{r}", last_row)
    rv = pt_value("G", f"$A{r}", last_row)
    rk = pt_value("K", f"$A{r}", last_row)
    real_draw = f"{rh}={rv}"
    pred_exact = f"AND($E{r}={rh},$F{r}={rv})"
    pred_draw = f"$E{r}=$F{r}"
    diff = f"($E{r}-$F{r})=({rh}-{rv})"
    clas_ok = f'AND($H{r}<>"",$H{r}={rk})'
    ko_no_draw = (
        f"IF({pred_exact},{PTS_KO_EXACT},"
        f"IF({diff},{PTS_KO_DIFF},"
        f"IF({clas_ok},{PTS_KO_CLAS},0)))"
    )
    ko_draw = (
        f"IF({pred_exact},IF({clas_ok},{PTS_KO_EXACT},{PTS_KO_EMP_EX_FAIL}),"
        f"IF(AND({pred_draw},{diff}),IF({clas_ok},{PTS_KO_EMP_DIFF_OK},{PTS_KO_EMP_DIFF_FAIL}),"
        f"IF({clas_ok},{PTS_KO_CLAS},0)))"
    )
    return f"IF({real_draw},{ko_draw},{ko_no_draw})"


def points_formula(row: int, last_row: int) -> str:
    """Puntuación no acumulativa: solo el tier más alto (grupos vs eliminatorias)."""
    r = row
    rh = pt_value("F", f"$A{r}", last_row)
    rv = pt_value("G", f"$A{r}", last_row)
    rk = pt_value("K", f"$A{r}", last_row)
    st = pt_value("H", f"$A{r}", last_row)
    fase = pt_value("J", f"$A{r}", last_row)
    win = (
        f"OR(AND($E{r}>$F{r},{rh}>{rv}),"
        f"AND($E{r}=$F{r},{rh}={rv}),"
        f"AND($E{r}<$F{r},{rh}<{rv}))"
    )
    grupos = (
        f"IF(AND($E{r}={rh},$F{r}={rv}),{PTS_GR_EXACT},"
        f"IF(($E{r}-$F{r})=({rh}-{rv}),{PTS_GR_DIFF},"
        f"IF({win},{PTS_GR_WIN},0)))"
    )
    ko = ko_points_formula(r, last_row)
    return (
        f"=IFERROR(IF({st}<>\"Finalizado\",\"\","
        f"IF(OR(NOT(ISNUMBER($E{r})),NOT(ISNUMBER($F{r}))),\"\","
        f"IF({fase}=\"Grupos\",{grupos},{ko}))),\"\")"
    )


def pronosticos_fase_formula(row: int, last_row: int) -> str:
    return f"=IFERROR({pt_value('J', f'$A{row}', last_row)},\"\")"


def pronosticos_clasif_ok_formula(row: int, last_row: int) -> str:
    r = row
    rk = pt_value("K", f"$A{r}", last_row)
    st = pt_value("H", f"$A{r}", last_row)
    fase = pt_value("J", f"$A{r}", last_row)
    return (
        f"=IF(AND({fase}<>\"Grupos\",{st}=\"Finalizado\","
        f"$H{r}<>\"\",$H{r}={rk}),1,0)"
    )


def round_bonus_formula(
    player_cell: str,
    fase_name: str,
    n_expected: int,
    bonus_cell: str,
    last_row: int,
    pron_last: int,
) -> str:
    fr = PT_FIRST_ROW
    lr = last_row
    fin = (
        f'COUNTIFS(Partidos!$J${fr}:$J${lr},"{fase_name}",'
        f'Partidos!$H${fr}:$H${lr},"Finalizado")'
    )
    correct = (
        f"SUMIFS(Pronosticos!$J${PT_FIRST_ROW}:$J${pron_last},Pronosticos!$D${PT_FIRST_ROW}:$D${pron_last},"
        f"{player_cell},Pronosticos!$I${PT_FIRST_ROW}:$I${pron_last},\"{fase_name}\")"
    )
    return f"=IF(AND({fin}={n_expected},{fin}>0,{correct}={n_expected}),{bonus_cell},0)"


def apuestas_points_formula(player: str, last_row: int) -> str:
    sheet = excel_sheet_name(player)
    first = special_bets_first_row(last_row)
    parts = [
        (
            f"IF(AND(Resumen!$B${SPECIAL_RESULTS_FIRST_ROW + i}<>\"\","
            f"'{sheet}'!$B${first + i}<>\"\","
            f"'{sheet}'!$B${first + i}=Resumen!$B${SPECIAL_RESULTS_FIRST_ROW + i}),"
            f"{PTS_APUESTA},0)"
        )
        for i in range(len(SPECIAL_BET_CATEGORIES))
    ]
    return "=" + "+".join(parts)


def special_bets_first_row(last_row: int) -> int:
    return last_row + 4


def exactos_count_formula(player_cell: str, pron_last: int) -> str:
    return (
        f"={count_pts_formula(player_cell, PTS_GR_EXACT, pron_last)}"
        f"+{count_pts_formula(player_cell, PTS_KO_EXACT, pron_last)}"
    )


def parciales_count_formula(player_cell: str, pron_last: int) -> str:
    tiers = [PTS_GR_DIFF, PTS_GR_WIN, PTS_KO_DIFF, PTS_KO_CLAS]
    return "=" + "+".join(
        count_pts_formula(player_cell, t, pron_last) for t in tiers
    )


def team_lookup_formula(row: int, team_col: str, last_row: int) -> str:
    return (
        f"=IFERROR({pt_value(team_col, f'$A{row}', last_row)},\"\")"
    )


def sort_key_formula(row: int) -> str:
    """Clave única de orden: puntos totales > exactos > menos fallos (cols E,F,H)."""
    return f"=E{row}*1000000+F{row}*1000+(100-H{row})"


def ranked_stat_formula(
    pos: int, p_first: int, p_last: int, src_col: str, order_col: str = "J"
) -> str:
    """Ranking pos (1=líder) usando columna de orden única."""
    order = f"${order_col}${p_first}:${order_col}${p_last}"
    src = f"${src_col}${p_first}:${src_col}${p_last}"
    return f"=IFERROR(INDEX({src},MATCH(LARGE({order},{pos}),{order},0)),\"\")"


def medal_formula(row: int) -> str:
    """Medallas en texto (compatible con todos los Excel; sin emojis)."""
    return f'=IF(A{row}=1,"ORO",IF(A{row}=2,"PLATA",IF(A{row}=3,"BRONCE","")))'


def set_tab_color(ws, color: str) -> None:
    ws.sheet_properties.tabColor = color


def gold_fill() -> PatternFill:
    return PatternFill("solid", fgColor=GOLD)


def banner_fill() -> PatternFill:
    return PatternFill("solid", fgColor=GREEN_DARK)


def dash_fill() -> PatternFill:
    return PatternFill("solid", fgColor=BG_DASH)


def apply_banner(ws, cell_range: str, title: str, subtitle: str = "", two_row: bool = True) -> None:
    """Cabecera legible sin merge: el título (fila 1) y la descripción (fila 2)
    se escriben en la primera celda y se desbordan sobre celdas vacías con fondo
    de color. Así el texto se lee completo sin combinar celdas."""
    start, end = cell_range.split(":")
    start_col = ws[start].column
    end_col = ws[end].column
    row = ws[start].row

    for col in range(start_col, end_col + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=NAVY)
    title_cell = ws.cell(row=row, column=start_col, value=title)
    title_cell.font = Font(name="Calibri", size=18, bold=True, color=WHITE)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 30

    if two_row:
        srow = row + 1
        for col in range(start_col, end_col + 1):
            ws.cell(row=srow, column=col).fill = PatternFill("solid", fgColor=NAVY_MID)
        if subtitle:
            sub = ws.cell(row=srow, column=start_col, value=subtitle)
            sub.font = Font(name="Calibri", size=10, italic=True, color=WHITE)
            sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[srow].height = 18
    elif subtitle and start_col + 1 <= end_col:
        sub = ws.cell(row=row, column=start_col + 1, value=subtitle)
        sub.font = Font(name="Calibri", size=9, italic=True, color=WHITE)
        sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def style_kpi(ws, label_cell: str, value_cell: str, label: str, value_formula: str) -> None:
    """Tarjeta KPI estilo consulting: etiqueta pequeña, número grande."""
    lc = ws[label_cell]
    lc.value = label.upper()
    lc.font = Font(name="Calibri", size=9, bold=False, color=GRAY_MID)
    lc.fill = PatternFill("solid", fgColor=WHITE)
    lc.alignment = Alignment(horizontal="left", vertical="bottom", indent=1)
    vc = ws[value_cell]
    vc.value = value_formula
    vc.font = Font(name="Calibri", size=20, bold=True, color=NAVY)
    vc.fill = PatternFill("solid", fgColor=WHITE)
    vc.alignment = Alignment(horizontal="left", vertical="top", indent=1)
    bottom = Side(style="thin", color=BLUE_ACCENT)
    for c in (lc, vc):
        c.border = Border(bottom=bottom)


def resolve_asset(*names: str) -> Path | None:
    for name in names:
        path = ASSETS / name
        if path.exists() and path.stat().st_size > 1000:
            return path
    return None


def add_image_scaled(
    ws,
    anchor: str,
    *filenames: str,
    max_width_px: int = 900,
    max_height_px: int = 320,
    allow_upscale: bool = False,
) -> None:
    """Inserta imagen manteniendo proporción (evita deformaciones)."""
    path = resolve_asset(*filenames)
    if path is None:
        return
    nw, nh = max_width_px, max_height_px
    try:
        from PIL import Image as PILImage

        with PILImage.open(path) as im:
            w, h = im.size
            scale = min(max_width_px / w, max_height_px / h)
            if not allow_upscale:
                scale = min(scale, 1.0)
            nw, nh = int(w * scale), int(h * scale)
    except Exception:
        pass
    img = XLImage(str(path))
    img.width = nw
    img.height = nh
    ws.add_image(img, anchor)


INPUT_FILL = PatternFill("solid", fgColor="FFF8E1")
READONLY_FILL = PatternFill("solid", fgColor=GRAY_LIGHT)


def polish_sheet(ws, hide_grid: bool = True) -> None:
    if hide_grid:
        ws.sheet_view.showGridLines = False


def style_chart_mckinsey(chart, *, horizontal: bool = False) -> None:
    """Gráficos limpios: sin rejilla, leyenda abajo, ejes legibles."""
    for axis in (chart.x_axis, chart.y_axis):
        axis.majorGridlines = None
        axis.minorGridlines = None
        axis.tickLblPos = "nextTo"
        axis.delete = False
    if horizontal:
        chart.y_axis.tickLblPos = "nextTo"
        chart.x_axis.tickLblPos = "low"
    else:
        chart.x_axis.tickLblPos = "low"
        chart.y_axis.tickLblPos = "low"
    if chart.legend:
        chart.legend.position = "b"
        chart.legend.overlay = False


CHART_SERIES_COLORS = ["0066CC", "C5A028", "2E7D32", "B71C1C", "6A1B9A"]


def color_chart_series(chart, colors: list[str]) -> None:
    """Asigna colores sólidos a cada serie para gráficos limpios y legibles."""
    for s, color in zip(chart.series, colors):
        s.graphicalProperties.solidFill = color
        s.graphicalProperties.line.solidFill = color


def excel_sheet_name(name: str) -> str:
    """Escapa comillas simples en nombres de hoja para fórmulas Excel."""
    return name.replace("'", "''")


def pull_from_player_sheet(pron_row: int, player: str, goals_col: str, pt_last: int) -> str:
    """Referencia directa a la hoja del jugador (sin INDIRECT — evita corrupción en Excel)."""
    sheet = excel_sheet_name(player)
    r = pron_row
    c = goals_col
    fr = PT_FIRST_ROW
    return (
        f"=IFERROR(INDEX('{sheet}'!{c}${fr}:{c}${pt_last},"
        f"MATCH($A{r},'{sheet}'!$A${fr}:$A${pt_last},0)),\"\")"
    )


def build_player_sheet(
    wb: Workbook,
    player: str,
    matches: list[dict],
    last_row: int,
    tab_color: str,
    demo_predictions: dict[int, dict[str, tuple[int, int]]] | None = None,
) -> None:
    ws = wb.create_sheet(player)
    set_tab_color(ws, tab_color)
    polish_sheet(ws)

    apply_banner(
        ws,
        "A1:H1",
        f"MIS PRONÓSTICOS — {player.upper()}",
        "Rellena goles en todos los partidos y Clasificado solo en eliminatorias; abajo van tus apuestas especiales.",
    )
    ws.row_dimensions[1].height = 44

    headers = [
        "Partido",
        "Fecha",
        "Fase",
        "Equipo local",
        "Equipo visitante",
        "Mis goles local",
        "Mis goles visit.",
        "Clasificado",
    ]
    apply_header_row(ws, headers, row=HEADER_ROW)

    first = PT_FIRST_ROW
    for i, m in enumerate(matches):
        r = first + i
        mid = m["id"]
        ko = match_is_ko(m)
        ws.cell(row=r, column=1, value=mid)
        ws.cell(row=r, column=2, value=m["fecha"])
        ws.cell(row=r, column=3, value=m.get("fase", ""))
        ws.cell(row=r, column=4, value=team_lookup_formula(r, "D", last_row))
        ws.cell(row=r, column=5, value=team_lookup_formula(r, "E", last_row))

        pred = (demo_predictions or {}).get(mid, {}).get(player)
        for col_idx, val in ((6, pred[0] if pred else None), (7, pred[1] if pred else None)):
            cell = ws.cell(row=r, column=col_idx)
            if val is not None:
                cell.value = val
            cell.fill = INPUT_FILL
            cell.alignment = Alignment(horizontal="center")

        cl_cell = ws.cell(row=r, column=8)
        if ko:
            cl_cell.fill = INPUT_FILL
            cl_cell.alignment = Alignment(horizontal="center")
        else:
            cl_cell.value = "-"
            cl_cell.fill = READONLY_FILL
            cl_cell.alignment = Alignment(horizontal="center")

        for col in range(1, 6):
            ws.cell(row=r, column=col).alignment = Alignment(vertical="center")

    last_data = first + len(matches) - 1
    style_data_range(ws, first, last_data, 8)
    for i, m in enumerate(matches):
        r = first + i
        for col_idx in (3, 4, 5):
            ws.cell(row=r, column=col_idx).fill = REFERENCE_FILL
        for col_idx in (6, 7):
            ws.cell(row=r, column=col_idx).fill = INPUT_FILL
            ws.cell(row=r, column=col_idx).alignment = Alignment(horizontal="center")
        cl_cell = ws.cell(row=r, column=8)
        cl_cell.fill = INPUT_FILL if match_is_ko(m) else READONLY_FILL
        cl_cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = f"A{first}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:H{last_data}"

    special_title_row = special_bets_first_row(last_row) - 2
    special_header_row = special_title_row + 1
    special_first = special_header_row + 1
    section_title(ws, f"A{special_title_row}", "Mis apuestas especiales", f"C{special_title_row}")
    apply_header_row(ws, ["Categoría", "Mi apuesta", "Resultado oficial"], row=special_header_row)
    for i, cat in enumerate(SPECIAL_BET_CATEGORIES):
        r = special_first + i
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2).fill = INPUT_FILL
        ws.cell(row=r, column=3, value=f'=IFERROR(Resumen!$B${SPECIAL_RESULTS_FIRST_ROW + i},"")')
        ws.cell(row=r, column=3).fill = READONLY_FILL
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
    style_data_range(ws, special_first, special_first + len(SPECIAL_BET_CATEGORIES) - 1, 3)
    for r in range(special_first, special_first + len(SPECIAL_BET_CATEGORIES)):
        ws.cell(row=r, column=2).fill = INPUT_FILL
        ws.cell(row=r, column=3).fill = READONLY_FILL

    for col, w in zip("ABCDEFGH", [30, 11, 18, 22, 22, 12, 12, 14]):
        ws.column_dimensions[col].width = w


def ultimos_partido_formula(col: str, nth: int, last_row: int) -> str:
    """Últimos partidos: fórmulas simples INDEX+COUNTIF (Excel 2010+, sin LARGE/MATCH)."""
    fr = PT_FIRST_ROW
    n_fin = f'COUNTIF(Partidos!$H${fr}:$H${last_row},"Finalizado")'
    pos = f"{n_fin}-{nth - 1}"
    return (
        f"=IF({nth}<={n_fin},INDEX(Partidos!${col}${fr}:${col}${last_row},{pos}),\"\")"
    )


def ultimos_score_formula(nth: int, last_row: int) -> str:
    fr = PT_FIRST_ROW
    n_fin = f'COUNTIF(Partidos!$H${fr}:$H${last_row},"Finalizado")'
    pos = f"{n_fin}-{nth - 1}"
    return (
        f"=IF({nth}<={n_fin},INDEX(Partidos!$F${fr}:$F${last_row},{pos})&\"-\"&"
        f"INDEX(Partidos!$G${fr}:$G${last_row},{pos}),\"\")"
    )


def validate_workbook_formulas(wb: Workbook) -> None:
    """Evita guardar fórmulas inválidas que Excel repara/borra (p. ej. IF(=IFERROR...))."""
    bad: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                if "(=" in v or ",=" in v or v.startswith("=="):
                    bad.append(f"{ws.title}!{cell.coordinate}: {v[:120]}")
                if v.count("(") != v.count(")"):
                    bad.append(f"{ws.title}!{cell.coordinate}: paréntesis desbalanceados")
    if bad:
        raise ValueError(
            "Fórmulas inválidas detectadas (Excel las borraría al abrir):\n"
            + "\n".join(bad[:10])
        )


def strip_workbook_data_validations(wb: Workbook) -> None:
    """openpyxl escribe extensiones de validación que Excel repara al abrir."""
    for ws in wb.worksheets:
        ws.data_validations.dataValidation.clear()


def sanitize_xlsx_xml(path: Path) -> None:
    """Quita dataValidations del XML (evita aviso WORKBOOK REPAIRED / AutoSave off)."""
    tmp = path.with_suffix(".tmp.xlsx")
    pat = re.compile(r"<dataValidations[^>]*>.*?</dataValidations>", re.DOTALL)
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                text = data.decode("utf-8")
                text = pat.sub("", text)
                data = text.encode("utf-8")
            zout.writestr(item, data)
    replace_tmp_xlsx(tmp, path)


def replace_tmp_xlsx(tmp: Path, path: Path) -> None:
    """OneDrive/Excel can briefly lock the target after save; the workbook is still usable."""
    try:
        tmp.replace(path)
    except PermissionError:
        tmp.unlink(missing_ok=True)


def inject_safe_list_validations(path: Path, players: list[str], matches: list[dict]) -> None:
    """
    Inserta dataValidations 'list' mínimas en el XML del .xlsx (sin usar openpyxl).
    Esto evita el bug de Excel 'WORKBOOK REPAIRED' que provocaban validaciones
    escritas por openpyxl.

    Aplica la lista _Lists!$A$1:$A$2 solo a filas de eliminatorias:
    - Partidos: columna K
    - Cada hoja jugador: columna G
    """
    import xml.etree.ElementTree as ET

    ko_rows = [PT_FIRST_ROW + i for i, match in enumerate(matches) if match_is_ko(match)]
    player_sqref = " ".join(f"H{row}" for row in ko_rows)
    partidos_sqref = " ".join(f"K{row}" for row in ko_rows)

    def _sheet_map(z: zipfile.ZipFile) -> dict[str, str]:
        # Map hoja -> xl/worksheets/sheetN.xml
        wb = ET.fromstring(z.read("xl/workbook.xml"))
        ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
        rid_to_target = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels.findall("r:Relationship", rns)
        }
        out: dict[str, str] = {}
        for sh in wb.findall("m:sheets/m:sheet", ns):
            name = sh.attrib["name"]
            rid = sh.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            target = rid_to_target.get(rid or "")
            if target:
                out[name] = "xl/" + target.lstrip("/")
        return out

    def _inject(worksheet_xml: str, sqref: str) -> str:
        if "<dataValidations" in worksheet_xml:
            worksheet_xml = re.sub(
                r"<dataValidations[^>]*>.*?</dataValidations>", "", worksheet_xml, flags=re.DOTALL
            )
        dv = (
            '<dataValidations count="1">'
            f'<dataValidation type="list" allowBlank="1" showDropDown="1" sqref="{sqref}">'
            "<formula1>_Lists!$A$1:$A$2</formula1>"
            "</dataValidation>"
            "</dataValidations>"
        )
        return worksheet_xml.replace("</worksheet>", dv + "</worksheet>")

    tmp = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(path, "r") as zin:
        sheet_targets = _sheet_map(zin)
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename in sheet_targets.values():
                    # Determine sqref by sheet name
                    sheet_name = next((k for k, v in sheet_targets.items() if v == item.filename), None)
                    if sheet_name == "Partidos":
                        sqref = partidos_sqref
                    elif sheet_name in players:
                        sqref = player_sqref
                    else:
                        sqref = None
                    if sqref:
                        txt = data.decode("utf-8")
                        txt = _inject(txt, sqref)
                        # sanity: well-formed XML
                        ET.fromstring(txt.encode("utf-8"))
                        data = txt.encode("utf-8")
                zout.writestr(item, data)
    replace_tmp_xlsx(tmp, path)


def estado_formula(row: int) -> str:
    r = row
    return f'=IF(AND(ISNUMBER(F{r}),ISNUMBER(G{r})),"Finalizado","Pendiente")'


def build_partidos(wb: Workbook, matches: list[dict]) -> None:
    ws = wb.create_sheet("Partidos")
    set_tab_color(ws, TAB_CORE)
    polish_sheet(ws)
    apply_banner(
        ws,
        "A1:K1",
        "CALENDARIO DE PARTIDOS",
        "Rellena resultados reales en amarillo; Clasificado solo aplica en eliminatorias.",
    )
    ws.row_dimensions[1].height = 40
    headers = [
        "Nº partido",
        "Fecha",
        "Hora",
        "Equipo local",
        "Equipo visitante",
        "Resultado local",
        "Resultado visitante",
        "Estado",
        "Jornada",
        "Fase",
        "Clasificado",
    ]
    apply_header_row(ws, headers, row=HEADER_ROW)
    n = len(matches)

    for i, m in enumerate(matches, start=PT_FIRST_ROW):
        mid = m["id"]
        ko = match_is_ko(m)
        ws.cell(row=i, column=1, value=mid)
        ws.cell(row=i, column=2, value=m["fecha"])
        ws.cell(row=i, column=3, value=m["hora"])
        ws.cell(row=i, column=4, value=m["local"])
        ws.cell(row=i, column=5, value=m["visitante"])
        if mid in DEMO_RESULTS:
            hl, hv = DEMO_RESULTS[mid]
            ws.cell(row=i, column=6, value=hl)
            ws.cell(row=i, column=7, value=hv)
        for col in (6, 7):
            ws.cell(row=i, column=col).fill = INPUT_FILL
        cl_cell = ws.cell(row=i, column=11)
        if ko:
            cl_cell.fill = INPUT_FILL
            cl_cell.alignment = Alignment(horizontal="center")
        else:
            cl_cell.value = "-"
            cl_cell.fill = READONLY_FILL
            cl_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=8, value=estado_formula(i))
        ws.cell(row=i, column=9, value=m.get("jornada", ""))
        ws.cell(row=i, column=10, value=m.get("fase", ""))

    last_data = pt_last_row(n)
    style_data_range(ws, PT_FIRST_ROW, last_data, 11)
    for i, m in enumerate(matches, start=PT_FIRST_ROW):
        for col in (4, 5, 9, 10):
            ws.cell(row=i, column=col).fill = REFERENCE_FILL
        for col in (6, 7):
            ws.cell(row=i, column=col).fill = INPUT_FILL
            ws.cell(row=i, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=8).fill = FORMULA_FILL
        cl_cell = ws.cell(row=i, column=11)
        if match_is_ko(m):
            cl_cell.fill = INPUT_FILL
        else:
            cl_cell.fill = READONLY_FILL
        cl_cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = f"A{PT_FIRST_ROW}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:K{last_data}"

    ws.conditional_formatting.add(
        f"H{PT_FIRST_ROW}:H{last_data}",
        CellIsRule(operator="equal", formula=['"Finalizado"'], fill=PatternFill("solid", fgColor=GREEN_PTS)),
    )
    ws.conditional_formatting.add(
        f"H{PT_FIRST_ROW}:H{last_data}",
        CellIsRule(operator="equal", formula=['"Pendiente"'], fill=PatternFill("solid", fgColor=GRAY_LIGHT)),
    )

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 9
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 14


def build_lists(wb: Workbook) -> None:
    ws = wb.create_sheet("_Lists")
    ws.sheet_state = "hidden"
    ws["A1"] = "Local"
    ws["A2"] = "Visitante"


def build_puntuacion(wb: Workbook, cfg: dict) -> None:
    ws = wb.create_sheet("Puntuacion")
    set_tab_color(ws, TAB_WORK)
    polish_sheet(ws)
    apply_banner(
        ws,
        "A1:D1",
        "CONFIGURACIÓN DE PUNTUACIÓN",
        "Edita solo los puntos en amarillo; el resto del libro se recalcula solo.",
        two_row=False,
    )
    ws.row_dimensions[1].height = 40

    g = cfg["grupos"]
    ko = cfg["eliminatorias"]
    bon = cfg["bonus_ronda"]
    ap = cfg.get("apuesta_especial", 10)

    rows = [
        ("", "", ""),
        ("FASE DE GRUPOS (90 min)", "", ""),
        ("Resultado exacto", g["exacto"], "pts"),
        ("Diferencia de goles correcta", g["diferencia"], "pts"),
        ("Ganador o empate correcto", g["ganador_empate"], "pts"),
        ("", "", ""),
        ("ELIMINATORIAS (×2)", "", ""),
        ("Resultado exacto (90 min)", ko["exacto"], "pts"),
        ("Diferencia de goles (90 min)", ko["diferencia"], "pts"),
        ("Clasificado correcto", ko["clasificado"], "pts"),
        ("Empate exacto, fallo clasificado", ko.get("empate_exacto_falla_clasificado", 6), "pts"),
        ("Empate dif., acierto clasificado", ko.get("empate_diferencia_acierta_clasificado", 6), "pts"),
        ("Empate dif., fallo clasificado", ko.get("empate_diferencia_falla_clasificado", 2), "pts"),
        ("", "", ""),
        ("BONUS RONDA PERFECTA", "", ""),
        ("Octavos (8 clasificados)", bon["octavos"], "pts"),
        ("Cuartos (4 clasificados)", bon["cuartos"], "pts"),
        ("Semifinales (2 clasificados)", bon["semifinal"], "pts"),
        ("", "", ""),
        ("APUESTAS ESPECIALES", "", ""),
        ("Cada acierto", ap, "pts"),
    ]
    for i, (label, val, unit) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label)
        if val != "":
            c = ws.cell(row=i, column=2, value=val)
            c.fill = INPUT_FILL
            c.alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=3, value=unit)
        if label and not val and not unit:
            ws.cell(row=i, column=1).font = Font(bold=True, color=NAVY)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 8


def build_pronosticos(
    wb: Workbook, matches: list[dict], players: list[str], last_row: int
) -> tuple:
    pt_last = last_row
    ws = wb.create_sheet("Pronosticos")
    set_tab_color(ws, TAB_CORE)
    polish_sheet(ws)
    apply_banner(
        ws,
        "A1:J1",
        "PRONÓSTICOS (consolidado)",
        "Vista automática de todos los pronósticos; no hace falta editar esta hoja.",
    )
    ws.row_dimensions[1].height = 40
    headers = [
        "Partido",
        "Equipo local",
        "Equipo visitante",
        "Jugador",
        "Goles local",
        "Goles visitante",
        "Puntos",
        "Clasificado",
        "Fase",
        "Clasif OK",
    ]
    apply_header_row(ws, headers, row=HEADER_ROW)

    row = PT_FIRST_ROW
    pron_rows: list[int] = []
    for m in matches:
        for player in players:
            ws.cell(row=row, column=1, value=m["id"])
            ws.cell(row=row, column=2, value=team_lookup_formula(row, "D", last_row))
            ws.cell(row=row, column=3, value=team_lookup_formula(row, "E", last_row))
            ws.cell(row=row, column=4, value=player)

            ws.cell(row=row, column=5, value=pull_from_player_sheet(row, player, "F", pt_last))
            ws.cell(row=row, column=6, value=pull_from_player_sheet(row, player, "G", pt_last))
            ws.cell(row=row, column=8, value=pull_from_player_sheet(row, player, "H", pt_last))
            for col in (5, 6, 8):
                ws.cell(row=row, column=col).fill = REFERENCE_FILL

            ws.cell(row=row, column=7, value=points_formula(row, last_row))
            ws.cell(row=row, column=9, value=pronosticos_fase_formula(row, last_row))
            ws.cell(row=row, column=10, value=pronosticos_clasif_ok_formula(row, last_row))
            pron_rows.append(row)
            row += 1

    last = row - 1
    style_data_range(ws, PT_FIRST_ROW, last, 10)
    for r in range(PT_FIRST_ROW, last + 1):
        for col in (2, 3, 5, 6, 8, 9):
            ws.cell(row=r, column=col).fill = REFERENCE_FILL
        for col in (7, 10):
            ws.cell(row=r, column=col).fill = FORMULA_FILL
    ws.freeze_panes = f"A{PT_FIRST_ROW}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:H{last}"

    for pts, color in (
        (10, GREEN_PTS),
        (5, GREEN_PTS),
        (6, YELLOW),
        (3, YELLOW),
        (4, YELLOW),
        (2, YELLOW),
        (0, RED_SOFT),
    ):
        ws.conditional_formatting.add(
            f"G{PT_FIRST_ROW}:G{last}",
            CellIsRule(operator="equal", formula=[str(pts)], fill=PatternFill("solid", fgColor=color)),
        )

    for col, w in zip("ABCDEFGHIJ", [9, 20, 20, 14, 11, 14, 8, 12, 12, 8]):
        ws.column_dimensions[col].width = w
    ws.column_dimensions["I"].hidden = False
    ws.column_dimensions["J"].hidden = True

    return last, pron_rows


def jornada_points_formula(
    player_cell: str, match_ids: list[int], pron_last: int
) -> str:
    if not match_ids:
        return "0"
    parts = [
        (
            f"SUMIFS(Pronosticos!$G${PT_FIRST_ROW}:$G${pron_last},Pronosticos!$D${PT_FIRST_ROW}:$D${pron_last},"
            f"{player_cell},Pronosticos!$A${PT_FIRST_ROW}:$A${pron_last},{mid})"
        )
        for mid in match_ids
    ]
    return "=" + "+".join(parts)


def build_helpers(
    wb: Workbook, players: list[str], matches: list[dict], pron_last: int
) -> tuple[int, int, int, int, int]:
    """Construye _Helpers en zonas separadas (sin solapar fórmulas)."""
    n_matches = len(matches)
    n_players = len(players)
    p_first, p_last = player_range(players)
    pron_lr = pron_last
    last_row = pt_last_row(n_matches)
    fin_partidos = f'COUNTIF(Partidos!$H${PT_FIRST_ROW}:$H${last_row},"Finalizado")'

    jornada_ids: dict[int, list[int]] = {}
    for m in matches:
        jn = m.get("jornada", 0)
        jornada_ids.setdefault(jn, []).append(m["id"])
    max_jn = max(jornada_ids.keys()) if jornada_ids else 1

    ws = wb.create_sheet("_Helpers", 2)
    polish_sheet(ws)
    set_tab_color(ws, TAB_WORK)

    # --- Zona 1: estadísticas por jugador (filas 2..p_last, cols A:J) ---
    stat_hdr = [
        "Jugador",
        "Pts partidos",
        "Bonus",
        "Apuestas",
        "Total",
        "Exactos",
        "Parciales",
        "Fallos",
        "% aciertos",
        "Orden",
    ]
    apply_header_row(ws, stat_hdr, row=1)

    for i, name in enumerate(players):
        r = p_first + i
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=match_pts_sum_formula(f"A{r}", pron_lr))
        # Bonus por ronda (cols K:L:M ocultas)
        ws.cell(
            row=r,
            column=11,
            value=round_bonus_formula(f"A{r}", "Octavos", 8, BONUS_OCT, last_row, pron_lr),
        )
        ws.cell(
            row=r,
            column=12,
            value=round_bonus_formula(f"A{r}", "Cuartos", 4, BONUS_CUA, last_row, pron_lr),
        )
        ws.cell(
            row=r,
            column=13,
            value=round_bonus_formula(f"A{r}", "Semifinal", 2, BONUS_SEM, last_row, pron_lr),
        )
        ws.cell(row=r, column=3, value=f"=K{r}+L{r}+M{r}")
        ws.cell(row=r, column=4, value=apuestas_points_formula(name, last_row))
        ws.cell(row=r, column=5, value=f"=B{r}+C{r}+D{r}")
        ws.cell(row=r, column=6, value=exactos_count_formula(f"A{r}", pron_lr))
        ws.cell(row=r, column=7, value=parciales_count_formula(f"A{r}", pron_lr))
        ws.cell(
            row=r,
            column=8,
            value=(
                f"=COUNTIFS(Pronosticos!$D${PT_FIRST_ROW}:$D${pron_lr},A{r},"
                f"Pronosticos!$G${PT_FIRST_ROW}:$G${pron_lr},0)"
            ),
        )
        ws.cell(
            row=r,
            column=9,
            value=f'=IF({fin_partidos}=0,"",(F{r}+G{r})/{fin_partidos})',
        )
        ws.cell(row=r, column=10, value=sort_key_formula(r))

    for col in range(11, 14):
        ws.column_dimensions[get_column_letter(col)].hidden = True

    # --- Zona 2: ranking ordenado (debajo de stats) ---
    r_first = p_last + 3
    r_last = r_first + n_players - 1
    rank_hdr_row = r_first - 1
    apply_header_row(ws, stat_hdr[:6], row=rank_hdr_row)

    for pos in range(1, n_players + 1):
        sr = r_first + pos - 1
        ws.cell(row=sr, column=1, value=ranked_stat_formula(pos, p_first, p_last, "A", "J"))
        ws.cell(row=sr, column=2, value=ranked_stat_formula(pos, p_first, p_last, "E", "J"))
        ws.cell(row=sr, column=3, value=ranked_stat_formula(pos, p_first, p_last, "F", "J"))
        ws.cell(row=sr, column=4, value=ranked_stat_formula(pos, p_first, p_last, "G", "J"))
        ws.cell(row=sr, column=5, value=ranked_stat_formula(pos, p_first, p_last, "H", "J"))
        ws.cell(row=sr, column=6, value=ranked_stat_formula(pos, p_first, p_last, "I", "J"))

    # --- Zona 3: evolución acumulada por partido (cols A + jugadores B..) ---
    evo_hdr = r_last + 2
    ws.cell(row=evo_hdr, column=1, value="Partido")
    for j, p in enumerate(players):
        ws.cell(row=evo_hdr, column=2 + j, value=p)
    evo_steps = min(15, n_matches)
    for m in range(1, evo_steps + 1):
        r = evo_hdr + m
        ws.cell(row=r, column=1, value=m)
        for j, p in enumerate(players):
            ws.cell(
                row=r,
                column=2 + j,
                value=(
                    f'=SUMIFS(Pronosticos!$G${PT_FIRST_ROW}:$G${pron_lr},Pronosticos!$D${PT_FIRST_ROW}:$D${pron_lr},'
                    f'"{p}",Pronosticos!$A${PT_FIRST_ROW}:$A${pron_lr},"<="&$A{r})'
                ),
            )
    evo_end = evo_hdr + evo_steps

    # --- Zona 4: matriz jornadas (antes del resumen jornada) ---
    jm_row = evo_end + 2
    ws.cell(row=jm_row, column=1, value="Jugador")
    for jn in range(1, max_jn + 1):
        ws.cell(row=jm_row, column=1 + jn, value=f"J{jn}")
    for i, _name in enumerate(players):
        kr = p_first + i
        ws.cell(row=jm_row + 1 + i, column=1, value=f"=A{kr}")
        for jn in range(1, max_jn + 1):
            ids = jornada_ids.get(jn, [])
            ws.cell(
                row=jm_row + 1 + i,
                column=1 + jn,
                value=jornada_points_formula(f"A{kr}", ids, pron_lr),
            )
    jm_end = jm_row + n_players

    # --- Zona 5: mejor jugador por jornada ---
    jr = jm_end + 2
    apply_header_row(ws, ["Jornada", "Mejor jugador", "Puntos jornada"], row=jr)
    for jn in range(1, min(13, max_jn + 1)):
        r = jr + jn
        col_letter = get_column_letter(1 + jn)
        data_rng = f"{col_letter}{jm_row + 1}:{col_letter}{jm_end}"
        ws.cell(row=r, column=1, value=jn)
        ws.cell(row=r, column=3, value=f"=MAX({data_rng})")
        ws.cell(
            row=r,
            column=2,
            value=(
                f'=IFERROR(INDEX($A${p_first}:$A${p_last},'
                f"MATCH(C{r},{data_rng},0)),\"\")"
            ),
        )

    # --- Zona 6: datos para gráficos ---
    chart_row = jr + 14
    apply_header_row(ws, ["Jugador", "Puntos", "Exactos", "Parciales", "Fallos"], row=chart_row)
    for i in range(n_players):
        r = chart_row + 1 + i
        rank_row = r_first + i
        ws.cell(row=r, column=1, value=f"=$A${rank_row}")
        ws.cell(row=r, column=2, value=f"=$B${rank_row}")
        ws.cell(row=r, column=3, value=f"=$C${rank_row}")
        ws.cell(row=r, column=4, value=f"=$D${rank_row}")
        ws.cell(row=r, column=5, value=f"=$E${rank_row}")

    # --- Zona 7: evolución top 3 ---
    evo_top_row = chart_row + n_players + 2
    ws.cell(row=evo_top_row, column=1, value="Partidos disputados")
    for rank in range(3):
        ws.cell(row=evo_top_row, column=2 + rank, value=f"=$A${r_first + rank}")
        ws.cell(row=evo_top_row, column=2 + rank).font = Font(bold=True, color=NAVY)
    for m in range(1, evo_steps + 1):
        r = evo_top_row + m
        ws.cell(row=r, column=1, value=m)
        for rank in range(3):
            pl_row = r_first + rank
            ws.cell(
                row=r,
                column=2 + rank,
                value=(
                    f'=SUMIFS(Pronosticos!$G${PT_FIRST_ROW}:$G${pron_lr},Pronosticos!$D${PT_FIRST_ROW}:$D${pron_lr},'
                    f"$A${pl_row},Pronosticos!$A${PT_FIRST_ROW}:$A${pron_lr},\"<=\"&$A{r})"
                ),
            )
    evo_top_end = evo_top_row + evo_steps

    # --- Zona 8: últimos partidos (U:X, ocultas) ---
    ult_first = evo_top_end + 2
    for i in range(5):
        r = ult_first + i
        n = i + 1
        ws.cell(row=r, column=21, value=ultimos_partido_formula("D", n, last_row))
        ws.cell(row=r, column=22, value=ultimos_partido_formula("E", n, last_row))
        ws.cell(row=r, column=23, value=ultimos_score_formula(n, last_row))
        ws.cell(row=r, column=24, value=ultimos_partido_formula("B", n, last_row))
    for col in range(21, 25):
        ws.column_dimensions[get_column_letter(col)].hidden = True

    ws.column_dimensions["A"].width = 14
    for col in "BCDEFGHIJ":
        ws.column_dimensions[col].width = 11
    return chart_row, evo_top_row, evo_top_end, ult_first, r_first


def build_resumen(
    wb: Workbook,
    players: list[str],
    n_matches: int,
    chart_row: int,
    evo_top_row: int,
    evo_top_end: int,
    ult_first: int,
    rank_first: int,
) -> None:
    last_pt = pt_last_row(n_matches)
    r_first = rank_first
    p_first, p_last = player_range(players)
    chart_hdr = chart_row
    chart_data_start = chart_row + 1
    chart_data_end = chart_row + len(players)
    fin = f'COUNTIF(Partidos!$H${PT_FIRST_ROW}:$H${last_pt},"Finalizado")'

    ws = wb.create_sheet("Resumen")
    set_tab_color(ws, TAB_CORE)
    polish_sheet(ws)
    apply_banner(
        ws,
        "A1:N1",
        "RESUMEN — DASHBOARD",
        "Todo se calcula solo. Lo único editable aquí es, en amarillo, el resultado oficial de las apuestas especiales.",
    )

    n = len(players)
    CLASIF_FIRST = 10
    CLASIF_LAST = CLASIF_FIRST + n - 1

    # ===== Indicadores del torneo (KPIs) =====
    section_title(ws, "A4", "Indicadores del torneo", "N4")

    def kpi(col: int, label: str, formula, num_fmt: str | None = None) -> None:
        cl = get_column_letter(col)
        style_kpi(ws, f"{cl}5", f"{cl}6", label, formula)
        nb = ws.cell(row=5, column=col + 1)
        nb.fill = PatternFill("solid", fgColor=WHITE)
        nb.border = Border(bottom=Side(style="thin", color=BLUE_ACCENT))
        nb2 = ws.cell(row=6, column=col + 1)
        nb2.fill = PatternFill("solid", fgColor=WHITE)
        nb2.border = Border(bottom=Side(style="thin", color=BLUE_ACCENT))
        if num_fmt:
            ws[f"{cl}6"].number_format = num_fmt

    kpi(1, "Líder", f"=_Helpers!$A${r_first}")
    kpi(3, "Puntos del líder", f"=_Helpers!$B${r_first}")
    kpi(5, "Partidos jugados", f"={fin}")
    kpi(7, "Partidos pendientes", f"={n_matches}-{fin}")
    kpi(9, "Progreso torneo", f"=IF({fin}=0,0,{fin}/{n_matches})", "0%")
    kpi(11, "Media pts/partido líder", f"=IF({fin}=0,\"\",ROUND(_Helpers!$B${r_first}/{fin},2))")
    kpi(13, "Jugadores", str(n))
    ws.row_dimensions[5].height = 15
    ws.row_dimensions[6].height = 26

    # ===== Clasificación general (zona izquierda A:I) =====
    section_title(ws, "A8", "Clasificación general", "I8")
    apply_header_row(
        ws,
        ["Pos.", "Jugador", "Puntos", "Exactos", "Parciales", "Fallos", "% aciertos", "Medalla", "vs Líder"],
        row=9,
    )
    for i in range(n):
        r = CLASIF_FIRST + i
        sr = r_first + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=f"=_Helpers!$A${sr}")
        ws.cell(row=r, column=3, value=f"=_Helpers!$B${sr}")
        ws.cell(row=r, column=4, value=f"=_Helpers!$C${sr}")
        ws.cell(row=r, column=5, value=f"=_Helpers!$D${sr}")
        ws.cell(row=r, column=6, value=f"=_Helpers!$E${sr}")
        ws.cell(row=r, column=7, value=f"=_Helpers!$F${sr}")
        ws.cell(row=r, column=8, value=medal_formula(r))
        ws.cell(row=r, column=9, value=f"=IF(A{r}=1,0,$C${CLASIF_FIRST}-C{r})")
    style_block(ws, CLASIF_FIRST, CLASIF_LAST, 1, 9)
    for i in range(n):
        r = CLASIF_FIRST + i
        ws.row_dimensions[r].height = 20
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=r, column=7).number_format = "0.0%"
        for col in (3, 4, 5, 6, 7, 8, 9):
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(2, 10):
            ws.cell(row=r, column=col).fill = FORMULA_FILL
    for pos, color in ((1, GOLD), (2, SILVER), (3, BRONZE)):
        ws.conditional_formatting.add(
            f"A{CLASIF_FIRST}:I{CLASIF_LAST}",
            FormulaRule(formula=[f"$A{CLASIF_FIRST}={pos}"], fill=PatternFill("solid", fgColor=color)),
        )

    # ===== Últimos partidos (zona derecha K:N) =====
    section_title(ws, "K8", "Últimos partidos jugados", "N8")
    for c, htxt in enumerate(["Fecha", "Local", "Visitante", "Resultado"], start=11):
        cell = ws.cell(row=9, column=c, value=htxt)
        cell.font = header_font()
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
    for i in range(5):
        r = CLASIF_FIRST + i
        sr = ult_first + i
        ws.cell(row=r, column=11, value=f"=_Helpers!$X${sr}")
        ws.cell(row=r, column=12, value=f"=_Helpers!$U${sr}")
        ws.cell(row=r, column=13, value=f"=_Helpers!$V${sr}")
        ws.cell(row=r, column=14, value=f"=_Helpers!$W${sr}")
    style_block(ws, CLASIF_FIRST, CLASIF_FIRST + 4, 11, 14)
    for r in range(CLASIF_FIRST, CLASIF_FIRST + 5):
        ws.row_dimensions[r].height = 20
        for col in range(11, 15):
            ws.cell(row=r, column=col).fill = REFERENCE_FILL
        ws.cell(row=r, column=14).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ===== Análisis por jugador (selector) =====
    keys = f"_Helpers!$A${p_first}:$A${p_last}"
    match_pts = f"_Helpers!$B${p_first}:$B${p_last}"
    bonus_pts = f"_Helpers!$C${p_first}:$C${p_last}"
    special_pts = f"_Helpers!$D${p_first}:$D${p_last}"
    vals = f"_Helpers!$E${p_first}:$E${p_last}"
    sel = "$C$23"
    section_title(ws, "A22", "Análisis por jugador", "I22")
    lbl = ws.cell(row=23, column=1, value="Elige un jugador:")
    lbl.font = Font(bold=True, color=NAVY)
    sel_cell = ws.cell(row=23, column=3, value=players[0])
    sel_cell.fill = INPUT_FILL
    sel_cell.alignment = Alignment(horizontal="center")
    sel_cell.border = thin_border()
    hint = ws.cell(row=23, column=4, value="(escribe el nombre exacto de una pestaña)")
    hint.font = Font(size=9, italic=True, color=GRAY_MID)
    # MECE: solo el desglose que la tabla de clasificación no muestra
    breakdown_metrics = [
        ("Posición actual", f"=IFERROR(MATCH({sel},_Helpers!$A${r_first}:$A${r_first + n - 1},0),\"\")"),
        ("Puntos totales", f"=IFERROR({idx_match(sel, keys, vals)},0)"),
        ("Puntos por partidos", f"=IFERROR({idx_match(sel, keys, match_pts)},0)"),
        ("Bonus de rondas", f"=IFERROR({idx_match(sel, keys, bonus_pts)},0)"),
        ("Puntos apuestas especiales", f"=IFERROR({idx_match(sel, keys, special_pts)},0)"),
        ("Distancia al líder", f"=IFERROR(_Helpers!$B${r_first}-{idx_match(sel, keys, vals)},0)"),
    ]
    apply_header_row(ws, ["Métrica", "Valor"], row=24)
    ana_first = 25
    for i, (label, formula) in enumerate(breakdown_metrics):
        r = ana_first + i
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=formula)
    ana_last = ana_first + len(breakdown_metrics) - 1
    style_block(ws, ana_first, ana_last, 1, 2)
    for r in range(ana_first, ana_last + 1):
        ws.row_dimensions[r].height = 20
        ws.cell(row=r, column=2).fill = FORMULA_FILL
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ===== Apuestas especiales: resultado oficial (izq) + puntos (der) =====
    sp_first = SPECIAL_RESULTS_FIRST_ROW
    sp_title = sp_first - 2
    sp_header = sp_first - 1
    section_title(ws, f"A{sp_title}", "Apuestas especiales · resultado oficial", f"D{sp_title}")
    apply_header_row(ws, ["Categoría", "Resultado oficial", "Aciertos", "Estado"], row=sp_header)
    for i, cat in enumerate(SPECIAL_BET_CATEGORIES):
        r = sp_first + i
        ws.cell(row=r, column=1, value=cat)
        hit_parts = [
            f"IF('{excel_sheet_name(player)}'!$B${special_bets_first_row(last_pt) + i}=$B{r},1,0)"
            for player in players
        ]
        ws.cell(row=r, column=3, value=f"=IF($B{r}=\"\",0,{'+'.join(hit_parts)})")
        ws.cell(row=r, column=4, value=f'=IF($B{r}="","Pendiente","Cargado")')
    sp_results_last = sp_first + len(SPECIAL_BET_CATEGORIES) - 1
    style_block(ws, sp_first, sp_results_last, 1, 4)
    for r in range(sp_first, sp_results_last + 1):
        ws.row_dimensions[r].height = 24
        ws.cell(row=r, column=2).fill = INPUT_FILL
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=r, column=3).fill = FORMULA_FILL
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=r, column=4).fill = FORMULA_FILL
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    section_title(ws, f"F{sp_title}", "Puntos por apuestas especiales", f"I{sp_title}")
    for c, htxt in enumerate(["Jugador", "Puntos", "Aciertos", "Peso"], start=6):
        cell = ws.cell(row=sp_header, column=c, value=htxt)
        cell.font = header_font()
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
    for i, player in enumerate(players):
        r = sp_first + i
        sr = p_first + i
        ws.cell(row=r, column=6, value=f"=_Helpers!$A${sr}")
        ws.cell(row=r, column=7, value=f"=_Helpers!$D${sr}")
        ws.cell(row=r, column=8, value=f"=IF({PTS_APUESTA}=0,0,G{r}/{PTS_APUESTA})")
        ws.cell(row=r, column=9, value=f"={PTS_APUESTA}")
    sp_points_last = sp_first + n - 1
    style_block(ws, sp_first, sp_points_last, 6, 9)
    for r in range(sp_first, sp_points_last + 1):
        current_height = ws.row_dimensions[r].height or 0
        ws.row_dimensions[r].height = max(current_height, 20)
        for col in (6, 7, 8, 9):
            ws.cell(row=r, column=col).fill = FORMULA_FILL
        for col in (7, 8, 9):
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ===== Visualizaciones (rejilla 2x2, bien separada) =====
    h = wb["_Helpers"]
    section_title(ws, "A46", "Visualizaciones", "N46")
    cats = Reference(h, min_col=1, min_row=chart_data_start, max_row=chart_data_end)

    # 1) Ranking por puntos (barra horizontal, con etiquetas)
    pts_bar = BarChart()
    pts_bar.type = "bar"
    pts_bar.barDir = "bar"
    pts_bar.title = "Puntos totales por jugador"
    pts_bar.y_axis.title = None
    pts_bar.x_axis.title = "Puntos"
    pts_bar.add_data(
        Reference(h, min_col=2, min_row=chart_data_start, max_row=chart_data_end),
        titles_from_data=False,
    )
    pts_bar.series[0].title = SeriesLabel(v="Puntos")
    pts_bar.set_categories(cats)
    pts_bar.gapWidth = 55
    pts_bar.legend = None
    pts_bar.dataLabels = DataLabelList()
    pts_bar.dataLabels.showVal = True
    style_chart_mckinsey(pts_bar, horizontal=True)
    color_chart_series(pts_bar, [NAVY])
    pts_bar.width = 15.5
    pts_bar.height = 8.5
    ws.add_chart(pts_bar, "A47")

    # 2) Origen de puntos (apilado: partidos + bonus + apuestas)
    source = BarChart()
    source.type = "bar"
    source.barDir = "bar"
    source.grouping = "stacked"
    source.overlap = 100
    source.title = "Origen de los puntos por jugador"
    source.x_axis.title = "Puntos"
    source.set_categories(Reference(h, min_col=1, min_row=p_first, max_row=p_last))
    source.add_data(Reference(h, min_col=2, max_col=4, min_row=1, max_row=p_last), titles_from_data=True)
    source.gapWidth = 55
    style_chart_mckinsey(source, horizontal=True)
    color_chart_series(source, [BLUE_ACCENT, GOLD, NAVY])
    source.width = 15.5
    source.height = 8.5
    ws.add_chart(source, "H47")

    # 3) Evolución acumulada del Top 3
    line = LineChart()
    line.title = "Evolución acumulada — Top 3"
    line.y_axis.title = "Puntos acumulados"
    line.x_axis.title = "Partidos disputados"
    line.set_categories(Reference(h, min_col=1, min_row=evo_top_row + 1, max_row=evo_top_end))
    for col in (2, 3, 4):
        line.add_data(
            Reference(h, min_col=col, min_row=evo_top_row, max_row=evo_top_end),
            titles_from_data=True,
        )
    for s in line.series:
        s.smooth = False
        s.graphicalProperties.line.width = 28000
    style_chart_mckinsey(line)
    color_chart_series(line, [GOLD, BLUE_ACCENT, "2E7D32"])
    line.width = 15.5
    line.height = 8.5
    ws.add_chart(line, "A66")

    # 4) Composición de aciertos (apilado: exactos / parciales / fallos)
    breakdown = BarChart()
    breakdown.type = "bar"
    breakdown.barDir = "bar"
    breakdown.grouping = "stacked"
    breakdown.overlap = 100
    breakdown.title = "Composición de aciertos por jugador"
    breakdown.x_axis.title = "Número de pronósticos"
    breakdown.set_categories(cats)
    for col, title in ((3, "Exactos"), (4, "Parciales"), (5, "Fallos")):
        breakdown.add_data(
            Reference(h, min_col=col, min_row=chart_data_start, max_row=chart_data_end),
            titles_from_data=False,
        )
        breakdown.series[-1].title = SeriesLabel(v=title)
    breakdown.gapWidth = 55
    style_chart_mckinsey(breakdown, horizontal=True)
    color_chart_series(breakdown, ["2E7D32", GOLD, "B71C1C"])
    breakdown.width = 15.5
    breakdown.height = 8.5
    ws.add_chart(breakdown, "H66")

    resumen_widths = [32, 24, 12, 12, 12, 14, 14, 13, 13, 3, 14, 18, 18, 13]
    for col, w in zip("ABCDEFGHIJKLMN", resumen_widths):
        ws.column_dimensions[col].width = w
    for spacer in (3, 7, 21, 31, 45, 64):
        ws.row_dimensions[spacer].height = 8


def build_portada(wb: Workbook, players: list[str], group_label: str = "") -> None:
    ws = wb.create_sheet("Portada", 0)
    set_tab_color(ws, TAB_INTRO)
    ws.sheet_view.showGridLines = False

    for col in "ABCDEFGHIJKLMN":
        ws.column_dimensions[col].width = 12
    for r in range(1, 40):
        ws.row_dimensions[r].height = 22

    # Hero principal del Mundial 2026 (trofeo + balón + países anfitriones).
    add_image_scaled(
        ws,
        "A1",
        "wc2026_hero.jpg",
        "wc2026_crowd.jpg",
        "wc2026_stadium.jpg",
        "stadium.jpg",
        max_width_px=760,
        max_height_px=470,
        allow_upscale=True,
    )

    # Balón temático a la derecha del hero.
    add_image_scaled(
        ws,
        "I2",
        "wc2026_ball.jpg",
        "football.jpg",
        "wc2026_stadium.jpg",
        "stadium.jpg",
        max_width_px=430,
        max_height_px=300,
        allow_upscale=True,
    )

    # Banda de título inferior: textos centrados sin merge (Center Across Selection).
    for r in range(26, 39):
        for c in range(1, 15):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=NAVY)

    center_across(
        ws, 27, 1, 14, "PORRA MUNDIAL 2026",
        Font(name="Calibri", size=36, bold=True, color=WHITE), NAVY,
    )
    center_across(
        ws, 30, 1, 14, "FIFA WORLD CUP 26  ·  USA · CANADÁ · MÉXICO",
        Font(name="Calibri", size=16, bold=True, color=GOLD), NAVY,
    )
    subtitle = (
        f"Grupo {group_label} · Porra privada · Pronósticos, clasificación automática y apuestas especiales"
        if group_label
        else "Porra privada entre amigos · Pronósticos, clasificación automática y apuestas especiales"
    )
    center_across(
        ws, 33, 1, 14,
        subtitle,
        Font(name="Calibri", size=12, italic=True, color=WHITE), NAVY,
    )
    center_across(
        ws, 36, 1, 14, "Jugadores:  " + "   ·   ".join(players),
        Font(name="Calibri", size=11, color=BLUE_LIGHT, italic=True), NAVY,
    )
    ws.row_dimensions[27].height = 40


def build_instrucciones(wb: Workbook) -> None:
    ws = wb.create_sheet("Instrucciones")
    polish_sheet(ws)
    set_tab_color(ws, TAB_INTRO)
    apply_banner(ws, "A1:E1", "INSTRUCCIONES", "Guía paso a paso para usar la porra sin tocar fórmulas.")
    ws.row_dimensions[1].height = 36

    steps = [
        (
            "1",
            "Empieza por tu pestaña",
            "Busca la pestaña con tu nombre. Ahí rellenas todo lo tuyo: goles de cada partido, clasificado en eliminatorias y apuestas especiales. "
            "Solo escribe en celdas amarillas. Las celdas azules son fórmulas y las grises son referencias de otras pestañas.",
        ),
        (
            "2",
            "Pronósticos de partidos",
            "En fase de grupos rellena Mis goles local y Mis goles visit.; la columna Clasificado aparece con '-' y no se usa. "
            "En octavos, cuartos, semifinales y final, además de los goles a 90 minutos, elige Clasificado: Local o Visitante.",
        ),
        (
            "3",
            "Apuestas especiales",
            "En la parte inferior de tu pestaña escribe tus apuestas especiales antes de que empiece el torneo. "
            "El resultado oficial de esas apuestas se rellena después en Resumen, dentro del bloque Apuestas especiales.",
        ),
        (
            "4",
            "Resultados reales",
            "La persona que administre la porra usa Partidos. En grupos rellena Resultado local y Resultado visitante. "
            "En eliminatorias rellena también Clasificado, que es quien pasa la ronda aunque sea en prórroga o penaltis.",
        ),
        (
            "5",
            "Qué mirar durante el torneo",
            "Resumen es el dashboard principal: líder, clasificación completa, últimos partidos, análisis por jugador y apuestas especiales. "
            "Pronosticos consolida lo que ha puesto cada jugador y es de solo lectura.",
        ),
        (
            "6",
            "Configuración y hojas técnicas",
            "Puntuacion permite cambiar los puntos de cada regla. Helpers >> separa las hojas de trabajo. _Helpers contiene cálculos internos y no requiere input.",
        ),
    ]
    row = 3
    for num, title, desc in steps:
        ws.cell(row=row, column=1, value=num).font = Font(bold=True, size=14, color=NAVY)
        ws.cell(row=row, column=2, value=title).font = Font(bold=True, size=12)
        row += 1
        ws.cell(row=row, column=2, value=desc).alignment = Alignment(wrap_text=True)
        row += 2

    ws["A22"] = "Pestañas con input: tu nombre, Partidos, Resumen (solo resultados oficiales de apuestas especiales) y Puntuacion."
    ws["A23"] = "Pestañas sin input: Pronosticos, Helpers >> y _Helpers. Regla visual: amarillo = input; azul = fórmula; gris = referencia."
    ws["A24"] = "Requisitos: Microsoft Excel 2010 o superior (Windows/Mac). No usa macros."
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 84


def build_helper_separator(wb: Workbook) -> None:
    ws = wb.create_sheet("Helpers >>")
    set_tab_color(ws, TAB_WORK)
    polish_sheet(ws)
    ws["A1"] = "Helpers >>"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=GRAY_MID)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")


def main(group_id: str = "broshu") -> None:
    group = load_group(group_id)
    players = group["players"]
    output_path = group["excel_path"]
    demo_predictions = DEMO_PREDICTIONS_BY_GROUP.get(group_id, {})

    matches_path = ROOT / "data" / "matches_2026.json"
    if not matches_path.exists():
        import build_matches

        build_matches.main()
    matches = load_json(matches_path)
    n_matches = len(matches)

    wb = Workbook()
    wb.remove(wb.active)

    last_row = pt_last_row(n_matches)
    p_first, _ = player_range(players)
    scoring_cfg = load_scoring_config()

    player_tab_colors = [TAB_WORK]

    build_portada(wb, players, group_label=group["label"])
    build_lists(wb)
    build_puntuacion(wb, scoring_cfg)
    for i, player in enumerate(players):
        build_player_sheet(
            wb,
            player,
            matches,
            last_row,
            player_tab_colors[i % len(player_tab_colors)],
            demo_predictions=demo_predictions,
        )
    build_partidos(wb, matches)
    pron_last, _ = build_pronosticos(wb, matches, players, last_row)
    chart_row, evo_top_row, evo_top_end, ult_first, rank_first = build_helpers(
        wb, players, matches, pron_last
    )
    build_resumen(
        wb, players, n_matches, chart_row, evo_top_row, evo_top_end, ult_first, rank_first
    )
    build_instrucciones(wb)
    build_helper_separator(wb)

    wb.active = wb["Portada"]

    sheet_order = (
        [
            "Portada",
            "Instrucciones",
            "Resumen",
            "Partidos",
            "Pronosticos",
            "Helpers >>",
        ]
        + players
        + ["Puntuacion", "_Helpers", "_Lists"]
    )
    for idx, name in enumerate(sheet_order):
        ws = wb[name]
        wb._sheets.remove(ws)
        wb._sheets.insert(idx, ws)

    validate_workbook_formulas(wb)
    strip_workbook_data_validations(wb)

    from openpyxl.workbook.properties import CalcProperties

    wb.calculation = CalcProperties(fullCalcOnLoad=True, calcMode="auto")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    sanitize_xlsx_xml(output_path)
    inject_safe_list_validations(output_path, players, matches)
    print(f"Generado: {output_path}")
    print(f"  Grupo: {group['label']}")
    print(f"  Partidos: {n_matches}")
    print(f"  Pronosticos: {pron_last - 1} filas ({n_matches} x {len(players)} jugadores)")
    print(f"  Jugadores: {', '.join(players)}")

    try:
        import build_dashboard

        build_dashboard.main(group_id=group_id, open_browser=False)
    except Exception as exc:
        print(f"  Aviso: no se pudo generar {group['dashboard']} ({exc})")


if __name__ == "__main__":
    import sys

    for gid in resolve_group_ids(sys.argv):
        main(gid)
