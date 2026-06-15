"""Importa fixtures desde Excel de la porra + calendario FIFA → matches_2026.json."""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from fifa_schedule_2026 import (
    build_all_matches,
    jornada_for_group,
    lookup_pair_meta,
)

ROOT = Path(__file__).parent
OUTPUT = ROOT / "data" / "matches_2026.json"
DEFAULT_EXCEL = Path(
    r"c:\Users\Alvaro J Perez Triay\OneDrive - McKinsey & Company\Desktop"
    r"\Personal\3. Personal\4. Broshu\Porra_Mundial_2026_Correo.xlsx"
)
PT_FIRST_ROW = 4


def read_group_fixtures_from_excel(excel_path: Path) -> list[tuple[str, str]]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    fixtures: list[tuple[str, str]] = []
    for row in ws.iter_rows(min_row=PT_FIRST_ROW, values_only=True):
        mid, fase, local, visitante = row[0], row[1], row[2], row[3]
        if mid is None:
            break
        if int(mid) > 72:
            break
        if str(fase) != "Grupos":
            raise ValueError(f"Partido {mid}: se esperaba fase Grupos, got {fase!r}")
        fixtures.append((str(local).strip(), str(visitante).strip()))
    wb.close()
    if len(fixtures) != 72:
        raise ValueError(f"Se esperaban 72 partidos de grupos, se leyeron {len(fixtures)}")
    return fixtures


def build_matches_from_excel(excel_path: Path) -> list[dict]:
    group_fixtures = read_group_fixtures_from_excel(excel_path)
    matches: list[dict] = []

    for i, (local, visitante) in enumerate(group_fixtures, start=1):
        meta = lookup_pair_meta(local, visitante)
        grupo = meta["grupo"]
        matches.append(
            {
                "id": i,
                "fecha": meta["fecha"],
                "hora": meta["hora"],
                "local": local,
                "visitante": visitante,
                "grupo": grupo,
                "fase": "Grupos",
                "jornada": jornada_for_group(grupo),
            }
        )

    canonical = build_all_matches()
    for m in canonical:
        if m["id"] >= 73:
            matches.append(m)

    return matches


def main() -> None:
    excel_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_EXCEL
    if not excel_path.exists():
        print(f"Excel no encontrado: {excel_path}")
        print("Usando orden por defecto embebido (PORRA_GROUP_FIXTURES).")
        matches = build_all_matches()
    else:
        matches = build_matches_from_excel(excel_path)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT.open("w", encoding="utf-8") as f:
        json.dump(matches, f, ensure_ascii=False, indent=2)
    print(f"Written {len(matches)} matches to {OUTPUT}")
    print(f"  #1: {matches[0]['local']} vs {matches[0]['visitante']} ({matches[0]['fecha']})")
    print(f"  #104: {matches[-1]['local']} vs {matches[-1]['visitante']} ({matches[-1]['fecha']})")


if __name__ == "__main__":
    main()
