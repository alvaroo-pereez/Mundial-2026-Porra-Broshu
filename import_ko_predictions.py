"""
Importa pronósticos de eliminatorias desde porras de amigos al Excel maestro.
Octavos: empareja por ID de partido (89-96) con fallback para filas desordenadas.
"""
from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from config.groups import load_group
from import_r32_predictions import (
    COLS_NEW,
    COLS_OLD,
    _as_int,
    build_pair_index,
    detect_layout,
    pair_key,
    teams_match,
)
from octavos_fixtures import (
    OCTAVOS_TEMPLATE_PAIRS,
    template_id_to_calendar,
)

ROOT = Path(__file__).parent
PT_FIRST_ROW = 4
OCTAVOS_FIRST_ID = 89
OCTAVOS_LAST_ID = 96
PROTECT_LAST_ROW = 88  # filas 4-88 = grupos + dieciseisavos

PHASE_RANGES = {
    "dieciseisavos": (73, 88, "Dieci"),
    "octavos": (89, 96, "Octav"),
    "cuartos": (97, 100, "Cuart"),
    "semifinal": (101, 102, "Semif"),
}

# Plantilla Excel amigos (ids fila 89-96)
STANDARD_OCTAVOS = OCTAVOS_TEMPLATE_PAIRS

_TEMPLATE_PAIR_INDEX: dict[tuple[str, str], int] = {
    pair_key(loc, vis): mid for mid, (loc, vis) in OCTAVOS_TEMPLATE_PAIRS.items()
}

# (path, group_id, player_sheet_master, source_sheet, label)
IMPORTS_OCTAVOS: list[tuple[str, str, str, str, str]] = [
    (
        r"c:\Users\Alvaro J Perez Triay\AppData\Local\Temp\PORRA MUNDIAL NACH.xlsx",
        "broshu",
        "Nacho",
        "Nach",
        "Nacho",
    ),
    (
        r"c:\Users\Alvaro J Perez Triay\AppData\Local\Temp\Porra_Mundial_2026_Patricio.xlsx",
        "broshu",
        "Patri",
        "Patricio",
        "Patricio",
    ),
    (
        r"c:\Users\Alvaro J Perez Triay\AppData\Local\Temp\PORRA KIKOTA.xlsx",
        "broshu",
        "Kike",
        "Álvaro",
        "Kike",
    ),
    (
        r"c:\Users\Alvaro J Perez Triay\AppData\Local\Temp\Porra_Mundial_2026_Pepe (1).xlsx",
        "broshu",
        "Pepe",
        "Álvaro",
        "Pepe",
    ),
    (
        r"c:\Users\Alvaro J Perez Triay\AppData\Local\Temp\Porra_Mundial_2026_Correo_Quintero-8AVOS.xlsx",
        "broshu",
        "Quintero",
        "Jaime",
        "Quintero",
    ),
    (
        r"c:\Users\Alvaro J Perez Triay\AppData\Local\Temp\Porra_Mundial_2026_Correo.xlsx",
        "broshu",
        "Luis",
        "Álvaro",
        "Luis",
    ),
    (
        r"c:\Users\Alvaro J Perez Triay\AppData\Local\Temp\Porra_Mundial_2026_papi.xlsx",
        "papinenes",
        "Papá",
        "Álvaro",
        "Papa",
    ),
    (
        r"c:\Users\Alvaro J Perez Triay\AppData\Local\Temp\Porra_Mundial_2026_Diego.xlsx",
        "papinenes",
        "Diego",
        "Álvaro",
        "Diego",
    ),
]

IMPORTS_CUARTOS: list[tuple[str, str, str, str, str]] = [
    (
        r"c:\Users\Alvaro J Perez Triay\AppData\Local\Temp\Porra_Mundial_2026_Correo.xlsx",
        "broshu",
        "Luis",
        "Álvaro",
        "Luis",
    ),
    (
        r"c:\Users\Alvaro J Perez Triay\AppData\Local\Temp\Porra_Mundial_2026_papi.xlsx",
        "papinenes",
        "Papá",
        "Álvaro",
        "Papa",
    ),
    (
        r"c:\Users\Alvaro J Perez Triay\AppData\Local\Temp\Porra_Mundial_2026_Diego.xlsx",
        "papinenes",
        "Diego",
        "Álvaro",
        "Diego",
    ),
]

SOURCE_SHEET_FALLBACKS: dict[str, list[str]] = {
    "Nach": ["Nacho"],
    "Álvaro": ["Alvaro", "Luis", "Kike", "Pepe", "Papa", "Papá", "Diego"],
    "Patricio": ["Patri"],
    "Jaime": ["Quintero"],
}


@dataclass
class KoRow:
    row: int
    file_id: int | None
    local: str
    visit: str
    gh: int | None
    ga: int | None
    clasif_raw: object


@dataclass
class ImportResult:
    label: str
    player: str
    matched: dict[int, tuple[int, int, str | None]] = field(default_factory=dict)
    flags: list[str] = field(default_factory=list)


def open_source_sheet(wb, source_sheet: str):
    if source_sheet in wb.sheetnames:
        return wb[source_sheet]
    for alt in SOURCE_SHEET_FALLBACKS.get(source_sheet, []):
        if alt in wb.sheetnames:
            return wb[alt]
    raise KeyError(f"Hoja '{source_sheet}' no encontrada en {wb.sheetnames}")


def _row_matches_standard(row: KoRow, std_id: int) -> bool:
    if std_id not in STANDARD_OCTAVOS:
        return False
    std_loc, std_vis = STANDARD_OCTAVOS[std_id]
    matches = 0
    if teams_match(row.local, std_loc):
        matches += 1
    if teams_match(row.visit, std_vis):
        matches += 1
    return matches >= 1


def normalize_clasificado_ko(
    raw: object,
    friend_local: str,
    friend_visit: str,
    gh: int | None,
    ga: int | None,
) -> tuple[str | None, str | None]:
    """Clasificado usando equipos del archivo amigo (ids de slot, no master)."""
    if raw in (1, 1.0, "1"):
        return "Local", None
    if raw in (2, 2.0, "2"):
        return "Visitante", None

    if raw not in (None, "", "-"):
        s = str(raw).strip()
        sl = s.lower()
        if sl == "local":
            return "Local", None
        if sl == "visitante":
            return "Visitante", None
        if teams_match(s, friend_local):
            return "Local", f"clasif nombre '{s}' -> Local"
        if teams_match(s, friend_visit):
            return "Visitante", f"clasif nombre '{s}' -> Visitante"
        if gh is not None and ga is not None and gh != ga:
            side = "Local" if gh > ga else "Visitante"
            return side, f"clasif no reconocido '{s}' -> inferido {side} por marcador"

    if gh is not None and ga is not None:
        if gh > ga:
            return "Local", None
        if ga > gh:
            return "Visitante", None
        return None, f"empate {gh}-{ga} sin clasificado"

    return None, "goles incompletos"


def _row_matches_master_id(
    row: KoRow, mid: int, master_pairs: dict[int, tuple[str, str]]
) -> bool:
    pair = master_pairs.get(mid)
    if not pair:
        return False
    loc, vis = pair
    return teams_match(row.local, loc) and teams_match(row.visit, vis)


def resolve_master_id(
    row: KoRow,
    master_pairs: dict[int, tuple[str, str]],
    pair_index: dict[tuple[str, str], int],
    first_id: int,
    last_id: int,
) -> tuple[int | None, str]:
    """Resuelve id calendario master y método de emparejamiento."""
    template_id: int | None = None
    method = "none"

    if row.file_id is not None and first_id <= row.file_id <= last_id:
        mid_by_pair = pair_index.get(pair_key(row.local, row.visit))
        if not _row_matches_master_id(row, row.file_id, master_pairs):
            if mid_by_pair is not None:
                return mid_by_pair, "master_pair"
        if _row_matches_standard(row, row.file_id):
            template_id, method = row.file_id, "id"
        else:
            alt = _TEMPLATE_PAIR_INDEX.get(pair_key(row.local, row.visit))
            if alt is not None:
                template_id, method = alt, "template_pair"
            elif mid_by_pair is not None:
                return mid_by_pair, "master_pair"
            else:
                template_id, method = row.file_id, "id_fallback"
    else:
        alt = _TEMPLATE_PAIR_INDEX.get(pair_key(row.local, row.visit))
        if alt is not None:
            template_id, method = alt, "template_pair"
        else:
            mid = pair_index.get(pair_key(row.local, row.visit))
            if mid is not None:
                return mid, "master_pair"

    if template_id is not None:
        calendar_id = template_id_to_calendar(template_id)
        return calendar_id, method

    return None, "none"


def load_master_pairs(master_path: Path, first_id: int, last_id: int) -> dict[int, tuple[str, str]]:
    wb = load_workbook(master_path, data_only=True)
    pt = wb["Partidos"]
    pairs: dict[int, tuple[str, str]] = {}
    for mid in range(first_id, last_id + 1):
        row = PT_FIRST_ROW + mid - 1
        loc = pt.cell(row, 4).value
        vis = pt.cell(row, 5).value
        if not loc or not vis:
            continue
        loc_s, vis_s = str(loc).strip(), str(vis).strip()
        if loc_s.startswith("Ganador") or loc_s.startswith("Perdedor"):
            continue
        pairs[mid] = (loc_s, vis_s)
    wb.close()
    return pairs


def extract_ko_rows(ws, cols: dict[str, int], phase_token: str) -> list[KoRow]:
    rows: list[KoRow] = []
    for r in range(PT_FIRST_ROW, ws.max_row + 1):
        fase = ws.cell(r, cols["fase"]).value
        if not fase or phase_token not in str(fase):
            continue
        local = ws.cell(r, cols["local"]).value
        visit = ws.cell(r, cols["visit"]).value
        if not local or not visit:
            continue
        loc_s = str(local).strip().rstrip(",")
        vis_s = str(visit).strip().rstrip(",")
        if loc_s.startswith("Ganador") or loc_s.startswith("Perdedor"):
            continue
        raw_id = ws.cell(r, cols["id"]).value
        file_id = None
        if raw_id is not None:
            try:
                file_id = int(raw_id)
            except (TypeError, ValueError):
                pass
        gh = _as_int(ws.cell(r, cols["gh"]).value)
        ga = _as_int(ws.cell(r, cols["ga"]).value)
        cl = ws.cell(r, cols["cl"]).value
        rows.append(
            KoRow(
                row=r,
                file_id=file_id,
                local=loc_s,
                visit=vis_s,
                gh=gh,
                ga=ga,
                clasif_raw=cl,
            )
        )
    return rows


def parse_source(
    source_path: Path,
    master_pairs: dict[int, tuple[str, str]],
    pair_index: dict[tuple[str, str], int],
    label: str,
    phase_token: str,
    source_sheet: str,
    first_id: int,
    last_id: int,
) -> ImportResult:
    result = ImportResult(label=label, player="")
    if not source_path.exists():
        result.flags.append(f"ARCHIVO NO ENCONTRADO: {source_path}")
        return result

    wb = load_workbook(source_path, data_only=True)
    try:
        ws = open_source_sheet(wb, source_sheet)
    except KeyError as exc:
        result.flags.append(str(exc))
        wb.close()
        return result

    cols = detect_layout(ws)
    layout = "new" if cols is COLS_NEW else "old"
    result.flags.append(f"hoja fuente: {ws.title} (layout {layout})")

    for row in extract_ko_rows(ws, cols, phase_token):
        mid, method = resolve_master_id(
            row, master_pairs, pair_index, first_id, last_id
        )
        if mid is None:
            result.flags.append(
                f"NO MATCH fila {row.row}: {row.local} vs {row.visit} ({row.gh}-{row.ga})"
            )
            continue

        ml, mv = master_pairs[mid]
        gh, ga = row.gh, row.ga
        if gh is not None and ga is None:
            ga = 0
            result.flags.append(f"fila {row.row}: goles visitante vacío, asumido 0")
        elif ga is not None and gh is None:
            gh = 0
            result.flags.append(f"fila {row.row}: goles local vacío, asumido 0")

        if gh is None or ga is None:
            result.flags.append(f"INCOMPLETO master {mid}: goles {gh}-{ga}")
            continue

        cl, cl_flag = normalize_clasificado_ko(row.clasif_raw, row.local, row.visit, gh, ga)
        if cl_flag:
            result.flags.append(f"master {mid}: {cl_flag}")
        if cl is None:
            result.flags.append(f"REVISAR master {mid}: empate sin clasificado ({gh}-{ga})")

        if row.file_id is not None and row.file_id != mid:
            result.flags.append(
                f"fila {row.row} id {row.file_id} -> master {mid} ({method})"
            )
        elif method == "template_pair":
            result.flags.append(
                f"fila {row.row} reordenada -> master {mid} ({method})"
            )

        result.matched[mid] = (gh, ga, cl)

    wb.close()
    missing = [m for m in master_pairs if m not in result.matched]
    if missing:
        result.flags.append(f"FALTAN partidos (master resueltos): {missing}")
    return result


def snapshot_protected_cells(wb_path: Path, player: str) -> dict:
    wb = load_workbook(wb_path, data_only=True)
    if player not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb[player]
    snap: dict = {}
    for row in range(PT_FIRST_ROW, PT_FIRST_ROW + PROTECT_LAST_ROW - 1):
        for col in (6, 7, 8):
            snap[(row, col)] = ws.cell(row, col).value
    wb.close()
    return snap


def write_ko_batch_to_master(
    master_path: Path,
    batch: dict[str, dict[int, tuple[int, int, str | None]]],
    first_id: int,
    last_id: int,
) -> None:
    import time

    wb = load_workbook(master_path)
    for player, predictions in batch.items():
        if player not in wb.sheetnames:
            wb.close()
            raise SystemExit(f"Hoja '{player}' no existe en {master_path.name}")
        ws = wb[player]
        for mid, (gh, ga, cl) in predictions.items():
            if not (first_id <= mid <= last_id):
                continue
            row = PT_FIRST_ROW + mid - 1
            ws.cell(row, 6, value=gh)
            ws.cell(row, 7, value=ga)
            if cl:
                ws.cell(row, 8, value=cl)
    for attempt in range(5):
        try:
            wb.save(master_path)
            break
        except PermissionError:
            if attempt == 4:
                wb.close()
                raise
            time.sleep(2)
    wb.close()


def run_imports(phase: str, imports: list[tuple[str, str, str, str, str]], dry_run: bool) -> int:
    if phase not in PHASE_RANGES:
        raise SystemExit(f"Fase desconocida: {phase}. Usa: {', '.join(PHASE_RANGES)}")
    first_id, last_id, phase_token = PHASE_RANGES[phase]
    total_expected = last_id - first_id + 1
    require_all = phase in ("dieciseisavos", "octavos")

    groups_cache: dict[str, Path] = {}
    master_pairs_cache: dict[str, dict[int, tuple[str, str]]] = {}
    snapshots_before: dict[tuple[str, str], dict] = {}
    pending_writes: dict[str, dict[str, dict[int, tuple[int, int, str | None]]]] = {}
    errors = 0

    for source, group_id, player, source_sheet, label in imports:
        if not Path(source).exists():
            print(f"\n[{label}] ERROR: no existe {source}")
            errors += 1
            continue

        if group_id not in groups_cache:
            g = load_group(group_id)
            groups_cache[group_id] = g["excel_path"]
            master_pairs_cache[group_id] = load_master_pairs(
                g["excel_path"], first_id, last_id
            )

        master_path = groups_cache[group_id]
        master_pairs = master_pairs_cache[group_id]
        if not master_pairs:
            print(f"\n[{label}] ERROR: master sin partidos resueltos para {phase}")
            errors += 1
            continue

        pair_index = build_pair_index(master_pairs)

        key = (str(master_path), player)
        if not dry_run and key not in snapshots_before:
            snapshots_before[key] = snapshot_protected_cells(master_path, player)

        result = parse_source(
            Path(source),
            master_pairs,
            pair_index,
            label,
            phase_token,
            source_sheet,
            first_id,
            last_id,
        )
        result.player = player

        print(f"\n{'=' * 60}")
        print(f"{label} -> {player} ({master_path.name})")
        print(f"Emparejados: {len(result.matched)}/{len(master_pairs)} (master resueltos)")
        for flag in result.flags:
            print(f"  {flag}")
        for mid in sorted(result.matched):
            gh, ga, cl = result.matched[mid]
            ml, mv = master_pairs[mid]
            cl_s = cl or "?"
            print(f"  {mid:2d} {ml} vs {mv}: {gh}-{ga}  clasif={cl_s}")

        if not result.matched:
            errors += 1
            print("  *** SIN EMPAREJAMIENTOS — no se escribe ***")
            continue

        if require_all and len(result.matched) < total_expected:
            errors += 1
            print(f"  *** INCOMPLETO ({len(result.matched)}/{total_expected}) — no se escribe ***")
            continue

        if any(cl is None for _gh, _ga, cl in result.matched.values()):
            errors += 1
            print("  *** HAY EMPATES SIN CLASIFICADO — no se escribe ***")
            continue

        if dry_run:
            print("  (dry-run: sin escribir)")
        else:
            master_key = str(master_path)
            pending_writes.setdefault(master_key, {})[player] = result.matched
            print(f"  Pendiente escribir ({len(result.matched)} partidos)")

    if not dry_run and pending_writes:
        for master_key, batch in pending_writes.items():
            master_path = Path(master_key)
            print(f"\nGuardando {master_path.name} ({len(batch)} jugadores)...")
            write_ko_batch_to_master(master_path, batch, first_id, last_id)
            for player in batch:
                print(f"  {player}: Escrito OK")

    if not dry_run and snapshots_before:
        print(f"\n{'=' * 60}")
        print(f"Validación filas 4-{PROTECT_LAST_ROW} sin cambios en F,G,H:")
        for (master_str, player), before in snapshots_before.items():
            after = snapshot_protected_cells(Path(master_str), player)
            changed = [k for k, val in before.items() if after.get(k) != val]
            if changed:
                print(f"  {player}: ERROR — cambió en {len(changed)} celdas protegidas")
                errors += 1
            else:
                print(f"  {player}: OK")

    return errors


def main() -> None:
    parser = argparse.ArgumentParser(description="Importar pronósticos KO desde porras de amigos")
    parser.add_argument(
        "--phase",
        default="octavos",
        choices=list(PHASE_RANGES),
        help="Fase a importar (default: octavos)",
    )
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    imports_by_phase = {
        "octavos": IMPORTS_OCTAVOS,
        "cuartos": IMPORTS_CUARTOS,
    }
    imports = imports_by_phase.get(args.phase, [])
    if not imports:
        print(f"No hay lista IMPORTS definida para fase '{args.phase}'.")
        sys.exit(1)

    mode = "DRY-RUN" if args.dry_run else "IMPORT"
    print(f"Import KO predictions [{mode}] fase={args.phase}")
    errors = run_imports(args.phase, imports, args.dry_run)
    if errors:
        print(f"\nTerminado con {errors} error(es).")
        sys.exit(1)
    print("\nTerminado OK.")


if __name__ == "__main__":
    main()
