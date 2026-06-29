"""
Importa pronósticos de dieciseisavos (partidos 73-88) desde porras de amigos
al Excel maestro. Empareja por pareja de equipos, no por ID de partido.
"""
from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from config.groups import load_group
from worldcup_data import normalize_name, team_names_match

ROOT = Path(__file__).parent
PT_FIRST_ROW = 4
R32_FIRST_ID = 73
R32_LAST_ID = 88

# Alias extra para nombres usados en porras de amigos (no siempre en team_names_match)
EXTRA_ALIASES: dict[str, set[str]] = {
    normalize_name("Estados Unidos"): {
        normalize_name("EEUU"),
        normalize_name("EE.UU"),
        normalize_name("EE.UU."),
        normalize_name("USA"),
        normalize_name("United States"),
    },
    normalize_name("Países Bajos"): {
        normalize_name("Holanda"),
        normalize_name("Paises Bajos"),
        normalize_name("Netherlands"),
    },
    normalize_name("RD Congo"): {
        normalize_name("Congo"),
        normalize_name("RD congo"),
        normalize_name("Congo DR"),
    },
    normalize_name("Bosnia y Herzegovina"): {
        normalize_name("Bosnia"),
        normalize_name("Bosnia y herzegovina"),
    },
    normalize_name("Inglaterra"): {
        normalize_name("inglatera"),
        normalize_name("Inlgaterra"),
    },
    normalize_name("Sudáfrica"): {
        normalize_name("Sudafrica"),
    },
    normalize_name("Canadá"): {
        normalize_name("Canada"),
    },
    normalize_name("México"): {
        normalize_name("Mexico"),
    },
    normalize_name("España"): {
        normalize_name("Espana"),
    },
    normalize_name("Japón"): {
        normalize_name("Japon"),
    },
    normalize_name("Bélgica"): {
        normalize_name("Belgica"),
    },
}

IMPORTS: list[tuple[str, str, str, str]] = [
    # (source_path, group_id, player_sheet, label)
    (
        r"c:\Users\Alvaro J Perez Triay\AppData\Local\Temp\Porra_Mundial_2026_Simontorreglosa2000@gmail.com.xlsx",
        "broshu",
        "Simón",
        "Simon",
    ),
    (
        r"c:\Users\Alvaro J Perez Triay\AppData\Local\Temp\Porra_Mundial_2026_Correo 3.xlsx",
        "broshu",
        "Fer",
        "Fer",
    ),
    (
        r"c:\Users\Alvaro J Perez Triay\AppData\Local\Temp\PORRA MUNDIAL NACH.xlsx",
        "broshu",
        "Nacho",
        "Nacho",
    ),
    (
        r"c:\Users\Alvaro J Perez Triay\AppData\Local\Temp\Porra_Mundial_2026_Correo.xlsx",
        "broshu",
        "Luis",
        "Luis",
    ),
    (
        r"c:\Users\Alvaro J Perez Triay\AppData\Local\Temp\Porra_Mundial_2026_Pepe.xlsx",
        "broshu",
        "Pepe",
        "Pepe",
    ),
    (
        r"c:\Users\Alvaro J Perez Triay\AppData\Local\Temp\Porra_Mundial_2026_Correo_Quintero-16AVOS.xlsx",
        "broshu",
        "Quintero",
        "Quintero",
    ),
    (
        r"c:\Users\Alvaro J Perez Triay\AppData\Local\Temp\Porra_Mundial_2026_Correo (1) (1) (1).xlsx",
        "broshu",
        "Felipe",
        "Felipe",
    ),
    (
        r"c:\Users\Alvaro J Perez Triay\AppData\Local\Temp\Porra_Mundial_2026_Patricio.xlsx",
        "broshu",
        "Patri",
        "Patricio",
    ),
    (
        r"c:\Users\Alvaro J Perez Triay\AppData\Local\Temp\PORRA KIKOTA.xlsx",
        "broshu",
        "Kike",
        "Kike",
    ),
    (
        r"c:\Users\Alvaro J Perez Triay\AppData\Local\Temp\Porra_Mundial_2026_papi.xlsx",
        "papinenes",
        "Papá",
        "Papa",
    ),
    (
        r"c:\Users\Alvaro J Perez Triay\AppData\Local\Temp\Porra_Mundial_2026_Diego.xlsx",
        "papinenes",
        "Diego",
        "Diego",
    ),
]

SKIP_SHEETS = {
    "Partidos", "Resumen", "Pronosticos", "Puntuacion", "Instrucciones",
    "_Helpers", "_Lists", "Portada", "Inicio",
}

COLS_NEW = {"id": 1, "fase": 3, "local": 4, "visit": 5, "gh": 6, "ga": 7, "cl": 8}
COLS_OLD = {"id": 1, "fase": 2, "local": 3, "visit": 4, "gh": 5, "ga": 6, "cl": 7}


@dataclass
class R32Row:
    row: int
    file_id: int | None
    local: str
    visit: str
    gh: int | None
    ga: int | None
    clasif_raw: object


@dataclass
class ImportResult:
    label: str
    player: str
    matched: dict[int, tuple[int, int, str | None]] = field(default_factory=dict)
    flags: list[str] = field(default_factory=list)


def _canonical_norm(name: str) -> str:
    """Normaliza un nombre de equipo a una clave canónica para emparejar parejas."""
    n = normalize_name(name)
    for canonical, alts in EXTRA_ALIASES.items():
        if n in alts or n == canonical:
            return canonical
    # También comprobar contra team_names_match con el propio nombre
    return n


def teams_match(a: str, b: str) -> bool:
    if _canonical_norm(a) == _canonical_norm(b):
        return True
    if team_names_match(a, b):
        return True
    na, nb = _canonical_norm(a), _canonical_norm(b)
    return na == nb


def pair_key(local: str, visit: str) -> tuple[str, str]:
    a, b = _canonical_norm(local), _canonical_norm(visit)
    return tuple(sorted([a, b]))


def detect_layout(ws) -> dict[str, int]:
    h3 = str(ws.cell(3, 2).value or "")
    return COLS_NEW if "Fecha" in h3 else COLS_OLD


def find_source_sheet(wb):
    for name in wb.sheetnames:
        if name in SKIP_SHEETS:
            continue
        ws = wb[name]
        if ws.max_row and ws.max_row >= 80:
            return ws
    for name in wb.sheetnames:
        if name not in SKIP_SHEETS:
            return wb[name]
    return wb[wb.sheetnames[0]]


def load_master_r32(master_path: Path) -> dict[int, tuple[str, str]]:
    wb = load_workbook(master_path, data_only=True)
    pt = wb["Partidos"]
    pairs: dict[int, tuple[str, str]] = {}
    for mid in range(R32_FIRST_ID, R32_LAST_ID + 1):
        row = PT_FIRST_ROW + mid - 1
        loc = pt.cell(row, 4).value
        vis = pt.cell(row, 5).value
        if loc and vis:
            pairs[mid] = (str(loc), str(vis))
    wb.close()
    return pairs


def build_pair_index(master_pairs: dict[int, tuple[str, str]]) -> dict[tuple[str, str], int]:
    return {pair_key(loc, vis): mid for mid, (loc, vis) in master_pairs.items()}


def extract_r32_rows(ws, cols: dict[str, int]) -> list[R32Row]:
    rows: list[R32Row] = []
    for r in range(PT_FIRST_ROW, ws.max_row + 1):
        fase = ws.cell(r, cols["fase"]).value
        if not fase or "Dieci" not in str(fase):
            continue
        local = ws.cell(r, cols["local"]).value
        visit = ws.cell(r, cols["visit"]).value
        if not local or not visit:
            continue
        raw_id = ws.cell(r, cols["id"]).value
        file_id = None
        if raw_id is not None:
            try:
                file_id = int(raw_id)
            except (TypeError, ValueError):
                pass
        gh = _as_int(ws.cell(r, cols["gh"]).value)
        ga = _as_int(ws.cell(r, cols["ga"]).value)
        cl = ws.cell(r, cols["cl"]).value
        rows.append(
            R32Row(
                row=r,
                file_id=file_id,
                local=str(local).strip(),
                visit=str(visit).strip(),
                gh=gh,
                ga=ga,
                clasif_raw=cl,
            )
        )
    return rows


def _as_int(v) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def normalize_clasificado(
    raw: object,
    master_local: str,
    master_visit: str,
    gh: int | None,
    ga: int | None,
) -> tuple[str | None, str | None]:
    """Returns (clasificado, flag_or_none)."""
    if raw in (1, 1.0, "1"):
        return "Local", None
    if raw in (2, 2.0, "2"):
        return "Visitante", None

    if raw not in (None, "", "-"):
        s = str(raw).strip()
        sl = s.lower()
        if sl == "local":
            return "Local", None
        if sl == "visitante":
            return "Visitante", None
        if teams_match(s, master_local):
            return "Local", f"clasif nombre '{s}' -> Local"
        if teams_match(s, master_visit):
            return "Visitante", f"clasif nombre '{s}' -> Visitante"
        # typo en clasificado: inferir por goles si no es empate
        if gh is not None and ga is not None and gh != ga:
            side = "Local" if gh > ga else "Visitante"
            return side, f"clasif no reconocido '{s}' -> inferido {side} por marcador"

    if gh is not None and ga is not None:
        if gh > ga:
            return "Local", None
        if ga > gh:
            return "Visitante", None
        return None, f"empate {gh}-{ga} sin clasificado"

    return None, "goles incompletos"


def parse_source(
    source_path: Path,
    master_pairs: dict[int, tuple[str, str]],
    pair_index: dict[tuple[str, str], int],
    label: str,
) -> ImportResult:
    result = ImportResult(label=label, player="")
    if not source_path.exists():
        result.flags.append(f"ARCHIVO NO ENCONTRADO: {source_path}")
        return result

    wb = load_workbook(source_path, data_only=True)
    ws = find_source_sheet(wb)
    cols = detect_layout(ws)
    layout = "new" if cols is COLS_NEW else "old"
    result.flags.append(f"hoja fuente: {ws.title} (layout {layout})")

    for row in extract_r32_rows(ws, cols):
        pk = pair_key(row.local, row.visit)
        mid = pair_index.get(pk)
        if mid is None:
            result.flags.append(
                f"NO MATCH fila {row.row}: {row.local} vs {row.visit} ({row.gh}-{row.ga})"
            )
            continue

        ml, mv = master_pairs[mid]
        swapped = not (teams_match(row.local, ml) and teams_match(row.visit, mv))
        gh, ga = row.gh, row.ga
        if swapped:
            gh, ga = row.ga, row.gh
            result.flags.append(
                f"pareja invertida fila {row.row}: {row.local}-{row.visit} -> master {mid}"
            )

        if row.file_id is not None and row.file_id != mid:
            result.flags.append(f"ID desfasado fila {row.row}: archivo {row.file_id} -> master {mid}")

        if gh is None or ga is None:
            result.flags.append(f"INCOMPLETO master {mid}: goles {gh}-{ga}")
            continue

        cl, cl_flag = normalize_clasificado(row.clasif_raw, ml, mv, gh, ga)
        if cl_flag:
            result.flags.append(f"master {mid}: {cl_flag}")
        if cl is None:
            result.flags.append(f"REVISAR master {mid}: empate sin clasificado ({gh}-{ga})")

        result.matched[mid] = (gh, ga, cl)

    wb.close()
    missing = [m for m in master_pairs if m not in result.matched]
    if missing:
        result.flags.append(f"FALTAN partidos: {missing}")
    return result


def snapshot_r32_cells(wb_path: Path, player: str) -> dict[tuple[int, int], object]:
    wb = load_workbook(wb_path, data_only=True)
    if player not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb[player]
    snap: dict[tuple[int, int], object] = {}
    for mid in range(R32_FIRST_ID, R32_LAST_ID + 1):
        row = PT_FIRST_ROW + mid - 1
        for col in (6, 7, 8):
            snap[(row, col)] = ws.cell(row, col).value
    for row in range(PT_FIRST_ROW, PT_FIRST_ROW + 72):
        for col in (6, 7, 8):
            snap[("grp", row, col)] = ws.cell(row, col).value
    wb.close()
    return snap


def write_r32_to_master(
    master_path: Path,
    player: str,
    predictions: dict[int, tuple[int, int, str | None]],
    dry_run: bool,
) -> None:
    if dry_run:
        return
    wb = load_workbook(master_path)
    if player not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"Hoja '{player}' no existe en {master_path.name}")
    ws = wb[player]
    for mid, (gh, ga, cl) in predictions.items():
        row = PT_FIRST_ROW + mid - 1
        ws.cell(row, 6, value=gh)
        ws.cell(row, 7, value=ga)
        if cl:
            ws.cell(row, 8, value=cl)
    wb.save(master_path)
    wb.close()


def run_imports(dry_run: bool) -> int:
    groups_cache: dict[str, Path] = {}
    master_pairs_cache: dict[str, dict[int, tuple[str, str]]] = {}
    errors = 0

    # Agrupar por master para snapshot previo
    snapshots_before: dict[tuple[str, str], dict] = {}

    for source, group_id, player, label in IMPORTS:
        if not Path(source).exists():
            print(f"\n[{label}] ERROR: no existe {source}")
            errors += 1
            continue

        if group_id not in groups_cache:
            g = load_group(group_id)
            groups_cache[group_id] = g["excel_path"]
            master_pairs_cache[group_id] = load_master_r32(g["excel_path"])

        master_path = groups_cache[group_id]
        master_pairs = master_pairs_cache[group_id]
        pair_index = build_pair_index(master_pairs)

        key = (str(master_path), player)
        if not dry_run and key not in snapshots_before:
            snapshots_before[key] = snapshot_r32_cells(master_path, player)

        result = parse_source(Path(source), master_pairs, pair_index, label)
        result.player = player

        print(f"\n{'=' * 60}")
        print(f"{label} -> {player} ({master_path.name})")
        print(f"Emparejados: {len(result.matched)}/16")
        for flag in result.flags:
            print(f"  {flag}")
        for mid in sorted(result.matched):
            gh, ga, cl = result.matched[mid]
            ml, mv = master_pairs[mid]
            cl_s = cl or "?"
            print(f"  {mid:2d} {ml} vs {mv}: {gh}-{ga}  clasif={cl_s}")

        if len(result.matched) < 16:
            errors += 1
            print("  *** INCOMPLETO — no se escribe ***")
            continue

        if any(cl is None for _gh, _ga, cl in result.matched.values()):
            errors += 1
            print("  *** HAY EMPATES SIN CLASIFICADO — no se escribe ***")
            continue

        write_r32_to_master(master_path, player, result.matched, dry_run)
        if dry_run:
            print("  (dry-run: sin escribir)")
        else:
            print("  Escrito OK")

    if not dry_run and snapshots_before:
        print(f"\n{'=' * 60}")
        print("Validación grupos (filas 4-75 sin cambios en F,G,H):")
        for (master_str, player), before in snapshots_before.items():
            after = snapshot_r32_cells(Path(master_str), player)
            changed = []
            for key, val in before.items():
                if isinstance(key, tuple) and key[0] == "grp":
                    if after.get(key) != val:
                        changed.append(key)
            if changed:
                print(f"  {player}: ERROR — cambió fase grupos en {len(changed)} celdas")
                errors += 1
            else:
                print(f"  {player}: OK")

    return errors


def main() -> None:
    parser = argparse.ArgumentParser(description="Importar dieciseisavos desde porras de amigos")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Solo informe, no escribe en Excel maestro",
    )
    args = parser.parse_args()
    mode = "DRY-RUN" if args.dry_run else "IMPORT"
    print(f"Import R32 predictions [{mode}]")
    errors = run_imports(args.dry_run)
    if errors:
        print(f"\nTerminado con {errors} error(es).")
        sys.exit(1)
    print("\nTerminado OK.")


if __name__ == "__main__":
    main()
