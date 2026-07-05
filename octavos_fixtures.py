"""Equipos canónicos de octavos (openfootball num == id calendario)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from config.groups import list_groups, load_group
from worldcup_data import load_calendar, save_json, MATCHES_JSON

PT_FIRST_ROW = 4
OCTAVOS_FIRST_ID = 89
OCTAVOS_LAST_ID = 96

# Id calendario → (local, visitante) en español
OCTAVOS_CANONICAL: dict[int, tuple[str, str]] = {
    89: ("Paraguay", "Francia"),
    90: ("Canadá", "Marruecos"),
    91: ("Brasil", "Noruega"),
    92: ("México", "Inglaterra"),
    93: ("Portugal", "España"),
    94: ("Estados Unidos", "Bélgica"),
    95: ("Argentina", "Egipto"),
    96: ("Suiza", "Colombia"),
}

# Id plantilla Excel amigos → id calendario (swap 04/07)
OCTAVOS_TEMPLATE_TO_CALENDAR: dict[int, int] = {
    89: 90,
    90: 89,
    **{i: i for i in range(91, 97)},
}

# Plantilla amigos (id fila Excel) → pareja predicha
OCTAVOS_TEMPLATE_PAIRS: dict[int, tuple[str, str]] = {
    89: ("Canada", "Marruecos"),
    90: ("Paraguay", "Francia"),
    91: ("Brasil", "Noruega"),
    92: ("México", "Inglaterra"),
    93: ("Portugal", "España"),
    94: ("Estados Unidos", "Bélgica"),
    95: ("Argentina", "Egipto"),
    96: ("Suiza", "Colombia"),
}

# Orden lista usuario → id calendario (para entrada manual)
OCTAVOS_LIST_ORDER: list[tuple[int, str, str]] = [
    (90, "Canadá", "Marruecos"),
    (89, "Paraguay", "Francia"),
    (91, "Brasil", "Noruega"),
    (92, "México", "Inglaterra"),
    (93, "Portugal", "España"),
    (94, "Estados Unidos", "Bélgica"),
    (95, "Argentina", "Egipto"),
    (96, "Suiza", "Colombia"),
]


def template_id_to_calendar(template_id: int) -> int:
    return OCTAVOS_TEMPLATE_TO_CALENDAR.get(template_id, template_id)


def apply_octavos_to_json(dry_run: bool = False) -> list[int]:
    calendar = load_calendar()
    changed: list[int] = []
    cal_by_id = {m["id"]: m for m in calendar}
    for mid, (local, visitante) in OCTAVOS_CANONICAL.items():
        m = cal_by_id.get(mid)
        if not m:
            continue
        if m["local"] == local and m["visitante"] == visitante:
            continue
        print(f"  JSON #{mid}: {m['local']} vs {m['visitante']} -> {local} vs {visitante}")
        m["local"] = local
        m["visitante"] = visitante
        changed.append(mid)
    if changed and not dry_run:
        save_json(MATCHES_JSON, calendar)
    return changed


def apply_octavos_to_excel(excel_path: Path, dry_run: bool = False) -> list[int]:
    wb = load_workbook(excel_path)
    pt = wb["Partidos"]
    changed: list[int] = []
    for mid, (local, visitante) in OCTAVOS_CANONICAL.items():
        row = PT_FIRST_ROW + mid - 1
        old_l, old_v = pt.cell(row, 4).value, pt.cell(row, 5).value
        if old_l == local and old_v == visitante:
            continue
        print(f"  #{mid}: {old_l} vs {old_v} -> {local} vs {visitante}")
        if not dry_run:
            pt.cell(row, 4, value=local)
            pt.cell(row, 5, value=visitante)
        changed.append(mid)
    if changed and not dry_run:
        wb.save(excel_path)
    wb.close()
    return changed


def apply_octavos_canonical_teams(dry_run: bool = False) -> list[int]:
    """Fija equipos octavos 89-96 en JSON y todos los Excel maestros."""
    print("Aplicando equipos canónicos octavos (89-96)...")
    all_changed: set[int] = set()
    print("matches_2026.json:")
    all_changed.update(apply_octavos_to_json(dry_run))
    for group_id in list_groups():
        excel_path = load_group(group_id)["excel_path"]
        if not excel_path.exists():
            continue
        print(f"Excel {group_id} ({excel_path.name}):")
        all_changed.update(apply_octavos_to_excel(excel_path, dry_run))
    return sorted(all_changed)
