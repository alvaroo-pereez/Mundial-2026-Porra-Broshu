"""Propaga ganadores del cuadro → nombres reales en matches_2026.json y Excel."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

from config.groups import list_groups, load_group
from ko_bracket import compute_resolved_teams, is_ko_placeholder
from cuartos_fixtures import CUARTOS_CANONICAL, apply_cuartos_canonical_teams
from octavos_fixtures import OCTAVOS_CANONICAL, apply_octavos_canonical_teams
from worldcup_data import load_calendar, save_json, MATCHES_JSON

PT_FIRST_ROW = 4


def read_results_from_excel(excel_path: Path) -> dict[int, dict]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    pt = wb["Partidos"]
    results: dict[int, dict] = {}
    for row in range(PT_FIRST_ROW, pt.max_row + 1):
        mid = pt.cell(row, 1).value
        if mid is None:
            break
        try:
            mid = int(mid)
        except (TypeError, ValueError):
            continue
        home, away = pt.cell(row, 6).value, pt.cell(row, 7).value
        if home is None or away is None:
            continue
        entry: dict = {"home": int(home), "away": int(away)}
        cl = pt.cell(row, 11).value
        if cl in ("Local", "Visitante"):
            entry["clasificado"] = cl
        results[mid] = entry
    wb.close()
    return results


def apply_teams_to_excel(excel_path: Path, teams: dict[int, tuple[str, str]], dry_run: bool) -> list[int]:
    wb = load_workbook(excel_path)
    pt = wb["Partidos"]
    changed: list[int] = []
    for mid, (local, visitante) in sorted(teams.items()):
        row = PT_FIRST_ROW + mid - 1
        old_l, old_v = pt.cell(row, 4).value, pt.cell(row, 5).value
        if old_l == local and old_v == visitante:
            continue
        if mid in OCTAVOS_CANONICAL or mid in CUARTOS_CANONICAL:
            continue
        cal_local = str(old_l or "")
        if not is_ko_placeholder(cal_local) and mid < 97:
            continue
        print(f"  #{mid}: {old_l} vs {old_v} -> {local} vs {visitante}")
        if not dry_run:
            pt.cell(row, 4, value=local)
            pt.cell(row, 5, value=visitante)
        changed.append(mid)
    if changed and not dry_run:
        wb.save(excel_path)
    wb.close()
    return changed


def apply_teams_to_json(teams: dict[int, tuple[str, str]], dry_run: bool) -> list[int]:
    calendar = load_calendar()
    changed: list[int] = []
    cal_by_id = {m["id"]: m for m in calendar}
    for mid, (local, visitante) in sorted(teams.items()):
        m = cal_by_id.get(mid)
        if not m:
            continue
        if m["local"] == local and m["visitante"] == visitante:
            continue
        if mid in OCTAVOS_CANONICAL or mid in CUARTOS_CANONICAL:
            continue
        if not is_ko_placeholder(m["local"]) and mid < 97:
            continue
        print(f"  JSON #{mid}: {m['local']} vs {m['visitante']} -> {local} vs {visitante}")
        m["local"] = local
        m["visitante"] = visitante
        changed.append(mid)
    if changed and not dry_run:
        save_json(MATCHES_JSON, calendar)
    return changed


def propagate_ko_teams(dry_run: bool = False) -> list[int]:
    """Propaga equipos usando resultados del Excel del primer grupo disponible."""
    oct_changed = apply_octavos_canonical_teams(dry_run=dry_run)
    cuart_changed = apply_cuartos_canonical_teams(dry_run=dry_run)
    calendar = load_calendar()
    source_excel: Path | None = None
    for group_id in list_groups():
        path = load_group(group_id)["excel_path"]
        if path.exists():
            source_excel = path
            break
    if source_excel is None:
        print("No hay Excel maestro en output/.")
        return sorted(set(oct_changed) | set(cuart_changed))

    print(f"Resultados fuente: {source_excel.name}")
    results = read_results_from_excel(source_excel)
    resolved = compute_resolved_teams(results, calendar)
    all_changed: set[int] = set(oct_changed) | set(cuart_changed)
    if not resolved:
        print("Sin equipos nuevos que propagar (cuartos+).")
        return sorted(all_changed)

    print("Actualizando matches_2026.json (cuartos+):")
    json_changed = apply_teams_to_json(resolved, dry_run)
    all_changed.update(json_changed)
    for group_id in list_groups():
        excel_path = load_group(group_id)["excel_path"]
        if not excel_path.exists():
            continue
        print(f"Excel {group_id} ({excel_path.name}):")
        excel_changed = apply_teams_to_excel(excel_path, resolved, dry_run)
        all_changed.update(excel_changed)

    return sorted(all_changed)


def main() -> None:
    parser = argparse.ArgumentParser(description="Propaga equipos KO en JSON y Excel")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    changed = propagate_ko_teams(dry_run=args.dry_run)
    if args.dry_run:
        print(f"(dry-run: {len(changed)} partidos cambiarían)")
    else:
        print(f"Listo: {len(changed)} partidos actualizados.")
    sys.exit(0)


if __name__ == "__main__":
    main()
