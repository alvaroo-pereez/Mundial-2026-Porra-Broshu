"""
Reordena dieciseisavos (IDs 73–88) en Excel según matches_2026.json cronológico.
Migra resultados y pronósticos por pareja de equipos, no por ID antiguo.
"""
from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from config.groups import list_groups, load_group
from import_r32_predictions import EXTRA_ALIASES, _canonical_norm
from worldcup_data import normalize_name, team_names_match

ROOT = Path(__file__).parent
MATCHES_JSON = ROOT / "data" / "matches_2026.json"
PT_FIRST_ROW = 4
R32_FIRST = 73
R32_LAST = 88


@dataclass
class PartidosSnap:
    res_h: int | None
    res_a: int | None
    clasificado: object


@dataclass
class PredSnap:
    gh: int | None
    ga: int | None
    clasif: object


def pair_key(local: str, visit: str) -> tuple[str, str]:
    a, b = _canonical_norm(str(local)), _canonical_norm(str(visit))
    return tuple(sorted([a, b]))


def teams_match(a: str, b: str) -> bool:
    if _canonical_norm(a) == _canonical_norm(b):
        return True
    return team_names_match(a, b)


def load_r32_calendar() -> list[dict]:
    matches = json.loads(MATCHES_JSON.read_text(encoding="utf-8"))
    r32 = [m for m in matches if R32_FIRST <= m["id"] <= R32_LAST]
    if len(r32) != 16:
        raise SystemExit(f"Se esperaban 16 R32 en JSON, hay {len(r32)}")
    return sorted(r32, key=lambda m: m["id"])


def snapshot_workbook(wb, players: list[str]) -> tuple[dict, dict]:
    pt = wb["Partidos"]
    partidos: dict[tuple[str, str], PartidosSnap] = {}
    preds: dict[tuple[str, str], dict[str, PredSnap]] = {}

    for mid in range(R32_FIRST, R32_LAST + 1):
        row = PT_FIRST_ROW + mid - 1
        local = pt.cell(row, 4).value
        visit = pt.cell(row, 5).value
        if not local or not visit:
            continue
        pk = pair_key(str(local), str(visit))
        partidos[pk] = PartidosSnap(
            res_h=_as_int(pt.cell(row, 6).value),
            res_a=_as_int(pt.cell(row, 7).value),
            clasificado=pt.cell(row, 11).value,
        )
        preds[pk] = {}
        for player in players:
            if player not in wb.sheetnames:
                continue
            ws = wb[player]
            preds[pk][player] = PredSnap(
                gh=_as_int(ws.cell(row, 6).value),
                ga=_as_int(ws.cell(row, 7).value),
                clasif=ws.cell(row, 8).value,
            )
    return partidos, preds


def _as_int(v) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def clear_r32_data(wb, players: list[str]) -> None:
    pt = wb["Partidos"]
    for mid in range(R32_FIRST, R32_LAST + 1):
        row = PT_FIRST_ROW + mid - 1
        for col in (6, 7):
            pt.cell(row, col, value=None)
        pt.cell(row, 11, value=None)
        for player in players:
            if player not in wb.sheetnames:
                continue
            ws = wb[player]
            for col in (6, 7, 8):
                ws.cell(row, col, value=None)


def apply_r32(wb, players: list[str], calendar: list[dict], partidos_snap, preds_snap, dry_run: bool) -> list[str]:
    flags: list[str] = []
    pt = wb["Partidos"]

    for m in calendar:
        mid = m["id"]
        row = PT_FIRST_ROW + mid - 1
        pk = pair_key(m["local"], m["visitante"])
        ps = partidos_snap.get(pk)
        if ps is None:
            flags.append(f"Partidos: sin snapshot para #{mid} {m['local']} vs {m['visitante']}")
        pp = preds_snap.get(pk, {})

        if dry_run:
            continue

        pt.cell(row, 1, value=mid)
        pt.cell(row, 2, value=m["fecha"])
        pt.cell(row, 3, value=m["hora"])
        pt.cell(row, 4, value=m["local"])
        pt.cell(row, 5, value=m["visitante"])
        pt.cell(row, 9, value=m.get("jornada", 13))
        pt.cell(row, 10, value=m.get("fase", "Dieciseisavos"))

        if ps:
            if ps.res_h is not None:
                pt.cell(row, 6, value=ps.res_h)
            if ps.res_a is not None:
                pt.cell(row, 7, value=ps.res_a)
            if ps.clasificado not in (None, ""):
                pt.cell(row, 11, value=ps.clasificado)

        for player in players:
            if player not in wb.sheetnames:
                continue
            ws = wb[player]
            ws.cell(row, 1, value=mid)
            ws.cell(row, 2, value=m["fecha"])
            ws.cell(row, 3, value=m.get("fase", "Dieciseisavos"))
            pred = pp.get(player)
            if pred is None:
                flags.append(f"{player}: sin pronóstico para {m['local']} vs {m['visitante']}")
                continue
            if pred.gh is not None:
                ws.cell(row, 6, value=pred.gh)
            if pred.ga is not None:
                ws.cell(row, 7, value=pred.ga)
            if pred.clasif not in (None, "", "-"):
                ws.cell(row, 8, value=pred.clasif)

    return flags


def verify_group_rows_unchanged(before: dict, after_wb, players: list[str]) -> list[str]:
    errors = []
    for player in players:
        if player not in after_wb.sheetnames:
            continue
        ws = after_wb[player]
        for key, val in before.get(player, {}).items():
            row, col = key
            if ws.cell(row, col).value != val:
                errors.append(f"{player} fila {row} col {col} cambió")
    return errors


def snapshot_group_rows(wb, players: list[str]) -> dict[str, dict]:
    out: dict[str, dict] = {}
    for player in players:
        if player not in wb.sheetnames:
            continue
        ws = wb[player]
        cells = {}
        for row in range(PT_FIRST_ROW, PT_FIRST_ROW + 72):
            for col in (6, 7, 8):
                cells[(row, col)] = ws.cell(row, col).value
        out[player] = cells
    return out


def process_excel(path: Path, group_id: str, dry_run: bool) -> int:
    group = load_group(group_id)
    players = group["players"]
    calendar = load_r32_calendar()

    print(f"\n{path.name} ({group_id})")
    wb = load_workbook(path)
    group_before = snapshot_group_rows(wb, players)
    partidos_snap, preds_snap = snapshot_workbook(wb, players)
    print(f"  Snapshot: {len(partidos_snap)} parejas Partidos, {len(preds_snap)} con pronósticos")

    flags = apply_r32(wb, players, calendar, partidos_snap, preds_snap, dry_run=True)
    for f in flags:
        print(f"  FLAG: {f}")

    missing_pairs = [
        pair_key(m["local"], m["visitante"])
        for m in calendar
        if pair_key(m["local"], m["visitante"]) not in partidos_snap
    ]
    if missing_pairs:
        print(f"  ERROR: faltan {len(missing_pairs)} parejas en snapshot")
        wb.close()
        return 1

    if dry_run:
        print("  (dry-run: sin escribir)")
        print("  Orden nuevo:")
        for m in calendar:
            print(f"    {m['id']:2d} {m['fecha']} {m['hora']} {m['local']} vs {m['visitante']}")
        wb.close()
        return 0

    clear_r32_data(wb, players)
    write_flags = apply_r32(wb, players, calendar, partidos_snap, preds_snap, dry_run=False)
    for f in write_flags:
        print(f"  AVISO: {f}")

    group_errors = verify_group_rows_unchanged(group_before, wb, players)
    if group_errors:
        for e in group_errors:
            print(f"  ERROR: {e}")
        wb.close()
        return 1

    wb.save(path)
    wb.close()
    print("  Escrito OK")
    return 0


def main() -> None:
    parser = argparse.ArgumentParser(description="Reordenar R32 en Excel por calendario cronológico")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    errors = 0
    for gid in list_groups():
        path = load_group(gid)["excel_path"]
        errors += process_excel(path, gid, dry_run=args.dry_run)

    if errors:
        sys.exit(1)
    print("\nListo.")


if __name__ == "__main__":
    main()
