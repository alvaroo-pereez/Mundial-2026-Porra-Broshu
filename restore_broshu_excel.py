"""
Restaura pronósticos y resultados del Excel Broshu desde un dashboard HTML
con datos embebidos (snapshot previo a la regeneración).
"""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).parent
PT_FIRST_ROW = 4
SPECIAL_RESULTS_FIRST_ROW = 34

DASHBOARD_SNAPSHOT = ROOT / "output" / "_recovered_dashboard_broshu.html"
MATCHES_JSON = ROOT / "data" / "matches_2026.json"
TARGET = ROOT / "output" / "Porra_Mundial_2026.xlsx"


def load_dashboard_data(path: Path) -> dict:
    html = path.read_text(encoding="utf-8")
    m = re.search(r"const DATA = (\{.*?\});\s*\n", html, re.DOTALL)
    if not m:
        raise SystemExit(f"No se encontró DATA en {path}")
    return json.loads(m.group(1))


def read_special_categories(ws) -> list[str]:
    cats: list[str] = []
    row = SPECIAL_RESULTS_FIRST_ROW
    while True:
        val = ws.cell(row, 1).value
        if val is None or str(val).strip() == "":
            break
        cats.append(str(val).strip())
        row += 1
    return cats


def restore_workbook(src_xlsx: Path, data: dict) -> None:
    wb = load_workbook(src_xlsx)
    n_matches = len(json.loads(MATCHES_JSON.read_text(encoding="utf-8")))
    last_row = PT_FIRST_ROW + n_matches - 1
    special_first_row = last_row + 4

    # Partidos — resultados reales
    if "Partidos" in wb.sheetnames:
        pt = wb["Partidos"]
        for m in data.get("matches", []):
            mid = m["id"]
            row = PT_FIRST_ROW + mid - 1
            if m.get("res_h") is not None:
                pt.cell(row, 6, value=int(m["res_h"]))
            if m.get("res_a") is not None:
                pt.cell(row, 7, value=int(m["res_a"]))
            if m.get("clasificado"):
                pt.cell(row, 11, value=m["clasificado"])

    # Apuestas especiales oficiales (Resumen)
    specials_official: dict[str, str] = {}
    for s in data.get("home", {}).get("specials_summary", []):
        if s.get("oficial"):
            specials_official[s["categoria"]] = s["oficial"]

    if "Resumen" in wb.sheetnames:
        resumen = wb["Resumen"]
        for idx, cat in enumerate(read_special_categories(resumen)):
            if cat in specials_official:
                resumen.cell(SPECIAL_RESULTS_FIRST_ROW + idx, 2, value=specials_official[cat])

    # Pronósticos por jugador
    for player_name, pdata in data.get("players", {}).items():
        if player_name not in wb.sheetnames:
            print(f"  Aviso: hoja '{player_name}' no existe, se omite")
            continue
        ws = wb[player_name]
        by_id = {m["id"]: m for m in pdata.get("matches", [])}
        for mid in range(1, n_matches + 1):
            row = PT_FIRST_ROW + mid - 1
            m = by_id.get(mid)
            if not m:
                continue
            if m.get("pred_h") is not None:
                ws.cell(row, 6, value=int(m["pred_h"]))
            if m.get("pred_a") is not None:
                ws.cell(row, 7, value=int(m["pred_a"]))
            pc = m.get("pred_clasificado")
            if pc not in (None, ""):
                ws.cell(row, 8, value=pc)

        # Apuestas especiales: mapear por índice (el snapshot HTML tenía
        # etiquetas desordenadas pero los valores en la posición correcta).
        specs = pdata.get("specials", [])
        for idx, spec in enumerate(specs):
            pred = spec.get("prediccion", "")
            if pred:
                ws.cell(special_first_row + idx, 2, value=pred)

    # Portada: quitar subtítulo de grupo añadido en la regeneración
    if "Portada" in wb.sheetnames:
        portada = wb["Portada"]
        for r in range(30, 37):
            val = portada.cell(r, 1).value
            if val and "Grupo Broshu" in str(val):
                portada.cell(
                    r,
                    1,
                    value="Porra privada entre amigos · Pronósticos, clasificación automática y apuestas especiales",
                )
                break

    return wb


def main() -> None:
    if not DASHBOARD_SNAPSHOT.exists():
        raise SystemExit(f"No existe snapshot: {DASHBOARD_SNAPSHOT}")

    if not TARGET.exists():
        raise SystemExit(
            "No existe output/Porra_Mundial_2026.xlsx. "
            "Ejecuta: py generate_porra.py --group broshu --no-open"
        )

    data = load_dashboard_data(DASHBOARD_SNAPSHOT)
    print(f"Snapshot: {data['meta'].get('generated_at', '?')}")

    wb = restore_workbook(TARGET, data)
    wb.save(TARGET)
    wb.close()
    print(f"Restaurado: {TARGET}")

    filled = sum(
        1
        for p in data["players"].values()
        for m in p["matches"]
        if m.get("pred_h") is not None
    )
    print(f"Celdas de pronóstico restauradas (total filas con goles): {filled}")


if __name__ == "__main__":
    main()
