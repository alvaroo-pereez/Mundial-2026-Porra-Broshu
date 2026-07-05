"""Sincroniza resultados openfootball → Excel → data.json del dashboard."""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

from config.groups import list_groups, load_group
from propagate_ko_teams import propagate_ko_teams
from worldcup_data import collect_finished_updates, load_calendar, match_is_ko

ROOT = Path(__file__).parent
PT_FIRST_ROW = 4


def read_excel_result(pt, match_id: int) -> tuple:
    row = PT_FIRST_ROW + match_id - 1
    return pt.cell(row, 6).value, pt.cell(row, 7).value, pt.cell(row, 11).value


def write_excel_result(
    pt,
    match_id: int,
    home: int,
    away: int,
    clasificado: str | None,
    is_ko: bool,
) -> bool:
    row = PT_FIRST_ROW + match_id - 1
    old_h, old_a, old_cl = pt.cell(row, 6).value, pt.cell(row, 7).value, pt.cell(row, 11).value
    new_cl = clasificado if is_ko and clasificado else (old_cl if is_ko else "-")

    changed = (
        old_h != home
        or old_a != away
        or (is_ko and old_cl != new_cl)
    )
    if not changed:
        return False

    pt.cell(row, 6, value=home)
    pt.cell(row, 7, value=away)
    if is_ko and clasificado:
        pt.cell(row, 11, value=clasificado)
    return True


def collect_api_updates() -> dict[int, dict]:
    return collect_finished_updates()


def count_excel_finished(excel_path: Path, total_matches: int) -> int:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    pt = wb["Partidos"]
    count = 0
    for mid in range(1, total_matches + 1):
        row = PT_FIRST_ROW + mid - 1
        if pt.cell(row, 6).value is not None and pt.cell(row, 7).value is not None:
            count += 1
    wb.close()
    return count


def apply_to_excel(excel_path: Path, updates: dict[int, dict], dry_run: bool) -> bool:
    calendar = load_calendar()
    calendar_by_id = {m["id"]: m for m in calendar}
    wb = load_workbook(excel_path)
    pt = wb["Partidos"]
    changed = False

    for mid, result in sorted(updates.items()):
        m = calendar_by_id[mid]
        is_ko = match_is_ko(m)
        old_h, old_a, old_cl = read_excel_result(pt, mid)
        home, away = result["home"], result["away"]
        clasif = result.get("clasificado")

        if dry_run:
            if old_h != home or old_a != away or (is_ko and clasif and old_cl != clasif):
                print(
                    f"  #{mid} {m['local']} vs {m['visitante']}: "
                    f"{old_h}-{old_a} -> {home}-{away}"
                    + (f" clasif={clasif}" if is_ko and clasif else "")
                )
                changed = True
            continue

        if write_excel_result(pt, mid, home, away, clasif, is_ko):
            print(f"  Actualizado #{mid}: {home}-{away}" + (f" ({clasif})" if clasif else ""))
            changed = True

    if changed and not dry_run:
        wb.save(excel_path)
    return changed


def rebuild_dashboards() -> None:
    import build_dashboard

    for group_id in list_groups():
        build_dashboard.main(group_id=group_id, open_browser=False)


def main() -> None:
    dry_run = "--dry-run" in sys.argv
    any_changed = False
    teams_changed = False

    if not dry_run:
        print("Propagando equipos eliminatorias...")
        propagated = propagate_ko_teams(dry_run=False)
        if propagated:
            print(f"  Equipos actualizados: {propagated}")
            teams_changed = True

    print("Consultando openfootball/worldcup.json...")
    updates = collect_api_updates()
    print(f"Partidos finalizados (emparejados): {len(updates)}")

    missing_excels: list[Path] = []
    for group_id in list_groups():
        group = load_group(group_id)
        excel_path = group["excel_path"]
        if not excel_path.exists():
            missing_excels.append(excel_path)
            continue
        print(f"Grupo {group_id} ({excel_path.name}):")
        if apply_to_excel(excel_path, updates, dry_run):
            any_changed = True

    if missing_excels:
        for path in missing_excels:
            print(f"  Error: no existe {path}")
        print("Los Excel deben estar en output/ y subidos al repositorio.")
        sys.exit(1)

    if not dry_run:
        propagated_after = propagate_ko_teams(dry_run=False)
        if propagated_after:
            print(f"Re-propagación tras resultados: {propagated_after}")
            teams_changed = True
            updates_after = collect_finished_updates()
            for group_id in list_groups():
                group = load_group(group_id)
                excel_path = group["excel_path"]
                if excel_path.exists():
                    if apply_to_excel(excel_path, updates_after, dry_run=False):
                        any_changed = True

    if (any_changed or teams_changed) and not dry_run:
        print("Regenerando dashboards...")
        rebuild_dashboards()
        print("Listo.")
    elif not any_changed and not teams_changed:
        total_matches = len(load_calendar())
        for group_id in list_groups():
            group = load_group(group_id)
            excel_path = group["excel_path"]
            if excel_path.exists():
                in_excel = count_excel_finished(excel_path, total_matches)
                print(
                    f"Sin cambios en {group_id}. "
                    f"openfootball: {len(updates)} con resultado | Excel: {in_excel} con resultado"
                )
        if not any(load_group(g)["excel_path"].exists() for g in list_groups()):
            print("Sin cambios.")
    elif dry_run:
        print("(dry-run: no se escribió nada)")

    sys.exit(0)


if __name__ == "__main__":
    main()
