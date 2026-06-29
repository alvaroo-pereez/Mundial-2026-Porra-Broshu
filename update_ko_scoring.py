"""Parchea puntuación KO (empates) en Puntuacion + fórmulas Pronosticos."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

from config.groups import list_groups, load_group
from generate_porra import (
    PT_FIRST_ROW,
    points_formula,
    pt_last_row,
)
from scoring import load_scoring_config

FORMULA_REPLACEMENTS = [
    ("Puntuacion!$B$19", "Puntuacion!$B$22"),
    ("Puntuacion!$B$16", "Puntuacion!$B$19"),
    ("Puntuacion!$B$15", "Puntuacion!$B$18"),
    ("Puntuacion!$B$14", "Puntuacion!$B$17"),
]


def _patch_formula_refs(wb) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith("="):
                    new_val = val
                    for old, new in FORMULA_REPLACEMENTS:
                        new_val = new_val.replace(old, new)
                    if new_val != val:
                        cell.value = new_val


def sync_puntuacion_values(ws) -> bool:
    """Escribe B9–B14 desde config/scoring.json (idempotente)."""
    ko = load_scoring_config()["eliminatorias"]
    mapping = [
        (9, ko["exacto"]),
        (10, ko["diferencia"]),
        (11, ko["clasificado"]),
        (12, ko.get("empate_exacto_falla_clasificado", 4)),
        (13, ko.get("empate_diferencia_acierta_clasificado", 4)),
        (14, ko.get("empate_diferencia_falla_clasificado", 2)),
    ]
    changed = False
    for row, pts in mapping:
        if ws.cell(row, 2).value != pts:
            ws.cell(row, 2, value=pts)
            changed = True
    return changed


def _update_puntuacion(ws) -> bool:
    ko = load_scoring_config()["eliminatorias"]
    if ws.cell(12, 1).value and "Empate" in str(ws.cell(12, 1).value):
        return False
    ws.insert_rows(12, 2)
    rows = [
        ("Empate exacto, fallo clasificado", ko.get("empate_exacto_falla_clasificado", 6)),
        ("Empate dif., acierto clasificado", ko.get("empate_diferencia_acierta_clasificado", 6)),
        ("Empate dif., fallo clasificado", ko.get("empate_diferencia_falla_clasificado", 2)),
    ]
    for i, (label, pts) in enumerate(rows):
        r = 12 + i
        ws.cell(r, 1, value=label)
        ws.cell(r, 2, value=pts)
        ws.cell(r, 3, value="pts")
    return True


def _update_pronosticos(ws, last_row: int) -> int:
    n = 0
    for row in range(PT_FIRST_ROW, ws.max_row + 1):
        mid = ws.cell(row, 1).value
        if mid is None:
            break
        try:
            mid = int(mid)
        except (TypeError, ValueError):
            continue
        if mid < 73:
            continue
        ws.cell(row, 7, value=points_formula(row, last_row))
        n += 1
    return n


def patch_excel(excel_path: Path, dry_run: bool = False) -> None:
    if not excel_path.exists():
        print(f"  No existe: {excel_path}")
        return
    wb = load_workbook(excel_path)
    n_matches = 104
    last_row = pt_last_row(n_matches)

    if "Puntuacion" in wb.sheetnames:
        ws_p = wb["Puntuacion"]
        if _update_puntuacion(ws_p):
            print("  Puntuacion: filas empate añadidas")
            _patch_formula_refs(wb)
            print("  Referencias bonus/apuestas actualizadas")
        if sync_puntuacion_values(ws_p):
            print("  Puntuacion: valores KO sincronizados (B9-B14)")

    if "Pronosticos" in wb.sheetnames:
        n = _update_pronosticos(wb["Pronosticos"], last_row)
        print(f"  Pronosticos: {n} filas KO (partido >= 73) actualizadas")

    if not dry_run:
        wb.save(excel_path)
    wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Parchea puntuación KO en Excel")
    parser.add_argument("excel", nargs="*", help="Rutas .xlsx")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    paths = [Path(p) for p in args.excel] if args.excel else [
        load_group(g)["excel_path"] for g in list_groups()
    ]
    for path in paths:
        print(path.name)
        patch_excel(path, dry_run=args.dry_run)
    print("Listo." if not args.dry_run else "(dry-run)")
    sys.exit(0)


if __name__ == "__main__":
    main()
