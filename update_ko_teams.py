"""Actualiza equipos de dieciseisavos (partidos 73–88) en hoja Partidos."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

from config.groups import list_groups, load_group
from fifa_schedule_2026 import KNOCKOUT_FIXTURES, R32_RESOLVED

PT_FIRST_ROW = 4
R32_FIRST = 73
R32_LAST = 88


def update_partidos(excel_path: Path, dry_run: bool = False) -> int:
    if not excel_path.exists():
        print(f"  No existe: {excel_path}")
        return 0
    teams_by_id = {
        ko["id"]: R32_RESOLVED[i]
        for i, ko in enumerate(KNOCKOUT_FIXTURES)
        if R32_FIRST <= ko["id"] <= R32_LAST
    }
    wb = load_workbook(excel_path)
    if "Partidos" not in wb.sheetnames:
        print(f"  Sin hoja Partidos: {excel_path.name}")
        wb.close()
        return 0
    ws = wb["Partidos"]
    changed = 0
    for row in range(PT_FIRST_ROW, ws.max_row + 1):
        mid = ws.cell(row, 1).value
        if mid is None:
            break
        try:
            mid = int(mid)
        except (TypeError, ValueError):
            continue
        if mid not in teams_by_id:
            continue
        local, visitante = teams_by_id[mid]
        old_l, old_v = ws.cell(row, 4).value, ws.cell(row, 5).value
        if old_l != local or old_v != visitante:
            print(f"  #{mid}: {old_l} vs {old_v} -> {local} vs {visitante}")
            if not dry_run:
                ws.cell(row, 4, value=local)
                ws.cell(row, 5, value=visitante)
            changed += 1
    if changed and not dry_run:
        wb.save(excel_path)
    wb.close()
    return changed


def main() -> None:
    parser = argparse.ArgumentParser(description="Actualiza equipos R32 en Excel")
    parser.add_argument("excel", nargs="*", help="Rutas .xlsx (por defecto output/ de cada grupo)")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    paths: list[Path] = [Path(p) for p in args.excel] if args.excel else []
    if not paths:
        paths = [load_group(g)["excel_path"] for g in list_groups()]

    total = 0
    for path in paths:
        print(path.name)
        total += update_partidos(path, dry_run=args.dry_run)
    if args.dry_run:
        print(f"(dry-run: {total} partidos cambiarían)")
    else:
        print(f"Listo: {total} partidos actualizados.")
    sys.exit(0)


if __name__ == "__main__":
    main()
