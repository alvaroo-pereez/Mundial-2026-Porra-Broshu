"""Comprueba que el pipeline de automatización está listo para ejecutarse."""
from __future__ import annotations

import argparse
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from api_football import FIXTURE_MAP_PATH, fetch_all_fixtures, get_api_key, load_json
from config.groups import list_groups, load_group

ROOT = Path(__file__).parent
TOURNAMENT_START = date(2026, 6, 11)
PT_FIRST_ROW = 4


def _log(level: str, message: str) -> None:
    print(f"[{level}] {message}")


def check_excel_files() -> list[str]:
    errors: list[str] = []
    for group_id in list_groups():
        group = load_group(group_id)
        excel_path = group["excel_path"]
        if not excel_path.exists():
            errors.append(f"Falta Excel del grupo {group_id}: {excel_path}")
            continue
        try:
            wb = load_workbook(excel_path, read_only=True, data_only=True)
            if "Partidos" not in wb.sheetnames:
                errors.append(f"{excel_path.name}: no tiene hoja 'Partidos'")
            wb.close()
        except Exception as exc:
            errors.append(f"No se puede leer {excel_path.name}: {exc}")
    return errors


def check_fixture_map(ci_mode: bool) -> list[str]:
    issues: list[str] = []
    if not FIXTURE_MAP_PATH.exists():
        issues.append(f"Falta {FIXTURE_MAP_PATH.relative_to(ROOT)}")
        return issues

    raw = load_json(FIXTURE_MAP_PATH)
    if not isinstance(raw, dict):
        issues.append("api_fixture_map.json no es un objeto JSON")
        return issues

    if raw:
        _log("OK", f"Mapeo API: {len(raw)} partidos")
        return issues

    msg = (
        "MAPA VACÍO: ejecuta py bootstrap_fixture_map.py "
        "(requiere fixtures WC 2026 en API-Football)"
    )
    if ci_mode and date.today() < TOURNAMENT_START:
        _log("WARN", msg)
    else:
        issues.append(msg)
    return issues


def check_api_key(test_api: bool) -> list[str]:
    errors: list[str] = []
    try:
        get_api_key()
        _log("OK", "API_FOOTBALL_KEY configurada")
    except SystemExit:
        errors.append("API_FOOTBALL_KEY no configurada")
        return errors

    if not test_api:
        return errors

    try:
        fixtures = fetch_all_fixtures()
        _log("OK", f"API-Football responde ({len(fixtures)} fixtures liga 1 / 2026)")
    except Exception as exc:
        errors.append(f"API-Football no responde: {exc}")
    return errors


def check_predictions_sample() -> None:
    """Informa si hay pronósticos rellenados (sanity check, no bloquea)."""
    for group_id in list_groups():
        group = load_group(group_id)
        excel_path = group["excel_path"]
        if not excel_path.exists():
            continue
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        players = group["players"]
        filled = 0
        for name in players:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            for row in range(PT_FIRST_ROW, PT_FIRST_ROW + 104):
                home = ws.cell(row, 4).value
                away = ws.cell(row, 5).value
                if home is not None and away is not None:
                    filled += 1
        wb.close()
        _log("OK", f"Grupo {group_id}: pronósticos en hojas de jugadores (muestras filas: {filled})")


def main() -> None:
    parser = argparse.ArgumentParser(description="Verifica el pipeline de automatización")
    parser.add_argument(
        "--ci",
        action="store_true",
        help="Modo CI: mapa vacío es warning antes del 11/06/2026",
    )
    parser.add_argument(
        "--skip-api",
        action="store_true",
        help="No llama a API-Football (solo comprueba la clave)",
    )
    args = parser.parse_args()

    _log("INFO", "Verificando automatización porra Mundial 2026...")
    errors: list[str] = []
    errors.extend(check_excel_files())
    errors.extend(check_fixture_map(ci_mode=args.ci))
    if not args.skip_api:
        errors.extend(check_api_key(test_api=True))
    else:
        _log("INFO", "Comprobación API omitida (--skip-api)")

    try:
        check_predictions_sample()
    except Exception as exc:
        _log("WARN", f"No se pudo comprobar pronósticos: {exc}")

    if errors:
        for err in errors:
            _log("ERROR", err)
        sys.exit(1)

    _log("INFO", "Todo listo para automatizar.")


if __name__ == "__main__":
    main()
