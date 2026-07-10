"""Escribe pronósticos cuartos manuales (id calendario 97-100)."""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

from config.groups import list_groups, load_group

PT_FIRST_ROW = 4

CUARTOS_LIST_ORDER: list[tuple[int, str, str]] = [
    (97, "Francia", "Marruecos"),
    (98, "Noruega", "Inglaterra"),
    (99, "España", "Bélgica"),
    (100, "Argentina", "Suiza"),
]

# (gh, ga, clasificado) por id calendario
BROSHU_MANUAL: dict[str, dict[int, tuple[int, int, str]]] = {
    "Pepe": {
        97: (2, 1, "Local"),
        99: (1, 0, "Local"),
        98: (2, 2, "Visitante"),
        100: (2, 0, "Local"),
    },
    "Quintero": {
        97: (3, 1, "Local"),
        99: (1, 0, "Local"),
        98: (1, 1, "Visitante"),
        100: (2, 1, "Local"),
    },
    "Nacho": {
        97: (2, 2, "Visitante"),
        99: (2, 0, "Local"),
        98: (0, 2, "Visitante"),
        100: (0, 1, "Visitante"),
    },
    "Fer": {
        97: (0, 0, "Visitante"),
        99: (3, 1, "Local"),
        98: (2, 0, "Local"),
        100: (2, 0, "Local"),
    },
    "Simón": {
        97: (1, 3, "Visitante"),
        99: (3, 0, "Local"),
        98: (2, 2, "Local"),
        100: (0, 0, "Visitante"),
    },
    "Felipe": {
        97: (3, 1, "Local"),
        99: (2, 2, "Local"),
        98: (3, 2, "Local"),
        100: (2, 1, "Local"),
    },
    "Kike": {
        97: (3, 1, "Local"),
        99: (2, 0, "Local"),
        98: (1, 2, "Visitante"),
        100: (1, 1, "Local"),
    },
    "Muni": {
        97: (2, 1, "Local"),
        99: (3, 0, "Local"),
        98: (3, 2, "Local"),
        100: (1, 1, "Local"),
    },
    "Patri": {
        97: (2, 1, "Local"),
        99: (3, 0, "Local"),
        98: (1, 1, "Visitante"),
        100: (1, 1, "Local"),
    },
    "Álvaro": {
        97: (1, 0, "Local"),
        99: (1, 0, "Local"),
        98: (1, 1, "Visitante"),
        100: (1, 0, "Local"),
    },
}

PAPINENES_MANUAL: dict[str, dict[int, tuple[int, int, str]]] = {
    "Álvaro": BROSHU_MANUAL["Álvaro"],
}


def write_predictions(
    excel_path: Path,
    players: dict[str, dict[int, tuple[int, int, str]]],
    dry_run: bool,
) -> None:
    wb = load_workbook(excel_path)
    for player, preds in players.items():
        if player not in wb.sheetnames:
            raise SystemExit(f"Hoja '{player}' no existe en {excel_path.name}")
        ws = wb[player]
        print(f"\n{player}:")
        for mid, (gh, ga, cl) in sorted(preds.items()):
            row = PT_FIRST_ROW + mid - 1
            loc = next((l for m, l, _ in CUARTOS_LIST_ORDER if m == mid), "?")
            vis = next((v for m, _, v in CUARTOS_LIST_ORDER if m == mid), "?")
            print(f"  #{mid} {loc} vs {vis}: {gh}-{ga} clasif={cl}")
            if not dry_run:
                ws.cell(row, 6, value=gh)
                ws.cell(row, 7, value=ga)
                ws.cell(row, 8, value=cl)
    if not dry_run:
        wb.save(excel_path)
    wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Pronósticos cuartos manuales")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--group",
        choices=["broshu", "papinenes", "all"],
        default="all",
    )
    args = parser.parse_args()
    mode = "DRY-RUN" if args.dry_run else "WRITE"
    groups = list_groups() if args.group == "all" else [args.group]

    for group_id in groups:
        excel = load_group(group_id)["excel_path"]
        players = BROSHU_MANUAL if group_id == "broshu" else PAPINENES_MANUAL
        if not players:
            continue
        print(f"\nwrite_cuartos_predictions [{mode}] {group_id} -> {excel.name}")
        write_predictions(excel, players, args.dry_run)

    print("\nTerminado OK.")


if __name__ == "__main__":
    main()
