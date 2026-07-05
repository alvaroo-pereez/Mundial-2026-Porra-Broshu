"""Escribe pronósticos octavos manuales (orden lista → id calendario)."""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

from config.groups import list_groups, load_group
from octavos_fixtures import OCTAVOS_LIST_ORDER

PT_FIRST_ROW = 4

# (gh, ga, clasificado) por id calendario
BROSHU_MANUAL: dict[str, dict[int, tuple[int, int, str]]] = {
    "Fer": {
        90: (1, 0, "Local"),
        89: (0, 0, "Local"),
        91: (0, 2, "Visitante"),
        92: (1, 0, "Local"),
        93: (0, 3, "Visitante"),
        94: (2, 0, "Local"),
        95: (3, 1, "Local"),
        96: (1, 3, "Visitante"),
    },
    "Felipe": {
        90: (1, 3, "Visitante"),
        89: (1, 1, "Visitante"),
        91: (2, 2, "Visitante"),
        92: (2, 2, "Visitante"),
        93: (2, 2, "Visitante"),
        94: (2, 1, "Local"),
        95: (3, 1, "Local"),
        96: (0, 1, "Visitante"),
    },
    "Muni": {
        90: (1, 1, "Local"),
        89: (1, 1, "Visitante"),
        91: (1, 1, "Visitante"),
        92: (1, 1, "Local"),
        93: (0, 3, "Visitante"),
        94: (2, 1, "Local"),
        95: (2, 0, "Local"),
        96: (1, 2, "Visitante"),
    },
    "Álvaro": {
        90: (1, 1, "Visitante"),
        89: (0, 2, "Visitante"),
        91: (1, 1, "Local"),
        92: (1, 1, "Local"),
        93: (1, 1, "Visitante"),
        94: (1, 1, "Local"),
        95: (2, 0, "Local"),
        96: (1, 1, "Visitante"),
    },
    "Simón": {
        90: (1, 3, "Visitante"),
        89: (0, 4, "Visitante"),
        91: (2, 3, "Visitante"),
        92: (0, 2, "Visitante"),
        93: (0, 3, "Visitante"),
        94: (1, 1, "Local"),
        95: (2, 0, "Local"),
        96: (0, 1, "Visitante"),
    },
}

PAPINENES_MANUAL: dict[str, dict[int, tuple[int, int, str]]] = {
    "Álvaro": BROSHU_MANUAL["Álvaro"],
}


def write_predictions(
    excel_path: Path,
    players: dict[str, dict[int, tuple[int, int, str]]],
    dry_run: bool,
) -> None:
    wb = load_workbook(excel_path)
    for player, preds in players.items():
        if player not in wb.sheetnames:
            raise SystemExit(f"Hoja '{player}' no existe en {excel_path.name}")
        ws = wb[player]
        print(f"\n{player}:")
        for mid, (gh, ga, cl) in sorted(preds.items()):
            row = PT_FIRST_ROW + mid - 1
            loc = next((l for m, l, _ in OCTAVOS_LIST_ORDER if m == mid), "?")
            vis = next((v for m, _, v in OCTAVOS_LIST_ORDER if m == mid), "?")
            print(f"  #{mid} {loc} vs {vis}: {gh}-{ga} clasif={cl}")
            if not dry_run:
                ws.cell(row, 6, value=gh)
                ws.cell(row, 7, value=ga)
                ws.cell(row, 8, value=cl)
    if not dry_run:
        wb.save(excel_path)
    wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Pronósticos octavos manuales")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--group",
        choices=["broshu", "papinenes", "all"],
        default="all",
    )
    args = parser.parse_args()
    mode = "DRY-RUN" if args.dry_run else "WRITE"
    groups = list_groups() if args.group == "all" else [args.group]

    for group_id in groups:
        excel = load_group(group_id)["excel_path"]
        players = BROSHU_MANUAL if group_id == "broshu" else PAPINENES_MANUAL
        if not players:
            continue
        print(f"\nwrite_octavos_predictions [{mode}] {group_id} -> {excel.name}")
        write_predictions(excel, players, args.dry_run)

    print("\nTerminado OK.")


if __name__ == "__main__":
    main()
