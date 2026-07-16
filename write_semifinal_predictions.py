"""Escribe pronósticos semifinal/final/tercer puesto (ids 101-104)."""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

from config.groups import list_groups, load_group

PT_FIRST_ROW = 4

MATCH_LABELS: dict[int, tuple[str, str]] = {
    101: ("Francia", "España"),
    102: ("Inglaterra", "Argentina"),
    103: ("Francia", "Inglaterra"),
    104: ("España", "Argentina"),
}

# (gh, ga, clasificado) por id calendario — semifinales
BROSHU_SEMIFINAL: dict[str, dict[int, tuple[int, int, str]]] = {
    "Pepe": {
        101: (2, 1, "Local"),
        102: (1, 2, "Visitante"),
    },
    "Quintero": {
        101: (1, 1, "Visitante"),
        102: (1, 1, "Local"),
    },
    "Muni": {
        101: (2, 3, "Visitante"),
        102: (1, 1, "Visitante"),
    },
    "Simón": {
        101: (2, 3, "Visitante"),
        102: (4, 2, "Local"),
    },
    "Patri": {
        101: (1, 1, "Visitante"),
        102: (1, 1, "Local"),
    },
    "Luis": {
        101: (2, 3, "Visitante"),
        102: (1, 0, "Local"),
    },
    "Nacho": {
        101: (4, 0, "Local"),
        102: (2, 1, "Local"),
    },
    "Fer": {
        101: (0, 7, "Visitante"),
        102: (0, 5, "Visitante"),
    },
    "Kike": {
        101: (2, 3, "Visitante"),
        102: (2, 1, "Local"),
    },
    "Álvaro": {
        101: (1, 1, "Local"),
        102: (1, 1, "Local"),
    },
}

PAPINENES_SEMIFINAL: dict[str, dict[int, tuple[int, int, str]]] = {
    "Álvaro": BROSHU_SEMIFINAL["Álvaro"],
    "Papá": {
        101: (1, 2, "Visitante"),
        102: (2, 1, "Local"),
    },
    "Diego": {
        101: (2, 1, "Local"),
        102: (1, 0, "Visitante"),
    },
}

# Solo Fer y Muni acertaron ambos finalistas
FINAL_TERCER: dict[str, dict[int, tuple[int, int, str]]] = {
    "Fer": {
        104: (10, 9, "Local"),
    },
    "Muni": {
        104: (6, 1, "Local"),
        103: (2, 1, "Local"),
    },
}

# Jugadores sin pronóstico en semifinales (celdas vacías)
CLEAR_SEMIFINAL: dict[str, list[str]] = {
    "broshu": ["Felipe"],
    "papinenes": [],
}

# Jugadores que NO puntúan final/tercer puesto (103-104 vacíos)
CLEAR_FINAL_IDS = (103, 104)


def merge_player_preds(
    semifinal: dict[str, dict[int, tuple[int, int, str]]],
    final_tercer: dict[str, dict[int, tuple[int, int, str]]],
) -> dict[str, dict[int, tuple[int, int, str] | None]]:
    all_players = set(semifinal) | set(final_tercer)
    merged: dict[str, dict[int, tuple[int, int, str] | None]] = {}
    for player in all_players:
        preds: dict[int, tuple[int, int, str] | None] = {}
        preds.update(semifinal.get(player, {}))
        preds.update(final_tercer.get(player, {}))
        merged[player] = preds
    return merged


def write_predictions(
    excel_path: Path,
    semifinal: dict[str, dict[int, tuple[int, int, str]]],
    final_tercer: dict[str, dict[int, tuple[int, int, str]]],
    clear_semifinal: list[str],
    dry_run: bool,
) -> None:
    wb = load_workbook(excel_path)
    all_players = set(wb.sheetnames) & (
        set(semifinal) | set(final_tercer) | set(clear_semifinal)
    )

    for player in sorted(all_players):
        if player not in wb.sheetnames:
            continue
        ws = wb[player]
        print(f"\n{player}:")

        preds = semifinal.get(player, {})
        finals = final_tercer.get(player, {})
        for mid in (101, 102):
            if player in clear_semifinal:
                print(f"  #{mid} (clear semifinal)")
                if not dry_run:
                    row = PT_FIRST_ROW + mid - 1
                    ws.cell(row, 6, value=None)
                    ws.cell(row, 7, value=None)
                    ws.cell(row, 8, value=None)
            elif mid in preds:
                gh, ga, cl = preds[mid]
                loc, vis = MATCH_LABELS[mid]
                print(f"  #{mid} {loc} vs {vis}: {gh}-{ga} clasif={cl}")
                if not dry_run:
                    row = PT_FIRST_ROW + mid - 1
                    ws.cell(row, 6, value=gh)
                    ws.cell(row, 7, value=ga)
                    ws.cell(row, 8, value=cl)

        for mid in CLEAR_FINAL_IDS:
            if mid in finals:
                gh, ga, cl = finals[mid]
                loc, vis = MATCH_LABELS[mid]
                print(f"  #{mid} {loc} vs {vis}: {gh}-{ga} clasif={cl}")
                if not dry_run:
                    row = PT_FIRST_ROW + mid - 1
                    ws.cell(row, 6, value=gh)
                    ws.cell(row, 7, value=ga)
                    ws.cell(row, 8, value=cl)
            else:
                print(f"  #{mid} (clear final/tercer)")
                if not dry_run:
                    row = PT_FIRST_ROW + mid - 1
                    ws.cell(row, 6, value=None)
                    ws.cell(row, 7, value=None)
                    ws.cell(row, 8, value=None)

    if not dry_run:
        wb.save(excel_path)
    wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Pronósticos semifinal/final/tercer puesto")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--group",
        choices=["broshu", "papinenes", "all"],
        default="all",
    )
    args = parser.parse_args()
    mode = "DRY-RUN" if args.dry_run else "WRITE"
    groups = list_groups() if args.group == "all" else [args.group]

    for group_id in groups:
        excel = load_group(group_id)["excel_path"]
        if group_id == "broshu":
            semi = BROSHU_SEMIFINAL
            final = FINAL_TERCER
        else:
            semi = PAPINENES_SEMIFINAL
            final = {}
        clear = CLEAR_SEMIFINAL.get(group_id, [])
        print(f"\nwrite_semifinal_predictions [{mode}] {group_id} -> {excel.name}")
        write_predictions(excel, semi, final, clear, args.dry_run)

    print("\nTerminado OK.")


if __name__ == "__main__":
    main()
